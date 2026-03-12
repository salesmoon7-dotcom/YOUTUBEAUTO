from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def read_excel_row(
    excel_path: str | Path, *, sheet_name: str, row_index: int
) -> dict[str, object]:
    workbook = load_workbook(Path(excel_path), read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        headers: list[str] = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value is not None else "")
        values = [cell.value for cell in sheet[row_index + 2]]
        row_map: dict[str, object] = {}
        for header, value in zip(headers, values, strict=False):
            if header:
                row_map[header] = value
        return row_map
    finally:
        workbook.close()


def read_excel_rows(
    excel_path: str | Path, *, sheet_name: str
) -> list[dict[str, object]]:
    workbook = load_workbook(Path(excel_path), read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        headers: list[str] = []
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value is not None else "")
        rows: list[dict[str, object]] = []
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            row_map: dict[str, object] = {}
            for header, value in zip(headers, values, strict=False):
                if header:
                    row_map[header] = value
            rows.append(row_map)
        return rows
    finally:
        workbook.close()
