from __future__ import annotations

import json
from pathlib import Path
from typing import cast

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from runtime_v2.excel.status_contract import (
    TERMINAL_STATUSES,
    can_transition_excel_status,
)


def _header_map(sheet: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for idx, cell in enumerate(row, start=1):
            typed_cell = cast(Cell, cell)
            if typed_cell.value is not None:
                headers[str(typed_cell.value).strip().lower()] = idx
    return headers


def _assign_cell(sheet: Worksheet, *, row_no: int, column_no: int, value: str) -> bool:
    cell = sheet.cell(row=row_no, column=column_no)
    if isinstance(cell, MergedCell):
        return False
    typed_cell = cast(Cell, cell)
    typed_cell.value = value
    return True


def merge_video_plan_to_excel(
    excel_path: str | Path,
    *,
    sheet_name: str,
    row_index: int,
    next_status: str,
    summary: str,
    reason_code: str,
) -> bool:
    workbook = load_workbook(Path(excel_path))
    try:
        try:
            sheet = cast(Worksheet, workbook[sheet_name])
        except KeyError:
            return False
        headers = _header_map(sheet)
        status_col = headers.get("status")
        summary_col = headers.get("video plan") or headers.get("script")
        reason_col = headers.get("reason code")
        if status_col is None:
            return False
        row_no = row_index + 2
        current_status = str(
            sheet.cell(row=row_no, column=status_col).value or ""
        ).strip()
        if current_status.lower() in TERMINAL_STATUSES:
            return False
        if not can_transition_excel_status(current_status, next_status):
            return False
        if not _assign_cell(
            sheet, row_no=row_no, column_no=status_col, value=next_status
        ):
            return False
        if summary_col is not None:
            if not _assign_cell(
                sheet, row_no=row_no, column_no=summary_col, value=summary
            ):
                return False
        if reason_col is not None:
            if not _assign_cell(
                sheet, row_no=row_no, column_no=reason_col, value=reason_code
            ):
                return False
        try:
            workbook.save(Path(excel_path))
        except OSError:
            return False
        return True
    finally:
        workbook.close()


def finalize_excel_status(
    excel_path: str | Path,
    *,
    sheet_name: str,
    row_index: int,
    next_status: str,
    result_path: str,
    reason_code: str,
) -> bool:
    return merge_video_plan_to_excel(
        excel_path,
        sheet_name=sheet_name,
        row_index=row_index,
        next_status=next_status,
        summary=result_path,
        reason_code=reason_code,
    )


def merge_stage1_handoff_to_excel(
    excel_path: str | Path,
    *,
    sheet_name: str,
    row_index: int,
    parsed_payload: dict[str, object],
) -> bool:
    workbook = load_workbook(Path(excel_path))
    try:
        try:
            sheet = cast(Worksheet, workbook[sheet_name])
        except KeyError:
            return False
        headers = _header_map(sheet)
        row_no = row_index + 2
        _write_optional_field(
            sheet, headers, row_no, "title", str(parsed_payload.get("title", ""))
        )
        _write_optional_field(
            sheet,
            headers,
            row_no,
            "title for thumb",
            str(parsed_payload.get("title_for_thumb", "")),
        )
        _write_optional_field(
            sheet,
            headers,
            row_no,
            "description",
            str(parsed_payload.get("description", "")),
        )
        keywords = parsed_payload.get("keywords", [])
        keyword_text = (
            ", ".join(
                [
                    str(item).strip()
                    for item in cast(list[object], keywords)
                    if str(item).strip()
                ]
            )
            if isinstance(keywords, list)
            else ""
        )
        _write_optional_field(sheet, headers, row_no, "keywords", keyword_text)
        voice_groups = parsed_payload.get("voice_groups", [])
        voice_text = (
            json.dumps(voice_groups, ensure_ascii=False)
            if isinstance(voice_groups, list)
            else ""
        )
        _write_optional_field(sheet, headers, row_no, "voice", voice_text)
        workbook.save(Path(excel_path))
        return True
    except OSError:
        return False
    finally:
        workbook.close()


def _write_optional_field(
    sheet: Worksheet,
    headers: dict[str, int],
    row_no: int,
    header_name: str,
    value: str,
) -> None:
    column_no = headers.get(header_name.strip().lower())
    if column_no is None:
        return
    _ = _assign_cell(sheet, row_no=row_no, column_no=column_no, value=value)


def _parsed_payload_from_video_plan(
    video_plan: dict[str, object],
) -> dict[str, object] | None:
    raw = video_plan.get("stage1_handoff")
    if not isinstance(raw, dict):
        return None
    contract = raw.get("contract")
    if not isinstance(contract, dict):
        return None
    return cast(dict[str, object], contract)
