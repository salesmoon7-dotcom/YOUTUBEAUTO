from __future__ import annotations

import json
import os
import queue
import threading
import tkinter as tk
from pathlib import Path
from time import sleep, time
from tkinter import filedialog, messagebox, ttk
from typing import Callable, cast
from uuid import uuid4

from openpyxl import load_workbook

from runtime_v2.browser.manager import open_browser_for_login
from runtime_v2.bootstrap import ensure_runtime_bootstrap
from runtime_v2.config import GpuWorkload, RuntimeConfig
from runtime_v2.control_plane import (
    run_control_loop_once,
    seed_control_job,
)
from runtime_v2.control_plane_feeder import seed_local_jobs
from runtime_v2.contracts.job_contract import JobContract
from runtime_v2.debug_log import (
    append_debug_event,
    debug_log_path,
    exception_payload,
    summarize_runtime_result,
)
from runtime_v2.cli import _run_excel_batch_mode
from runtime_v2.evidence import load_runtime_readiness
from runtime_v2.gpt.floor import load_gpt_status, write_gpt_status
from runtime_v2.gpu.lease import (
    Lease,
    build_gpu_health_payload,
    write_gpu_health_payload,
)
from runtime_v2.gui_adapter import build_gui_status_payload, write_gui_status
from runtime_v2.manager import seed_excel_row
from runtime_v2.soak_report import write_soak_report


BASE_DIR = Path(__file__).resolve().parent


def _read_json(path: Path) -> dict[str, object] | None:
    if not path.exists():
        return None
    try:
        payload = cast(object, json.loads(path.read_text(encoding="utf-8")))
    except (OSError, json.JSONDecodeError):
        return None
    if isinstance(payload, dict):
        raw_payload = cast(dict[object, object], payload)
        typed: dict[str, object] = {}
        for raw_key in raw_payload:
            typed[str(raw_key)] = raw_payload[raw_key]
        return typed
    return None


def _coerce_mapping(value: object) -> dict[str, object] | None:
    if not isinstance(value, dict):
        return None
    raw_payload = cast(dict[object, object], value)
    typed: dict[str, object] = {}
    for raw_key, raw_value in raw_payload.items():
        typed[str(raw_key)] = raw_value
    return typed


def _worker_error_code_mismatch_warning(
    result_payload: dict[str, object] | None,
) -> str:
    if result_payload is None:
        return ""
    metadata = _coerce_mapping(result_payload.get("metadata"))
    if metadata is None:
        return ""
    canonical_handoff = _coerce_mapping(metadata.get("canonical_handoff"))
    if canonical_handoff is None:
        return ""
    return str(canonical_handoff.get("warning_worker_error_code_mismatch", "")).strip()


def _read_json_list(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    try:
        payload = cast(object, json.loads(path.read_text(encoding="utf-8")))
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(payload, list):
        return []
    raw_payload = cast(list[object], payload)
    items: list[dict[str, object]] = []
    for entry in raw_payload:
        if isinstance(entry, dict):
            raw_entry = cast(dict[object, object], entry)
            typed: dict[str, object] = {}
            for raw_key in raw_entry:
                typed[str(raw_key)] = raw_entry[raw_key]
            items.append(typed)
    return items


def _read_jsonl_tail(path: Path, limit: int = 40) -> list[dict[str, object]]:
    if not path.exists():
        return []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return []
    records: list[dict[str, object]] = []
    for line in lines[-limit:]:
        if not line.strip():
            continue
        try:
            payload = cast(object, json.loads(line))
        except json.JSONDecodeError:
            continue
        typed_payload = _coerce_mapping(payload)
        if typed_payload is not None:
            records.append(typed_payload)
    return records


def _latest_final_output_record(path: Path) -> dict[str, object] | None:
    records = _read_jsonl_tail(path, limit=200)
    for record in reversed(records):
        if str(record.get("event", "")).strip() != "job_summary":
            continue
        if bool(record.get("final_output", False)):
            return record
    return None


def _archived_contract_count(root: Path) -> int:
    if not root.exists():
        return 0
    job_like = list(root.glob("*.job.json"))
    legacy_job_like = [
        path
        for path in root.glob("*.job.*.json")
        if not path.name.endswith(".reason.json")
    ]
    return len(job_like) + len(legacy_job_like)


def _to_int(value: object) -> int:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        try:
            return int(value)
        except ValueError:
            return 0
    return 0


def _to_float(value: object, default: float = 0.0) -> float:
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return default
    return default


def _display_worker_error_code(code: str) -> str:
    normalized = code.strip()
    if normalized == "BROWSER_RESTART_EXHAUSTED":
        return "BROWSER_RESTART_EXHAUSTED(restart budget exhausted)"
    return normalized


def _truncate_ui_text(text: str, limit: int = 64) -> str:
    normalized = " ".join(str(text).split())
    if len(normalized) <= limit:
        return normalized
    return normalized[: limit - 3].rstrip() + "..."


def _friendly_ready_text(readiness: dict[str, object], blockers: list[str]) -> str:
    if bool(readiness.get("ready", False)) and not blockers:
        return "준비되었습니다"
    joined = " ".join(blockers)
    if "GPT_FLOOR" in joined:
        return "GPT 준비 확인이 필요합니다"
    if "login_required" in joined:
        return "브라우저 로그인이 필요합니다"
    if "busy_lock" in joined:
        return "브라우저 작업이 끝날 때까지 기다려주세요"
    if "unknown_lock" in joined:
        return "브라우저 상태 확인이 필요합니다"
    return "실행 전 준비 상태를 확인해주세요"


def _friendly_evidence_text(evidence_summary: str) -> str:
    if "final_output=true" in evidence_summary:
        return "결과 파일이 준비되었습니다"
    if "failure_summary=" in evidence_summary:
        return "실패 요약이 기록되었습니다"
    if "pending terminal evidence" in evidence_summary:
        return "실행 결과를 기다리는 중입니다"
    return "아직 실행 기록이 없습니다"


def _runtime_config_from_text(runtime_root: str) -> RuntimeConfig:
    normalized = runtime_root.strip()
    if not normalized:
        return RuntimeConfig()
    return RuntimeConfig.from_root(Path(normalized).resolve())


def _readiness_blocker_messages(readiness: dict[str, object]) -> list[str]:
    blockers_obj = readiness.get("blockers", [])
    if not isinstance(blockers_obj, list):
        return []
    typed_blockers = cast(list[object], blockers_obj)
    messages: list[str] = []
    for raw_blocker in typed_blockers:
        if not isinstance(raw_blocker, dict):
            continue
        blocker = _coerce_mapping(cast(object, raw_blocker))
        if blocker is None:
            continue
        axis = str(blocker.get("axis", "unknown")).strip()
        code = str(blocker.get("code", "UNKNOWN")).strip()
        reason = str(blocker.get("reason", "")).strip()
        message = f"{axis}:{code}"
        if reason:
            message = f"{message}({reason})"
        messages.append(message)
    return messages


def _format_seed_summary(
    seed_result: dict[str, object] | None,
    *,
    excel_path: str,
    sheet_name: str,
    row_index: int,
) -> str:
    if seed_result is None:
        return "seed: not started"
    status = str(seed_result.get("status", "unknown")).strip()
    if status == "seeded":
        topic_spec = _coerce_mapping(seed_result.get("topic_spec"))
        row_ref = (
            "-"
            if topic_spec is None
            else str(topic_spec.get("row_ref", "-")).strip() or "-"
        )
        job_id = str(seed_result.get("job_id", "-")).strip() or "-"
        return f"seed: seeded row={row_ref} job={job_id}"
    if status == "no_work":
        return f"seed: no work excel={excel_path} sheet={sheet_name} row={row_index}"
    code = str(seed_result.get("code", "UNKNOWN")).strip()
    return f"seed: status={status} code={code}"


def _format_terminal_evidence_summary(
    result_payload: dict[str, object] | None,
    gui_payload: dict[str, object] | None,
) -> str:
    metadata = (
        None
        if result_payload is None
        else _coerce_mapping(result_payload.get("metadata"))
    )
    run_id = ""
    if metadata is not None:
        run_id = str(metadata.get("run_id", "")).strip()
    if not run_id and gui_payload is not None:
        run_id = str(gui_payload.get("run_id", "")).strip()
    final_output = False
    if metadata is not None:
        final_output = bool(metadata.get("final_output", False))
    if not final_output and gui_payload is not None:
        final_output = bool(gui_payload.get("final_output", False))
    final_artifact_path = ""
    if metadata is not None:
        final_artifact_path = str(metadata.get("final_artifact_path", "")).strip()
    if not final_artifact_path and gui_payload is not None:
        final_artifact_path = str(gui_payload.get("final_artifact_path", "")).strip()
    if final_output and final_artifact_path:
        return f"run_id={run_id or '-'} final_output=true path={final_artifact_path}"
    attached_failure_path = ""
    if metadata is not None:
        attached_failure_path = str(metadata.get("failure_summary_path", "")).strip()
    if not attached_failure_path and gui_payload is not None:
        attached_failure_path = str(gui_payload.get("failure_summary_path", "")).strip()
    if attached_failure_path:
        return f"run_id={run_id or '-'} failure_summary={attached_failure_path}"
    if run_id:
        return f"run_id={run_id} pending terminal evidence"
    return "evidence: latest result unavailable"


def _format_batch_summary(batch_result: dict[str, object] | None) -> str:
    if batch_result is None:
        return "batch: not started"
    status = str(batch_result.get("status", "unknown")).strip()
    code = str(batch_result.get("code", "UNKNOWN")).strip()
    selected_rows = (
        cast(list[object], batch_result.get("selected_rows", []))
        if isinstance(batch_result.get("selected_rows", []), list)
        else []
    )
    ticks = _to_int(batch_result.get("ticks", 0))
    if status == "ok":
        return f"batch: ok rows={len(selected_rows)} ticks={ticks}"
    if status == "no_work":
        return "batch: no work"
    return f"batch: status={status} code={code} rows={len(selected_rows)}"


def _format_soak_summary(report_path: str) -> str:
    normalized = report_path.strip()
    if not normalized:
        return "soak: report unavailable"
    return f"soak: report={normalized}"


def _blocking_browser_service_messages(
    browser_registry: dict[str, object] | None,
) -> list[str]:
    if browser_registry is None:
        return []
    sessions_obj = browser_registry.get("sessions", [])
    if not isinstance(sessions_obj, list):
        return []
    blocked: list[str] = []
    for raw_session in cast(list[object], sessions_obj):
        session = _coerce_mapping(raw_session)
        if session is None:
            continue
        status = str(session.get("status", "")).strip().lower()
        if status not in {"login_required", "busy_lock", "unknown_lock"}:
            continue
        service = str(session.get("service", "unknown")).strip().lower() or "unknown"
        blocked.append(f"{service}={status}")
    return blocked


class RuntimeV2ManagerGUI:
    SETTINGS_FILE: Path = (
        BASE_DIR / "system" / "runtime_v2" / "config" / "manager_gui_settings.json"
    )
    LOGIN_SERVICES: tuple[str, ...] = (
        "chatgpt",
        "genspark",
        "seaart",
        "geminigen",
        "canva",
    )

    def __init__(self) -> None:
        self.root: tk.Tk = tk.Tk()
        self.root.title("runtime_v2 매니저")
        self.root.geometry("1120x900")
        _ = self.root.configure(bg="#f4f4f4")
        _ = self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        _ = self.root.resizable(True, True)

        self.style: ttk.Style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass
        self.style.configure("ManagerHeader.TFrame", background="#203040")
        self.style.configure(
            "ManagerHeader.TLabel", background="#203040", foreground="#f4f7fb"
        )
        self.style.configure("AppBody.TFrame", background="#eef2f7")
        self.style.configure(
            "HeroCard.TFrame",
            background="#ffffff",
            relief=tk.GROOVE,
            borderwidth=1,
        )
        self.style.configure(
            "HeroTitle.TLabel",
            background="#ffffff",
            foreground="#1d2a3a",
            font=("Malgun Gothic", 16, "bold"),
        )
        self.style.configure(
            "HeroBody.TLabel",
            background="#ffffff",
            foreground="#52606d",
            font=("Malgun Gothic", 10),
        )
        self.style.configure(
            "Card.TFrame", background="#ffffff", relief=tk.GROOVE, borderwidth=1
        )
        self.style.configure(
            "CardTitle.TLabel",
            background="#ffffff",
            foreground="#223142",
            font=("Malgun Gothic", 10, "bold"),
        )
        self.style.configure(
            "CardBody.TLabel",
            background="#ffffff",
            foreground="#415161",
            font=("Malgun Gothic", 10),
        )
        self.style.configure(
            "ManagerSection.TLabelframe", borderwidth=2, relief=tk.GROOVE
        )
        self.style.configure(
            "ManagerSection.TLabelframe.Label",
            font=("Malgun Gothic", 10, "bold"),
        )
        self.style.configure(
            "Simple.TNotebook",
            background="#eef2f7",
            borderwidth=0,
            tabmargins=[0, 0, 0, 0],
        )
        self.style.configure(
            "Simple.TNotebook.Tab",
            padding=(18, 10),
            font=("Malgun Gothic", 10, "bold"),
        )

        self.config: RuntimeConfig = RuntimeConfig()
        self.owner: str = "runtime_v2_gui"
        self.poll_ms: int = 1000
        self.loop_sleep_sec: float = 2.0

        self.running: bool = False
        self.paused: bool = False
        self.stop_requested: bool = False
        self.stop_event: threading.Event = threading.Event()
        self.pause_event: threading.Event = threading.Event()
        self.control_lock: threading.Lock = threading.Lock()
        self.worker_thread: threading.Thread | None = None
        self.action_thread: threading.Thread | None = None
        self.refresh_after_id: str | None = None

        default_excel_path = BASE_DIR / "4 머니.xlsx"
        self.runtime_root_text: tk.StringVar = tk.StringVar(value="")
        self.excel_path_text: tk.StringVar = tk.StringVar(
            value=str(default_excel_path.resolve())
            if default_excel_path.exists()
            else ""
        )
        self.poll_ms_text: tk.StringVar = tk.StringVar(value="1000")
        self.loop_sleep_text: tk.StringVar = tk.StringVar(value="2.0")
        self.sheet_name_text: tk.StringVar = tk.StringVar(value="Sheet1")
        self.row_index_text: tk.StringVar = tk.StringVar(value="0")
        self.batch_count_text: tk.StringVar = tk.StringVar(value="5")
        self.max_control_ticks_text: tk.StringVar = tk.StringVar(value="50")
        self.login_service_text: tk.StringVar = tk.StringVar(value="chatgpt")
        self.selected_workload: tk.StringVar = tk.StringVar(value="qwen3_tts")
        self.preset_gpt_var: tk.BooleanVar = tk.BooleanVar(value=True)
        self.preset_browser_var: tk.BooleanVar = tk.BooleanVar(value=True)
        self.preset_seed_var: tk.BooleanVar = tk.BooleanVar(value=True)
        self.preset_control_var: tk.BooleanVar = tk.BooleanVar(value=True)
        self.source_path_text: tk.StringVar = tk.StringVar(value="")
        self.audio_path_text: tk.StringVar = tk.StringVar(value="")
        self.script_text: tk.StringVar = tk.StringVar(value="")
        self.duration_text: tk.StringVar = tk.StringVar(value="8")
        self.model_name_text: tk.StringVar = tk.StringVar(value="")
        self.status_text: tk.StringVar = tk.StringVar(value="대기")
        self.warning_text: tk.StringVar = tk.StringVar(value="경고 없음")
        self.last_result_text: tk.StringVar = tk.StringVar(value="최근 실행 없음")
        self.home_warning_text: tk.StringVar = tk.StringVar(value="문제가 없습니다")
        self.home_result_text: tk.StringVar = tk.StringVar(
            value="아직 실행하지 않았습니다"
        )
        self.queue_text: tk.StringVar = tk.StringVar(value="queue: 0")
        self.gpt_text: tk.StringVar = tk.StringVar(value="gpt: unknown")
        self.gpu_text: tk.StringVar = tk.StringVar(value="gpu: unknown")
        self.browser_text: tk.StringVar = tk.StringVar(value="browser: unknown")
        self.artifact_text: tk.StringVar = tk.StringVar(value="artifacts: 0")
        self.readiness_text: tk.StringVar = tk.StringVar(value="readiness: unknown")
        self.browser_services_text: tk.StringVar = tk.StringVar(
            value="services: unknown"
        )
        self.latest_run_text: tk.StringVar = tk.StringVar(value="run_id: -")
        self.seed_result_summary_text: tk.StringVar = tk.StringVar(
            value="seed: not started"
        )
        self.batch_result_summary_text: tk.StringVar = tk.StringVar(
            value="batch: not started"
        )
        self.soak_summary_text: tk.StringVar = tk.StringVar(
            value="soak: report unavailable"
        )
        self.evidence_summary_text: tk.StringVar = tk.StringVar(
            value="evidence: latest result unavailable"
        )
        self.home_ready_text: tk.StringVar = tk.StringVar(
            value="준비 상태를 확인하는 중입니다"
        )
        self.home_recent_text: tk.StringVar = tk.StringVar(
            value="최근 실행 정보가 없습니다"
        )
        self.home_browser_text: tk.StringVar = tk.StringVar(
            value="브라우저 상태를 확인하는 중입니다"
        )
        self.browser_progress_value: tk.IntVar = tk.IntVar(value=0)
        self.queue_progress_value: tk.IntVar = tk.IntVar(value=0)
        self.gpt_progress_value: tk.IntVar = tk.IntVar(value=0)
        self.artifact_progress_value: tk.IntVar = tk.IntVar(value=0)
        self.log_filter_text: tk.StringVar = tk.StringVar(value="")
        self.log_level_filter_text: tk.StringVar = tk.StringVar(value="전체")
        self.log_autoscroll_var: tk.BooleanVar = tk.BooleanVar(value=True)
        self.log_recent_only_var: tk.BooleanVar = tk.BooleanVar(value=False)
        self.log_buffer: list[str] = []
        self.last_event_ts: float = 0.0
        self.ui_queue: queue.SimpleQueue[tuple[str, str]] = queue.SimpleQueue()
        self._json_cache: dict[tuple[str, int], tuple[float, int, object]] = {}
        self._sheet_cache: dict[str, tuple[float, int, list[str]]] = {}

        self.program_rows: dict[str, dict[str, object]] = {}
        self.home_action_buttons: dict[str, tk.Button] = {}
        self.queue_list: tk.Listbox | None = None
        self.sheet_listbox: tk.Listbox | None = None
        self.log_text: tk.Text | None = None

        self._load_settings()
        self._refresh_runtime_config()
        self._build_ui()
        self._apply_home_presets()
        self._refresh_sheet_list()
        self._ensure_snapshot_contracts()
        self.refresh_after_id = self.root.after(200, self.refresh_dashboard)

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding=10, style="AppBody.TFrame")
        main.pack(fill=tk.BOTH, expand=True)

        self._build_header_frame(main)

        notebook = ttk.Notebook(main, style="Simple.TNotebook")
        notebook.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        home = ttk.Frame(notebook, padding=14, style="AppBody.TFrame")
        progress = ttk.Frame(notebook, padding=14, style="AppBody.TFrame")
        logs = ttk.Frame(notebook, padding=14, style="AppBody.TFrame")
        notebook.add(home, text="홈")
        notebook.add(progress, text="진행")
        notebook.add(logs, text="로그")

        self._build_home_tab(home)
        self._build_progress_tab(progress)
        self._build_logs_tab(logs)

    def _build_home_tab(self, parent: ttk.Frame) -> None:
        hero = ttk.Frame(parent, padding=14, style="HeroCard.TFrame")
        hero.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(hero, text="제어 홈", style="HeroTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hero,
            text="레거시 매니저처럼 설정을 먼저 맞추고, 홈 탭에서 바로 실행을 시작하세요.",
            style="HeroBody.TLabel",
        ).pack(anchor="w", pady=(6, 0))

        card_row = ttk.Frame(parent, style="AppBody.TFrame")
        card_row.pack(fill=tk.X, pady=(0, 8))
        for column in range(3):
            _ = card_row.grid_columnconfigure(column, weight=1)
        self._build_status_card(card_row, 0, "현재 상태", self.status_text)
        self._build_status_card(card_row, 1, "주의/안내", self.home_warning_text)
        self._build_status_card(card_row, 2, "최근 결과", self.home_result_text)

        body = ttk.Frame(parent, style="AppBody.TFrame")
        body.pack(fill=tk.BOTH, expand=True)
        _ = body.grid_columnconfigure(0, weight=5, minsize=700)
        _ = body.grid_columnconfigure(1, weight=3, minsize=320)
        _ = body.grid_rowconfigure(0, weight=1)

        control = ttk.Frame(body, style="AppBody.TFrame")
        _ = control.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        self._build_control_frame(control)

        summary = ttk.LabelFrame(
            body, text="홈 요약", padding=10, style="ManagerSection.TLabelframe"
        )
        _ = summary.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        self._build_home_summary_frame(summary)

    def _build_progress_tab(self, parent: ttk.Frame) -> None:
        intro = ttk.Frame(parent, padding=14, style="HeroCard.TFrame")
        intro.pack(fill=tk.X, pady=(0, 12))
        ttk.Label(intro, text="실행 진행 현황", style="HeroTitle.TLabel").pack(
            anchor="w"
        )
        ttk.Label(
            intro,
            text="현재 준비 상태, 서브시스템, 큐 처리 흐름을 한곳에서 확인할 수 있습니다.",
            style="HeroBody.TLabel",
        ).pack(anchor="w", pady=(6, 0))

        card_row = ttk.Frame(parent, style="AppBody.TFrame")
        card_row.pack(fill=tk.X, pady=(0, 12))
        for column in range(3):
            _ = card_row.grid_columnconfigure(column, weight=1)
        self._build_status_card(card_row, 0, "실행 상태", self.status_text)
        self._build_status_card(card_row, 1, "주의/안내", self.warning_text)
        self._build_status_card(card_row, 2, "실행 증거", self.evidence_summary_text)

        content = ttk.Frame(parent, style="AppBody.TFrame")
        content.pack(fill=tk.BOTH, expand=True)

        split: ttk.Panedwindow = ttk.Panedwindow(content, orient=tk.HORIZONTAL)
        split.pack(fill=tk.BOTH, expand=True)
        left = ttk.Frame(split, padding=(0, 0, 8, 0), style="AppBody.TFrame")
        right = ttk.Frame(split, padding=(8, 0, 0, 0), style="AppBody.TFrame")
        _ = split.add(left, weight=3)
        _ = split.add(right, weight=4)

        self._build_overview_frame(left)
        self._build_progress_meter_frame(left)
        self._build_programs_frame(left)
        self._build_queue_frame(right)

    def _build_logs_tab(self, parent: ttk.Frame) -> None:
        intro = ttk.Frame(parent, padding=14, style="HeroCard.TFrame")
        intro.pack(fill=tk.X, pady=(0, 12))
        ttk.Label(intro, text="실행 로그", style="HeroTitle.TLabel").pack(anchor="w")
        ttk.Label(
            intro,
            text="최근 이벤트와 작업 요약 로그를 확인하는 전용 화면입니다.",
            style="HeroBody.TLabel",
        ).pack(anchor="w", pady=(6, 0))

        status_strip = ttk.Frame(parent, style="AppBody.TFrame")
        status_strip.pack(fill=tk.X, pady=(0, 12))
        _ = status_strip.grid_columnconfigure(0, weight=1)
        _ = status_strip.grid_columnconfigure(1, weight=1)
        self._build_status_card(status_strip, 0, "최근 실행", self.last_result_text)
        self._build_status_card(status_strip, 1, "최근 Run ID", self.latest_run_text)

        filter_row = ttk.Frame(parent, style="AppBody.TFrame")
        filter_row.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(filter_row, text="검색:").pack(side=tk.LEFT)
        filter_entry = ttk.Entry(filter_row, textvariable=self.log_filter_text)
        filter_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 8))
        _ = filter_entry.bind("<KeyRelease>", self._on_log_filter_changed)
        ttk.Label(filter_row, text="레벨:").pack(side=tk.LEFT)
        level_combo = ttk.Combobox(
            filter_row,
            textvariable=self.log_level_filter_text,
            state="readonly",
            width=10,
        )
        level_combo["values"] = ("전체", "INFO", "WARNING", "ERROR")
        level_combo.pack(side=tk.LEFT)
        _ = level_combo.bind("<<ComboboxSelected>>", self._on_log_filter_changed)
        ttk.Checkbutton(
            filter_row,
            text="자동 스크롤",
            variable=self.log_autoscroll_var,
            command=self._on_log_filter_changed,
        ).pack(side=tk.LEFT, padx=(8, 4))
        ttk.Checkbutton(
            filter_row,
            text="최근 100개",
            variable=self.log_recent_only_var,
            command=self._on_log_filter_changed,
        ).pack(side=tk.LEFT)

        body = ttk.Frame(parent, style="AppBody.TFrame")
        body.pack(fill=tk.BOTH, expand=True)
        tool_row = ttk.Frame(parent, style="AppBody.TFrame")
        tool_row.pack(fill=tk.X, pady=(0, 10))
        tk.Button(
            tool_row,
            text="로그 지우기",
            width=11,
            command=self.clear_logs,
            bg="#f8fafc",
            fg="#1f2937",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side=tk.LEFT)
        tk.Button(
            tool_row,
            text="이벤트 파일 열기",
            width=15,
            command=self.open_events_file,
            bg="#eff6ff",
            fg="#1d4ed8",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side=tk.LEFT, padx=4)
        tk.Button(
            tool_row,
            text="디버그 폴더 열기",
            width=15,
            command=self.open_debug_dir,
            bg="#ecfeff",
            fg="#0f766e",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side=tk.LEFT, padx=4)
        tk.Button(
            tool_row,
            text="로그 복사",
            width=11,
            command=self.copy_visible_logs,
            bg="#f8fafc",
            fg="#1f2937",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side=tk.LEFT, padx=4)
        tk.Button(
            tool_row,
            text="로그 저장",
            width=11,
            command=self.save_visible_logs,
            bg="#eff6ff",
            fg="#1d4ed8",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side=tk.LEFT, padx=4)
        tk.Button(
            tool_row,
            text="에러만 추출",
            width=12,
            command=self.extract_error_logs,
            bg="#fee2e2",
            fg="#b91c1c",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side=tk.LEFT, padx=4)
        self._build_logs_frame(body)

    def _build_status_card(
        self, parent: ttk.Frame, column: int, title: str, text_var: tk.StringVar
    ) -> None:
        card = ttk.Frame(parent, padding=10, style="Card.TFrame")
        _ = card.grid(row=0, column=column, sticky="nsew", padx=4)
        ttk.Label(card, text=title, style="CardTitle.TLabel").pack(anchor="w")
        ttk.Label(
            card,
            textvariable=text_var,
            style="CardBody.TLabel",
            wraplength=180,
            justify=tk.LEFT,
        ).pack(anchor="w", pady=(8, 0))

    def _build_progress_meter_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(
            parent, text="진행 미터", padding=8, style="ManagerSection.TLabelframe"
        )
        frame.pack(fill=tk.X, pady=(0, 8))
        meters = (
            ("브라우저 준비", self.browser_progress_value, self.browser_text),
            ("GPT 준비", self.gpt_progress_value, self.gpt_text),
            ("큐 적재", self.queue_progress_value, self.queue_text),
            ("결과 산출", self.artifact_progress_value, self.artifact_text),
        )
        for title, value_var, text_var in meters:
            row = ttk.Frame(frame)
            row.pack(fill=tk.X, pady=3)
            ttk.Label(row, text=title, width=10).pack(side=tk.LEFT)
            ttk.Progressbar(
                row,
                mode="determinate",
                variable=value_var,
                maximum=100,
            ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 8))
            ttk.Label(row, textvariable=text_var, width=26).pack(side=tk.LEFT)

    def _build_quick_start_frame(self, parent: ttk.LabelFrame) -> None:
        ttk.Label(parent, text="Excel 파일", style="CardTitle.TLabel").pack(anchor="w")
        ttk.Entry(parent, textvariable=self.excel_path_text).pack(
            fill=tk.X, pady=(6, 12)
        )

        row = ttk.Frame(parent)
        row.pack(fill=tk.X, pady=(0, 12))
        ttk.Label(row, text="시트", style="CardTitle.TLabel").pack(anchor="w")
        sheet_box = ttk.Frame(parent)
        sheet_box.pack(fill=tk.X, pady=(6, 12))
        self.sheet_listbox = tk.Listbox(sheet_box, height=4, exportselection=False)
        self.sheet_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True)
        _ = self.sheet_listbox.bind("<<ListboxSelect>>", self._on_sheet_select)
        sheet_scroll = ttk.Scrollbar(
            sheet_box,
            orient=tk.VERTICAL,
            command=self._sheet_listbox_yview,
        )
        sheet_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        _ = self.sheet_listbox.configure(
            yscrollcommand=cast(Callable[..., object], sheet_scroll.set)
        )

        row2 = ttk.Frame(parent)
        row2.pack(fill=tk.X, pady=(0, 12))
        ttk.Label(row2, text="행", style="CardTitle.TLabel").pack(side=tk.LEFT)
        ttk.Entry(row2, textvariable=self.row_index_text, width=8).pack(
            side=tk.LEFT, padx=(8, 0)
        )

        browser = ttk.Frame(parent)
        browser.pack(fill=tk.X, pady=(0, 16))
        ttk.Label(browser, text="로그인 브라우저", style="CardTitle.TLabel").pack(
            side=tk.LEFT
        )
        combo = ttk.Combobox(
            browser,
            textvariable=self.login_service_text,
            state="readonly",
            width=14,
        )
        combo["values"] = self.LOGIN_SERVICES
        combo.pack(side=tk.LEFT, padx=(8, 0))

        action_col = ttk.Frame(parent)
        action_col.pack(fill=tk.X)
        tk.Button(
            action_col,
            text="1. 로그인 브라우저 열기",
            command=self.trigger_open_login_browser,
            bg="#ffffff",
            fg="#203040",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 11, "bold"),
            padx=14,
            pady=12,
        ).pack(fill=tk.X, pady=(0, 10))
        tk.Button(
            action_col,
            text="2. 입력한 행 불러오기",
            command=self.trigger_excel_seed,
            bg="#1e88e5",
            fg="#ffffff",
            relief=tk.FLAT,
            font=("Malgun Gothic", 11, "bold"),
            padx=14,
            pady=12,
        ).pack(fill=tk.X, pady=(0, 10))
        tk.Button(
            action_col,
            text="3. 한 단계 실행",
            command=self.trigger_control_once,
            bg="#00a86b",
            fg="#ffffff",
            relief=tk.FLAT,
            font=("Malgun Gothic", 12, "bold"),
            padx=14,
            pady=14,
        ).pack(fill=tk.X)

        helper = ttk.Frame(parent)
        helper.pack(fill=tk.X, pady=(16, 0))
        tk.Button(
            helper, text="새로고침", width=10, command=self.refresh_dashboard_now
        ).pack(side=tk.LEFT)

    def _build_home_summary_frame(self, parent: ttk.LabelFrame) -> None:
        items = (
            ("준비 상태", self.home_ready_text),
            ("최근 실행", self.home_recent_text),
            ("브라우저 상태", self.home_browser_text),
        )
        for title, var in items:
            card = ttk.Frame(parent, padding=8, style="Card.TFrame")
            card.pack(fill=tk.X, pady=(0, 8))
            ttk.Label(card, text=title, style="CardTitle.TLabel").pack(anchor="w")
            ttk.Label(
                card,
                textvariable=var,
                style="CardBody.TLabel",
                wraplength=240,
                justify=tk.LEFT,
            ).pack(anchor="w", pady=(6, 0))

    def _sheet_names_for_excel(self) -> list[str]:
        excel_path = self.excel_path_text.get().strip()
        if not excel_path:
            return []
        path = Path(excel_path)
        if not path.exists():
            return []
        stamp = self._file_stamp(path)
        cache_key = str(path.resolve())
        cached = self._sheet_cache.get(cache_key)
        if stamp is not None and cached is not None and cached[:2] == stamp:
            return list(cached[2])
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            names = [str(name) for name in workbook.sheetnames]
        finally:
            workbook.close()
        if stamp is not None:
            self._sheet_cache[cache_key] = (stamp[0], stamp[1], list(names))
        return names

    def _refresh_sheet_list(self) -> None:
        if self.sheet_listbox is None:
            return
        current = self.sheet_name_text.get().strip()
        self.sheet_listbox.delete(0, tk.END)
        names = self._sheet_names_for_excel()
        if not names:
            self.sheet_listbox.insert(tk.END, "시트를 찾을 수 없습니다")
            return
        selected_index = 0
        for index, name in enumerate(names):
            self.sheet_listbox.insert(tk.END, name)
            if name == current:
                selected_index = index
        self.sheet_listbox.selection_set(selected_index)
        self.sheet_listbox.activate(selected_index)
        self.sheet_name_text.set(names[selected_index])

    def _on_sheet_select(self, _event: tk.Event[tk.Listbox]) -> None:
        if self.sheet_listbox is None:
            return
        selected = cast(tuple[int, ...], self.sheet_listbox.curselection())
        if not selected:
            return
        value = cast(str, self.sheet_listbox.get(selected[0]))
        if value != "시트를 찾을 수 없습니다":
            self.sheet_name_text.set(value)

    def _sheet_listbox_yview(self, *args: str) -> tuple[float, float] | None:
        if self.sheet_listbox is None:
            return None
        return cast(tuple[float, float] | None, self.sheet_listbox.yview(*args))

    def _queue_list_yview(self, *args: str) -> tuple[float, float] | None:
        if self.queue_list is None:
            return None
        return cast(tuple[float, float] | None, self.queue_list.yview(*args))

    def _log_text_yview(self, *args: str) -> tuple[float, float] | None:
        if self.log_text is None:
            return None
        return cast(tuple[float, float] | None, self.log_text.yview(*args))

    def _build_header_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.Frame(parent, style="ManagerHeader.TFrame", padding=12)
        frame.pack(fill=tk.X)

        title_row = ttk.Frame(frame, style="ManagerHeader.TFrame")
        title_row.pack(fill=tk.X)
        ttk.Label(
            title_row,
            text="runtime_v2 매니저",
            style="ManagerHeader.TLabel",
            font=("Malgun Gothic", 16, "bold"),
        ).pack(side=tk.LEFT, anchor="w")
        ttk.Label(
            title_row,
            text="레거시 제어 흐름을 유지하면서 홈 / 진행 / 로그 3탭으로 분리했습니다.",
            style="ManagerHeader.TLabel",
            font=("Malgun Gothic", 10),
        ).pack(side=tk.RIGHT, anchor="e")

    def _build_control_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(
            parent,
            text="제어",
            padding=6,
            style="ManagerSection.TLabelframe",
        )
        frame.pack(fill=tk.X, pady=(0, 8))
        mode_row = ttk.Frame(frame)
        mode_row.pack(fill=tk.X, pady=(0, 4))
        for column in range(4):
            _ = mode_row.grid_columnconfigure(column, weight=1)
        ttk.Label(mode_row, text="작업군", style="CardTitle.TLabel").grid(
            row=0, column=0, sticky="w", padx=(2, 8)
        )
        workload_combo = ttk.Combobox(
            mode_row,
            textvariable=self.selected_workload,
            state="readonly",
            width=14,
        )
        workload_combo["values"] = ("qwen3_tts", "rvc", "kenburns")
        workload_combo.grid(row=0, column=1, sticky="ew", padx=(0, 8))
        ttk.Label(mode_row, text="로그인", style="CardTitle.TLabel").grid(
            row=0, column=2, sticky="w", padx=(2, 8)
        )
        login_combo = ttk.Combobox(
            mode_row,
            textvariable=self.login_service_text,
            state="readonly",
            width=14,
        )
        login_combo["values"] = self.LOGIN_SERVICES
        login_combo.grid(row=0, column=3, sticky="ew")

        settings_row = ttk.Frame(frame)
        settings_row.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(settings_row, text="갱신(ms):").pack(side=tk.LEFT)
        ttk.Entry(settings_row, textvariable=self.poll_ms_text, width=8).pack(
            side=tk.LEFT, padx=(4, 8)
        )
        ttk.Label(settings_row, text="루프 대기(s):").pack(side=tk.LEFT)
        ttk.Entry(settings_row, textvariable=self.loop_sleep_text, width=8).pack(
            side=tk.LEFT, padx=(4, 8)
        )
        ttk.Label(settings_row, text="최근 실행", style="CardTitle.TLabel").pack(
            side=tk.LEFT, padx=(12, 4)
        )
        ttk.Label(settings_row, textvariable=self.latest_run_text).pack(side=tk.LEFT)

        runtime_row = ttk.Frame(frame)
        runtime_row.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(runtime_row, text="런타임 루트:").pack(side=tk.LEFT, padx=(0, 4))
        ttk.Entry(runtime_row, textvariable=self.runtime_root_text).pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )
        tk.Button(
            runtime_row,
            text="상태 새로고침",
            width=14,
            command=self.refresh_dashboard_now,
        ).pack(side=tk.LEFT, padx=(8, 0))

        excel_row = ttk.Frame(frame)
        excel_row.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(excel_row, text="엑셀 파일:").pack(side=tk.LEFT, padx=(0, 4))
        ttk.Entry(excel_row, textvariable=self.excel_path_text).pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )

        row_select = ttk.Frame(frame)
        row_select.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(row_select, text="시트:").pack(side=tk.LEFT, padx=(0, 4))
        ttk.Entry(row_select, textvariable=self.sheet_name_text, width=16).pack(
            side=tk.LEFT
        )
        ttk.Label(row_select, text="행 번호(0부터):").pack(side=tk.LEFT, padx=(8, 4))
        ttk.Entry(row_select, textvariable=self.row_index_text, width=8).pack(
            side=tk.LEFT
        )
        ttk.Label(row_select, text="배치 수:").pack(side=tk.LEFT, padx=(8, 4))
        ttk.Entry(row_select, textvariable=self.batch_count_text, width=6).pack(
            side=tk.LEFT
        )
        ttk.Label(row_select, text="최대 Tick:").pack(side=tk.LEFT, padx=(8, 4))
        ttk.Entry(row_select, textvariable=self.max_control_ticks_text, width=6).pack(
            side=tk.LEFT
        )
        ttk.Label(row_select, text="로그인 브라우저:").pack(side=tk.LEFT, padx=(8, 4))
        browser_combo = ttk.Combobox(
            row_select, textvariable=self.login_service_text, state="readonly", width=12
        )
        browser_combo["values"] = self.LOGIN_SERVICES
        browser_combo.pack(side=tk.LEFT)

        preset_frame = ttk.LabelFrame(
            frame,
            text="빠른 프리셋",
            padding=6,
            style="ManagerSection.TLabelframe",
        )
        preset_frame.pack(fill=tk.X, pady=(0, 4))
        preset_row = ttk.Frame(preset_frame)
        preset_row.pack(fill=tk.X)
        ttk.Checkbutton(
            preset_row,
            text="GPT 준비",
            variable=self.preset_gpt_var,
            command=self._apply_home_presets,
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Checkbutton(
            preset_row,
            text="브라우저",
            variable=self.preset_browser_var,
            command=self._apply_home_presets,
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Checkbutton(
            preset_row,
            text="행 적재",
            variable=self.preset_seed_var,
            command=self._apply_home_presets,
        ).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Checkbutton(
            preset_row,
            text="제어 실행",
            variable=self.preset_control_var,
            command=self._apply_home_presets,
        ).pack(side=tk.LEFT)
        preset_button_row = ttk.Frame(preset_frame)
        preset_button_row.pack(fill=tk.X, pady=(6, 0))
        tk.Button(
            preset_button_row,
            text="전체 선택",
            width=10,
            command=lambda: self._apply_named_preset("all"),
            bg="#ecfeff",
            fg="#0f766e",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side=tk.LEFT)
        tk.Button(
            preset_button_row,
            text="최소 실행",
            width=10,
            command=lambda: self._apply_named_preset("minimal"),
            bg="#eff6ff",
            fg="#1d4ed8",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side=tk.LEFT, padx=4)
        tk.Button(
            preset_button_row,
            text="복구 모드",
            width=10,
            command=lambda: self._apply_named_preset("recovery"),
            bg="#fef3c7",
            fg="#92400e",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        ).pack(side=tk.LEFT)

        action_row = ttk.Frame(frame)
        action_row.pack(fill=tk.X)
        refresh_button = tk.Button(
            action_row,
            text="새로고침",
            width=10,
            command=self.refresh_dashboard_now,
            bg="#f8fafc",
            fg="#1f2937",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        )
        refresh_button.pack(side=tk.LEFT)
        self.home_action_buttons["refresh"] = refresh_button
        inbox_button = tk.Button(
            action_row,
            text="받은 작업 확인",
            width=12,
            command=self.trigger_scan_inbox,
            bg="#eff6ff",
            fg="#1d4ed8",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        )
        inbox_button.pack(side=tk.LEFT, padx=4)
        self.home_action_buttons["scan_inbox"] = inbox_button
        gpt_button = tk.Button(
            action_row,
            text="GPT 준비 실행",
            width=12,
            command=self.trigger_gpt_spawn,
            bg="#ecfeff",
            fg="#0f766e",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        )
        gpt_button.pack(side=tk.LEFT, padx=4)
        self.home_action_buttons["gpt_spawn"] = gpt_button
        browser_button = tk.Button(
            action_row,
            text="로그인 브라우저",
            width=12,
            command=self.trigger_open_login_browser,
            bg="#ffffff",
            fg="#334155",
            relief=tk.GROOVE,
            font=("Malgun Gothic", 9, "bold"),
        )
        browser_button.pack(side=tk.LEFT, padx=4)
        self.home_action_buttons["open_browser"] = browser_button

        action_row2 = ttk.Frame(frame)
        action_row2.pack(fill=tk.X, pady=(4, 0))
        seed_button = tk.Button(
            action_row2,
            text="행 불러오기",
            width=12,
            command=self.trigger_excel_seed,
            bg="#dbeafe",
            fg="#1d4ed8",
            relief=tk.FLAT,
            font=("Malgun Gothic", 10, "bold"),
        )
        seed_button.pack(side=tk.LEFT)
        self.home_action_buttons["seed"] = seed_button
        batch_button = tk.Button(
            action_row2,
            text="5행 배치",
            width=11,
            command=self.trigger_excel_batch,
            bg="#ede9fe",
            fg="#5b21b6",
            relief=tk.FLAT,
            font=("Malgun Gothic", 10, "bold"),
        )
        batch_button.pack(side=tk.LEFT, padx=4)
        self.home_action_buttons["batch"] = batch_button
        control_button = tk.Button(
            action_row2,
            text="한 단계 실행",
            width=11,
            command=self.trigger_control_once,
            bg="#dcfce7",
            fg="#166534",
            relief=tk.FLAT,
            font=("Malgun Gothic", 10, "bold"),
        )
        control_button.pack(side=tk.LEFT, padx=4)
        self.home_action_buttons["control"] = control_button
        recovery_button = tk.Button(
            action_row2,
            text="Recovery",
            width=11,
            command=self.trigger_recovery,
            bg="#fef3c7",
            fg="#92400e",
            relief=tk.FLAT,
            font=("Malgun Gothic", 10, "bold"),
        )
        recovery_button.pack(side=tk.LEFT, padx=4)
        self.home_action_buttons["recovery"] = recovery_button
        soak_button = tk.Button(
            action_row2,
            text="Soak Report",
            width=11,
            command=self.trigger_soak_report,
            bg="#fef3c7",
            fg="#92400e",
            relief=tk.FLAT,
            font=("Malgun Gothic", 10, "bold"),
        )
        soak_button.pack(side=tk.LEFT, padx=4)
        self.home_action_buttons["soak"] = soak_button
        start_button = tk.Button(
            action_row2,
            text="Start Loop",
            width=11,
            command=self.start_loop,
            bg="#00a86b",
            fg="#ffffff",
            relief=tk.FLAT,
            font=("Malgun Gothic", 10, "bold"),
        )
        start_button.pack(side=tk.LEFT, padx=4)
        self.home_action_buttons["start_loop"] = start_button

        advanced = ttk.LabelFrame(
            frame,
            text="레거시 수동 제어",
            padding=6,
            style="ManagerSection.TLabelframe",
        )
        advanced.pack(fill=tk.X, pady=(6, 0))

        top = ttk.Frame(advanced)
        top.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(top, text="수동 작업군:").pack(side=tk.LEFT)
        combo = ttk.Combobox(
            top, textvariable=self.selected_workload, state="readonly", width=16
        )
        combo["values"] = ("qwen3_tts", "rvc", "kenburns")
        combo.pack(side=tk.LEFT, padx=(6, 0))

        row = ttk.Frame(advanced)
        row.pack(fill=tk.X, pady=(0, 6))
        tk.Button(
            row, text="▶ Start", width=10, command=self.start_loop, fg="#1f7a1f"
        ).pack(side=tk.LEFT)
        tk.Button(
            row, text="⏸ Pause", width=10, command=self.toggle_pause, fg="#b36b00"
        ).pack(side=tk.LEFT, padx=4)
        tk.Button(
            row, text="■ Stop", width=10, command=self.stop_loop, fg="#b32020"
        ).pack(side=tk.LEFT)
        tk.Button(row, text="↻ Recovery", width=12, command=self.trigger_recovery).pack(
            side=tk.LEFT, padx=(16, 4)
        )

        seed = ttk.Frame(advanced)
        seed.pack(fill=tk.X)
        ttk.Label(seed, text="path:").pack(side=tk.LEFT, padx=(0, 4))
        ttk.Entry(seed, textvariable=self.source_path_text, width=26).pack(side=tk.LEFT)
        ttk.Label(seed, text="audio:").pack(side=tk.LEFT, padx=(8, 4))
        ttk.Entry(seed, textvariable=self.audio_path_text, width=24).pack(side=tk.LEFT)

        tune = ttk.Frame(advanced)
        tune.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(tune, text="duration:").pack(side=tk.LEFT, padx=(0, 4))
        ttk.Entry(tune, textvariable=self.duration_text, width=6).pack(side=tk.LEFT)
        ttk.Label(tune, text="model:").pack(side=tk.LEFT, padx=(8, 4))
        ttk.Entry(tune, textvariable=self.model_name_text, width=14).pack(side=tk.LEFT)

        prompt = ttk.Frame(advanced)
        prompt.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(prompt, text="script:").pack(side=tk.LEFT, padx=(0, 4))
        ttk.Entry(prompt, textvariable=self.script_text, width=64).pack(
            side=tk.LEFT, fill=tk.X, expand=True
        )

        seed_buttons = ttk.Frame(advanced)
        seed_buttons.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(seed_buttons, text="큐 삽입:").pack(side=tk.LEFT)
        tk.Button(
            seed_buttons,
            text="QWEN3",
            width=10,
            command=lambda: self.seed_job("qwen3_tts"),
        ).pack(side=tk.LEFT, padx=(6, 4))
        tk.Button(
            seed_buttons, text="RVC", width=10, command=lambda: self.seed_job("rvc")
        ).pack(side=tk.LEFT, padx=4)
        tk.Button(
            seed_buttons,
            text="KenBurns",
            width=10,
            command=lambda: self.seed_job("kenburns"),
        ).pack(side=tk.LEFT, padx=4)

    def _build_overview_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(
            parent, text="Master Status", padding=8, style="ManagerSection.TLabelframe"
        )
        frame.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(
            frame, textvariable=self.status_text, font=("Malgun Gothic", 10, "bold")
        ).pack(anchor="w")
        ttk.Label(frame, textvariable=self.warning_text, foreground="#b32020").pack(
            anchor="w", pady=(4, 0)
        )
        ttk.Label(frame, textvariable=self.last_result_text).pack(
            anchor="w", pady=(4, 0)
        )

        operator = ttk.Frame(frame)
        operator.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(operator, textvariable=self.readiness_text).pack(anchor="w")
        ttk.Label(operator, textvariable=self.browser_services_text).pack(
            anchor="w", pady=(2, 0)
        )
        ttk.Label(operator, textvariable=self.latest_run_text).pack(
            anchor="w", pady=(2, 0)
        )
        ttk.Label(operator, textvariable=self.seed_result_summary_text).pack(
            anchor="w", pady=(2, 0)
        )
        ttk.Label(operator, textvariable=self.batch_result_summary_text).pack(
            anchor="w", pady=(2, 0)
        )
        ttk.Label(operator, textvariable=self.soak_summary_text).pack(
            anchor="w", pady=(2, 0)
        )
        ttk.Label(operator, textvariable=self.evidence_summary_text).pack(
            anchor="w", pady=(2, 0)
        )

        grid = ttk.Frame(frame)
        grid.pack(fill=tk.X, pady=(6, 0))
        for index, var in enumerate(
            (
                self.browser_text,
                self.gpu_text,
                self.gpt_text,
                self.queue_text,
                self.artifact_text,
            )
        ):
            label = ttk.Label(grid, textvariable=var, relief=tk.GROOVE, padding=6)
            _ = label.grid(
                row=index // 2, column=index % 2, sticky="ew", padx=2, pady=2
            )
        _ = grid.grid_columnconfigure(0, weight=1)
        _ = grid.grid_columnconfigure(1, weight=1)

    def _build_programs_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(
            parent,
            text="Subsystems",
            padding=8,
            style="ManagerSection.TLabelframe",
        )
        frame.pack(fill=tk.BOTH, expand=False, pady=(0, 8))
        for program in (
            "GPT",
            "Genspark",
            "SeaArt",
            "Browser",
            "GeminiGen",
            "KenBurns",
            "QWEN3_TTS",
            "RVC",
            "Result",
            "GUI",
        ):
            card = ttk.Frame(frame, padding=6, style="Card.TFrame")
            card.pack(fill=tk.X, pady=3)
            top = ttk.Frame(card, style="Card.TFrame")
            top.pack(fill=tk.X)
            ttk.Label(top, text=program, width=12, style="CardTitle.TLabel").pack(
                side=tk.LEFT
            )
            badge = tk.Label(
                top,
                text="대기",
                width=10,
                bg="#dfe6ee",
                fg="#223142",
                font=("Malgun Gothic", 9, "bold"),
                padx=6,
                pady=2,
            )
            badge.pack(side=tk.LEFT, padx=(0, 8))
            indicator = tk.Canvas(
                top, width=12, height=12, highlightthickness=0, bg="#ffffff"
            )
            indicator.pack(side=tk.LEFT, padx=(0, 8))
            _ = indicator.create_oval(2, 2, 10, 10, fill="#cbd5e1", outline="#94a3b8")
            detail = ttk.Label(top, text="", style="CardBody.TLabel")
            detail.pack(side=tk.LEFT, fill=tk.X, expand=True)
            chip_row = ttk.Frame(card, style="Card.TFrame")
            chip_row.pack(fill=tk.X, pady=(6, 0))
            chip_labels: list[ttk.Label] = []
            for _index in range(3):
                chip = ttk.Label(
                    chip_row,
                    text="-",
                    relief=tk.GROOVE,
                    padding=(6, 2),
                    width=14,
                )
                chip.pack(side=tk.LEFT, padx=(0, 6))
                chip_labels.append(chip)
            progress = ttk.Progressbar(card, mode="determinate", maximum=100)
            progress.pack(fill=tk.X, pady=(6, 0))
            row_widgets: dict[str, object] = {
                "chips": chip_labels,
                "indicator": indicator,
                "status": badge,
                "detail": detail,
                "progress": progress,
            }
            self.program_rows[program] = row_widgets

    def _build_queue_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(
            parent,
            text="Queue / Live Operations",
            padding=8,
            style="ManagerSection.TLabelframe",
        )
        frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        queue_container = ttk.Frame(frame)
        queue_container.pack(fill=tk.BOTH, expand=True)
        self.queue_list = tk.Listbox(queue_container, height=18, font=("Consolas", 9))
        queue_scroll = ttk.Scrollbar(
            queue_container,
            orient=tk.VERTICAL,
            command=self._queue_list_yview,
        )
        _ = self.queue_list.configure(
            yscrollcommand=cast(Callable[..., object], queue_scroll.set)
        )
        self.queue_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        queue_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    def _build_logs_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(
            parent, text="Event Console", padding=8, style="ManagerSection.TLabelframe"
        )
        frame.pack(fill=tk.BOTH, expand=True)
        log_container = ttk.Frame(frame)
        log_container.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(
            log_container,
            height=14,
            font=("Consolas", 9),
            bg="#1f252b",
            fg="#eceff1",
        )
        log_scroll = ttk.Scrollbar(
            log_container,
            orient=tk.VERTICAL,
            command=self._log_text_yview,
        )
        _ = self.log_text.configure(
            yscrollcommand=cast(Callable[..., object], log_scroll.set),
            state=tk.DISABLED,
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

    def _append_log(self, message: str) -> None:
        timestamp = time()
        line = f"[{timestamp:.3f}] {message}"
        self.log_buffer.append(line)
        trimmed = False
        if len(self.log_buffer) > 300:
            self.log_buffer = self.log_buffer[-300:]
            trimmed = True
        self._render_log_buffer(force_full_render=trimmed)

    def _render_log_buffer(self, *, force_full_render: bool = False) -> None:
        if self.log_text is None:
            return
        filter_text = self.log_filter_text.get().strip().lower()
        level = self.log_level_filter_text.get().strip().upper()
        source_lines = (
            self.log_buffer[-100:]
            if self.log_recent_only_var.get()
            else self.log_buffer
        )
        filtered_lines: list[str] = []
        for line in source_lines:
            upper_line = line.upper()
            if (
                level == "WARNING"
                and "WARNING" not in upper_line
                and "WARN" not in upper_line
            ):
                continue
            if (
                level == "ERROR"
                and "ERROR" not in upper_line
                and "FAIL" not in upper_line
            ):
                continue
            if level == "INFO" and (
                "ERROR" in upper_line
                or "FAIL" in upper_line
                or "WARNING" in upper_line
                or "WARN" in upper_line
            ):
                continue
            if filter_text and filter_text not in line.lower():
                continue
            filtered_lines.append(line)
        _ = self.log_text.configure(state=tk.NORMAL)
        if force_full_render or True:
            self.log_text.delete("1.0", tk.END)
            self.log_text.insert(tk.END, "\n".join(filtered_lines))
        if self.log_autoscroll_var.get():
            self.log_text.see(tk.END)
        _ = self.log_text.configure(state=tk.DISABLED)

    def _on_log_filter_changed(self, _event: object | None = None) -> None:
        self._render_log_buffer(force_full_render=True)
        self._save_settings()

    def _apply_named_preset(self, preset_name: str) -> None:
        if preset_name == "all":
            self.preset_gpt_var.set(True)
            self.preset_browser_var.set(True)
            self.preset_seed_var.set(True)
            self.preset_control_var.set(True)
        elif preset_name == "minimal":
            self.preset_gpt_var.set(False)
            self.preset_browser_var.set(True)
            self.preset_seed_var.set(True)
            self.preset_control_var.set(True)
        elif preset_name == "recovery":
            self.preset_gpt_var.set(False)
            self.preset_browser_var.set(False)
            self.preset_seed_var.set(False)
            self.preset_control_var.set(False)
        self._apply_home_presets()

    def _apply_home_presets(self) -> None:
        selected_steps: list[str] = []
        if self.preset_gpt_var.get():
            selected_steps.append("GPT 준비")
        if self.preset_browser_var.get():
            selected_steps.append("브라우저 로그인")
        if self.preset_seed_var.get():
            selected_steps.append("행 적재")
        if self.preset_control_var.get():
            selected_steps.append("제어 실행")
        if selected_steps:
            self.home_warning_text.set("프리셋: " + " -> ".join(selected_steps))
        else:
            self.home_warning_text.set("프리셋이 선택되지 않았습니다")
        preset_states = {
            "gpt_spawn": self.preset_gpt_var.get(),
            "open_browser": self.preset_browser_var.get(),
            "seed": self.preset_seed_var.get(),
            "control": self.preset_control_var.get(),
        }
        for button_key, enabled in preset_states.items():
            button = self.home_action_buttons.get(button_key)
            if button is not None:
                _ = button.config(state=(tk.NORMAL if enabled else tk.DISABLED))
        self._save_settings()

    def _visible_log_text(self) -> str:
        if self.log_text is None:
            return ""
        return str(self.log_text.get("1.0", "end-1c"))

    def clear_logs(self) -> None:
        self.log_buffer.clear()
        if self.log_text is None:
            return
        _ = self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        _ = self.log_text.configure(state=tk.DISABLED)

    def open_events_file(self) -> None:
        path = self.config.control_plane_events_file
        if not path.exists():
            _ = messagebox.showwarning(
                "runtime_v2", "이벤트 로그 파일이 아직 없습니다."
            )
            return
        os.startfile(str(path))

    def open_debug_dir(self) -> None:
        path = self.config.debug_log_root
        path.mkdir(parents=True, exist_ok=True)
        os.startfile(str(path))

    def copy_visible_logs(self) -> None:
        visible = self._visible_log_text()
        self.root.clipboard_clear()
        self.root.clipboard_append(visible)
        self.last_result_text.set("현재 표시 중인 로그를 클립보드에 복사했습니다")

    def save_visible_logs(self) -> None:
        visible = self._visible_log_text()
        target = filedialog.asksaveasfilename(
            title="로그 저장",
            defaultextension=".log",
            filetypes=(
                ("Log files", "*.log"),
                ("Text files", "*.txt"),
                ("All files", "*.*"),
            ),
        )
        if not target:
            return
        _ = Path(target).write_text(visible, encoding="utf-8")
        self.last_result_text.set(f"로그를 저장했습니다: {target}")

    def extract_error_logs(self) -> None:
        self.log_level_filter_text.set("ERROR")
        self._on_log_filter_changed()
        self.last_result_text.set("에러/실패 로그만 표시 중입니다")

    def _file_stamp(self, path: Path) -> tuple[float, int] | None:
        if not path.exists():
            return None
        stat = path.stat()
        return (stat.st_mtime, stat.st_size)

    def _read_cached_json(self, path: Path) -> dict[str, object] | None:
        key = (str(path.resolve()), 0)
        stamp = self._file_stamp(path)
        if stamp is None:
            _ = self._json_cache.pop(key, None)
            return None
        cached = self._json_cache.get(key)
        if cached is not None and cached[:2] == stamp:
            payload = cached[2]
            return (
                cast(dict[str, object], payload) if isinstance(payload, dict) else None
            )
        payload = _read_json(path)
        self._json_cache[key] = (stamp[0], stamp[1], payload)
        return payload

    def _read_cached_json_list(self, path: Path) -> list[dict[str, object]]:
        key = (str(path.resolve()), 1)
        stamp = self._file_stamp(path)
        if stamp is None:
            _ = self._json_cache.pop(key, None)
            return []
        cached = self._json_cache.get(key)
        if cached is not None and cached[:2] == stamp:
            payload = cached[2]
            return (
                cast(list[dict[str, object]], payload)
                if isinstance(payload, list)
                else []
            )
        payload = _read_json_list(path)
        self._json_cache[key] = (stamp[0], stamp[1], payload)
        return payload

    def _read_cached_jsonl_tail(
        self, path: Path, *, limit: int = 40
    ) -> list[dict[str, object]]:
        key = (str(path.resolve()), limit + 1000)
        stamp = self._file_stamp(path)
        if stamp is None:
            _ = self._json_cache.pop(key, None)
            return []
        cached = self._json_cache.get(key)
        if cached is not None and cached[:2] == stamp:
            payload = cached[2]
            return (
                cast(list[dict[str, object]], payload)
                if isinstance(payload, list)
                else []
            )
        payload = _read_jsonl_tail(path, limit=limit)
        self._json_cache[key] = (stamp[0], stamp[1], payload)
        return payload

    def _summarize_control_result(self, result: dict[str, object]) -> str:
        summary = summarize_runtime_result(result)
        log_ref = debug_log_path(
            self.config.debug_log_root, str(summary.get("job_id", "gui-session"))
        )
        parts = [
            f"status={summary.get('status', '?')}",
            f"code={summary.get('code', '')}",
        ]
        if str(summary.get("job_id", "")):
            parts.append(f"job={summary.get('job_id', '')}")
        if str(summary.get("workload", "")):
            parts.append(f"workload={summary.get('workload', '')}")
        if str(summary.get("stage", "")):
            parts.append(f"stage={summary.get('stage', '')}")
        if str(summary.get("error_code", "")):
            parts.append(f"error={summary.get('error_code', '')}")
        parts.append(f"debug={log_ref}")
        return " ".join(parts)

    def _enqueue_ui(self, action: str, value: str) -> None:
        self.ui_queue.put((action, value))

    def _drain_ui_queue(self) -> None:
        while True:
            try:
                action, value = self.ui_queue.get_nowait()
            except queue.Empty:
                break
            if action == "log":
                self._append_log(value)
            elif action == "last_result":
                self.last_result_text.set(value)
            elif action == "status":
                self.status_text.set(value)
            elif action == "seed_summary":
                self.seed_result_summary_text.set(value)
            elif action == "batch_summary":
                self.batch_result_summary_text.set(value)
            elif action == "soak_summary":
                self.soak_summary_text.set(value)
            elif action == "refresh":
                self._refresh_dashboard_once()

    def _set_program_row(
        self,
        program: str,
        *,
        status: str | None = None,
        detail: str | None = None,
        chips: tuple[str, str, str] | None = None,
    ) -> None:
        row = self.program_rows.get(program)
        if row is None:
            return
        if status is not None:
            badge_bg = "#dfe6ee"
            badge_fg = "#223142"
            indicator_fill = "#cbd5e1"
            progress_value = 15
            normalized = status.strip().lower()
            if any(
                token in normalized for token in ("정상", "ready", "완료", "최종완료")
            ):
                badge_bg = "#d5f5e3"
                badge_fg = "#166534"
                indicator_fill = "#22c55e"
                progress_value = 100
            elif any(
                token in normalized
                for token in ("실행", "running", "seeded", "boot", "retry")
            ):
                badge_bg = "#dbeafe"
                badge_fg = "#1d4ed8"
                indicator_fill = "#3b82f6"
                progress_value = 70
            elif any(
                token in normalized
                for token in ("경고", "blocked", "fail", "error", "missing", "stale")
            ):
                badge_bg = "#fee2e2"
                badge_fg = "#b91c1c"
                indicator_fill = "#ef4444"
                progress_value = 35
            elif any(
                token in normalized for token in ("유휴", "대기", "idle", "unknown")
            ):
                badge_bg = "#e5e7eb"
                badge_fg = "#374151"
                indicator_fill = "#94a3b8"
                progress_value = 15
            badge = row["status"]
            progress = row["progress"]
            indicator = row.get("indicator")
            if isinstance(badge, tk.Label):
                _ = badge.config(text=status, bg=badge_bg, fg=badge_fg)
            if isinstance(progress, ttk.Progressbar):
                _ = progress.config(value=progress_value)
            if isinstance(indicator, tk.Canvas):
                _ = indicator.delete("all")
                _ = indicator.create_oval(
                    2, 2, 10, 10, fill=indicator_fill, outline=indicator_fill
                )
        if detail is not None:
            detail_label = row["detail"]
            if isinstance(detail_label, ttk.Label):
                _ = detail_label.config(text=detail)
        if chips is not None:
            chip_widgets = row.get("chips")
            if isinstance(chip_widgets, list):
                typed_chip_widgets = cast(list[ttk.Label], chip_widgets)
                for chip_widget, chip_text in zip(typed_chip_widgets, chips):
                    _ = chip_widget.config(text=chip_text)

    def _snapshot_warning(
        self, name: str, payload: dict[str, object] | None, fresh_sec: int
    ) -> str | None:
        if payload is None:
            return f"{name} snapshot missing"
        checked_at = _to_float(payload.get("checked_at", 0.0), 0.0)
        if checked_at <= 0:
            return f"{name} snapshot timestamp missing"
        age = max(0.0, time() - checked_at)
        if age > fresh_sec:
            return f"{name} snapshot stale ({int(age)}s)"
        return None

    def _recent_time_chip(self, raw_ts: object) -> str:
        ts = _to_float(raw_ts, 0.0)
        if ts <= 0:
            return "updated -"
        age = max(0, int(time() - ts))
        if age < 60:
            return f"updated {age}s"
        if age < 3600:
            return f"updated {age // 60}m"
        return f"updated {age // 3600}h"

    def _load_settings(self) -> None:
        payload = _read_json(self.SETTINGS_FILE)
        if payload is None:
            return
        self.runtime_root_text.set(str(payload.get("runtime_root", "")))
        self.poll_ms_text.set(str(payload.get("poll_ms", "1000")))
        self.loop_sleep_text.set(str(payload.get("loop_sleep_sec", "2.0")))
        self.excel_path_text.set(
            str(payload.get("excel_path", self.excel_path_text.get()))
        )
        self.sheet_name_text.set(str(payload.get("sheet_name", "Sheet1")))
        self.row_index_text.set(str(payload.get("row_index", "0")))
        self.batch_count_text.set(str(payload.get("batch_count", "5")))
        self.max_control_ticks_text.set(str(payload.get("max_control_ticks", "50")))
        login_service = str(payload.get("login_service", "chatgpt")).strip().lower()
        if login_service in self.LOGIN_SERVICES:
            self.login_service_text.set(login_service)
        workload = str(payload.get("selected_workload", "qwen3_tts"))
        if workload in {"qwen3_tts", "rvc", "kenburns"}:
            self.selected_workload.set(workload)
        self.source_path_text.set(str(payload.get("source_path", "")))
        self.audio_path_text.set(str(payload.get("audio_path", "")))
        self.script_text.set(str(payload.get("script_text", "")))
        self.duration_text.set(str(payload.get("duration_sec", "8")))
        self.model_name_text.set(str(payload.get("model_name", "")))
        self.preset_gpt_var.set(bool(payload.get("preset_gpt", True)))
        self.preset_browser_var.set(bool(payload.get("preset_browser", True)))
        self.preset_seed_var.set(bool(payload.get("preset_seed", True)))
        self.preset_control_var.set(bool(payload.get("preset_control", True)))
        self.log_filter_text.set(str(payload.get("log_filter_text", "")))
        log_level = str(payload.get("log_level_filter", "전체"))
        self.log_level_filter_text.set(log_level if log_level else "전체")
        self.log_autoscroll_var.set(bool(payload.get("log_autoscroll", True)))
        self.log_recent_only_var.set(bool(payload.get("log_recent_only", False)))

    def _save_settings(self) -> None:
        self.SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "runtime_root": self.runtime_root_text.get().strip(),
            "poll_ms": self.poll_ms_text.get().strip(),
            "loop_sleep_sec": self.loop_sleep_text.get().strip(),
            "excel_path": self.excel_path_text.get().strip(),
            "sheet_name": self.sheet_name_text.get().strip(),
            "row_index": self.row_index_text.get().strip(),
            "batch_count": self.batch_count_text.get().strip(),
            "max_control_ticks": self.max_control_ticks_text.get().strip(),
            "login_service": self.login_service_text.get().strip(),
            "selected_workload": self.selected_workload.get(),
            "source_path": self.source_path_text.get().strip(),
            "audio_path": self.audio_path_text.get().strip(),
            "script_text": self.script_text.get().strip(),
            "duration_sec": self.duration_text.get().strip(),
            "model_name": self.model_name_text.get().strip(),
            "preset_gpt": self.preset_gpt_var.get(),
            "preset_browser": self.preset_browser_var.get(),
            "preset_seed": self.preset_seed_var.get(),
            "preset_control": self.preset_control_var.get(),
            "log_filter_text": self.log_filter_text.get().strip(),
            "log_level_filter": self.log_level_filter_text.get().strip(),
            "log_autoscroll": self.log_autoscroll_var.get(),
            "log_recent_only": self.log_recent_only_var.get(),
        }
        _ = self.SETTINGS_FILE.write_text(
            json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8"
        )

    def _refresh_runtime_config(self, *, force: bool = False) -> None:
        action_running = (
            self.action_thread is not None and self.action_thread.is_alive()
        )
        if not force and (self.running or action_running):
            return
        self.config = _runtime_config_from_text(self.runtime_root_text.get())
        parsed_poll = _to_int(self.poll_ms_text.get().strip())
        self.poll_ms = parsed_poll if parsed_poll > 0 else 1000
        parsed_sleep = _to_float(self.loop_sleep_text.get().strip(), 2.0)
        self.loop_sleep_sec = parsed_sleep if parsed_sleep > 0 else 2.0

    def _parsed_row_index(self) -> int | None:
        raw = self.row_index_text.get().strip()
        if not raw:
            return 0
        try:
            row_index = int(raw)
        except ValueError:
            _ = messagebox.showwarning("runtime_v2", "row index는 정수여야 합니다.")
            return None
        if row_index < 0:
            _ = messagebox.showwarning("runtime_v2", "row index는 0 이상이어야 합니다.")
            return None
        return row_index

    def _start_background_action(
        self, action_name: str, action: Callable[[], None]
    ) -> None:
        if self.action_thread is not None and self.action_thread.is_alive():
            _ = messagebox.showwarning("runtime_v2", "다른 작업이 아직 실행 중입니다.")
            return

        def runner() -> None:
            self._enqueue_ui("status", f"{action_name} 실행중")
            try:
                action()
            except Exception as exc:
                _ = append_debug_event(
                    debug_log_path(self.config.debug_log_root, "gui-session"),
                    event="gui_manual_action_exception",
                    level="ERROR",
                    payload={
                        "owner": self.owner,
                        "action_name": action_name,
                        **exception_payload(exc),
                    },
                )
                self._enqueue_ui("log", f"{action_name} failed: {exc}")
                self._enqueue_ui("last_result", f"{action_name}: failed {exc}")
            finally:
                self._enqueue_ui("status", "실행중" if self.running else "대기")
                self._enqueue_ui("refresh", "")

        self.action_thread = threading.Thread(target=runner, daemon=True)
        self.action_thread.start()

    def refresh_dashboard_now(self) -> None:
        self._refresh_runtime_config(force=True)
        self._save_settings()
        self._refresh_sheet_list()
        self._refresh_dashboard_once()
        self._append_log("manual health refresh")

    def trigger_excel_seed(self) -> None:
        if self.running:
            _ = messagebox.showwarning(
                "runtime_v2",
                "manual Stage 5 seed 전에는 long-running loop를 먼저 중지하세요.",
            )
            return
        self._refresh_runtime_config(force=True)
        self._refresh_sheet_list()
        row_index = self._parsed_row_index()
        if row_index is None:
            return
        excel_path = self.excel_path_text.get().strip()
        if not excel_path:
            _ = messagebox.showwarning("runtime_v2", "excel path를 입력하세요.")
            return
        run_id = str(uuid4())
        seed_result = seed_excel_row(
            config=self.config,
            run_id=run_id,
            excel_path=excel_path,
            sheet_name=self.sheet_name_text.get().strip() or "Sheet1",
            row_index=row_index,
        )
        self._save_settings()
        summary = _format_seed_summary(
            seed_result,
            excel_path=excel_path,
            sheet_name=self.sheet_name_text.get().strip() or "Sheet1",
            row_index=row_index,
        )
        self.seed_result_summary_text.set(summary)
        self.latest_run_text.set(f"run_id: {run_id}")
        self.last_result_text.set(summary)
        self._append_log(summary)
        self._refresh_dashboard_once()

    def trigger_excel_batch(self) -> None:
        if self.running:
            _ = messagebox.showwarning(
                "runtime_v2",
                "batch 실행 전에는 long-running loop를 먼저 중지하세요.",
            )
            return
        self._refresh_runtime_config(force=True)
        batch_count = max(1, _to_int(self.batch_count_text.get().strip()))
        max_ticks = max(1, _to_int(self.max_control_ticks_text.get().strip()))
        excel_path = self.excel_path_text.get().strip()
        sheet_name = self.sheet_name_text.get().strip() or "Sheet1"
        if not excel_path:
            _ = messagebox.showwarning("runtime_v2", "excel path를 입력하세요.")
            return

        def action() -> None:
            result = _run_excel_batch_mode(
                owner=self.owner,
                config=self.config,
                run_id=str(uuid4()),
                excel_path=excel_path,
                sheet_name=sheet_name,
                batch_count=batch_count,
                max_control_ticks=max_ticks,
            )
            summary = _format_batch_summary(result)
            self._enqueue_ui("last_result", summary)
            self._enqueue_ui("log", summary)
            self._enqueue_ui("batch_summary", summary)
            self._enqueue_ui("refresh", "")

        self._start_background_action("excel batch", action)

    def trigger_soak_report(self) -> None:
        self._refresh_runtime_config(force=True)

        def action() -> None:
            report_path = write_soak_report(self.config)
            summary = _format_soak_summary(str(report_path.resolve()))
            self._enqueue_ui("last_result", summary)
            self._enqueue_ui("log", summary)
            self._enqueue_ui("soak_summary", summary)
            self._enqueue_ui("refresh", "")

        self._start_background_action("soak report", action)

    def _manual_control_stop_reasons(self) -> list[str]:
        self._refresh_runtime_config(force=True)
        readiness = load_runtime_readiness(self.config, completed=True)
        stop_reasons = _readiness_blocker_messages(readiness)
        browser_registry = self._read_cached_json(self.config.browser_registry_file)
        stop_reasons.extend(_blocking_browser_service_messages(browser_registry))
        return stop_reasons

    def trigger_control_once(self) -> None:
        if self.running:
            _ = messagebox.showwarning(
                "runtime_v2",
                "manual Stage 5 control 전에는 long-running loop를 먼저 중지하세요.",
            )
            return
        stop_reasons = self._manual_control_stop_reasons()
        if stop_reasons:
            message = " | ".join(stop_reasons[:4])
            self._append_log(f"control once blocked: {message}")
            self.last_result_text.set(f"control once blocked: {message}")
            _ = messagebox.showwarning("runtime_v2", f"No-Go 상태입니다: {message}")
            self._refresh_dashboard_once()
            return
        self._refresh_runtime_config(force=True)

        def action() -> None:
            result = self._run_control_once()
            summary_line = self._summarize_control_result(result)
            self._enqueue_ui("last_result", f"control once: {summary_line}")
            self._enqueue_ui("log", f"control once: {summary_line}")

        self._start_background_action("control once", action)

    def trigger_open_login_browser(self) -> None:
        service = self.login_service_text.get().strip().lower()
        if service not in self.LOGIN_SERVICES:
            _ = messagebox.showwarning(
                "runtime_v2", "지원하지 않는 브라우저 서비스입니다."
            )
            return

        def action() -> None:
            payload = open_browser_for_login(service)
            summary = (
                f"login browser: service={payload.get('service', service)} "
                f"port={payload.get('port', 0)} launched={payload.get('launched', False)}"
            )
            self._enqueue_ui("last_result", summary)
            self._enqueue_ui("log", summary)

        self._start_background_action("open login browser", action)

    def start_loop(self) -> None:
        if self.running:
            return
        self.running = True
        self.paused = False
        self.stop_requested = False
        self.stop_event.clear()
        self.pause_event.clear()
        self._save_settings()
        self.status_text.set("실행중")
        self._append_log("control loop started")
        self.worker_thread = threading.Thread(target=self._run_loop, daemon=True)
        self.worker_thread.start()

    def toggle_pause(self) -> None:
        if not self.running:
            return
        self.paused = not self.paused
        if self.paused:
            self.pause_event.set()
        else:
            self.pause_event.clear()
        self.status_text.set("일시정지" if self.paused else "실행중")
        self._append_log(
            "control loop paused" if self.paused else "control loop resumed"
        )

    def stop_loop(self) -> None:
        self.stop_requested = True
        self.running = False
        self.paused = False
        self.stop_event.set()
        self.pause_event.clear()
        self.status_text.set("중지")
        self._append_log("control loop stopped")

    def trigger_recovery(self) -> None:
        workload = self.selected_workload.get()
        self._append_log(f"manual recovery triggered for {workload}")
        result = self._run_control_once()
        self.last_result_text.set(
            f"recovery: {str(result.get('result', result))[:120]}"
        )
        self._refresh_dashboard_once()

    def trigger_gpt_spawn(self) -> None:
        status = load_gpt_status(self.config.gpt_status_file)
        if status is None:
            _ = messagebox.showwarning("runtime_v2", "gpt_status.json 이 없습니다.")
            return
        pending_boot_value = status.get("pending_boot", 0)
        pending_boot = (
            int(pending_boot_value)
            if isinstance(pending_boot_value, (int, float, str))
            else 0
        )
        status["pending_boot"] = pending_boot + 1
        status["last_spawn_at"] = round(time(), 3)
        _ = write_gpt_status(status, self.config.gpt_status_file)
        self._append_log("manual GPT spawn check triggered")
        self._refresh_dashboard_once()

    def trigger_scan_inbox(self) -> None:
        seeded_jobs = seed_local_jobs(self.config)
        self._append_log(f"scan inbox seeded={len(seeded_jobs)}")
        self._refresh_dashboard_once()

    def seed_job(self, workload: GpuWorkload) -> None:
        payload = self._build_job_payload(workload)
        if payload is None:
            return
        job = JobContract(
            job_id=f"{workload}-{uuid4()}",
            workload=workload,
            status="queued",
            payload=payload,
        )
        seed_control_job(job=job, config=self.config)
        self._save_settings()
        self._append_log(f"queued job: {job.job_id} payload={payload}")
        self._refresh_dashboard_once()

    def _build_job_payload(self, workload: GpuWorkload) -> dict[str, object] | None:
        payload: dict[str, object] = {}
        source_path = self.source_path_text.get().strip()
        audio_path = self.audio_path_text.get().strip()
        script_text = self.script_text.get().strip()
        duration_sec = _to_int(self.duration_text.get().strip())
        if workload == "qwen3_tts":
            if not script_text:
                _ = messagebox.showwarning("runtime_v2", "script 값을 입력하세요.")
                return None
            payload["script_text"] = script_text
            if source_path:
                payload["image_path"] = source_path
            return payload
        if not source_path:
            _ = messagebox.showwarning("runtime_v2", "path 값을 입력하세요.")
            return None
        payload["source_path"] = source_path
        if workload == "kenburns":
            payload["duration_sec"] = max(1, duration_sec)
        if workload == "rvc" and audio_path:
            payload["audio_path"] = audio_path
        model_name = self.model_name_text.get().strip()
        if workload == "rvc" and model_name:
            payload["model_name"] = model_name
        return payload

    def _run_loop(self) -> None:
        while self.running and not self.stop_event.is_set():
            try:
                if self.pause_event.is_set():
                    sleep(self.loop_sleep_sec)
                    continue
                result = self._run_control_once()
                result_summary = summarize_runtime_result(result)
                summary_line = self._summarize_control_result(result)
                self._enqueue_ui("last_result", f"last: {summary_line}")
                self._enqueue_ui("log", f"control result: {summary_line}")
                _ = append_debug_event(
                    debug_log_path(
                        self.config.debug_log_root,
                        str(result_summary.get("job_id", "gui-session")),
                    ),
                    event="gui_control_result",
                    level="ERROR"
                    if str(result.get("status", "")) == "failed"
                    else "INFO",
                    payload={
                        "owner": self.owner,
                        "result": result,
                    },
                )
                recovery = _coerce_mapping(result.get("recovery", {}))
                backoff_value = self.loop_sleep_sec
                if recovery is not None:
                    backoff_value = _to_float(
                        recovery.get("backoff_sec", self.loop_sleep_sec),
                        self.loop_sleep_sec,
                    )
                if result.get("code") == "NO_JOB":
                    sleep(max(self.loop_sleep_sec, 5.0))
                    continue
                sleep(max(self.loop_sleep_sec, backoff_value))
            except BaseException as exc:
                _ = append_debug_event(
                    debug_log_path(self.config.debug_log_root, "gui-session"),
                    event="gui_loop_exception",
                    level="ERROR",
                    payload={
                        "owner": self.owner,
                        **exception_payload(exc),
                    },
                )
                raise

    def refresh_dashboard(self) -> None:
        try:
            self._drain_ui_queue()
            self._refresh_dashboard_once()
        except Exception as exc:
            self._append_log(f"dashboard refresh failed: {exc}")
            self.warning_text.set(f"dashboard refresh failed: {exc}")
            _ = append_debug_event(
                debug_log_path(self.config.debug_log_root, "gui-session"),
                event="gui_refresh_exception",
                level="ERROR",
                payload={
                    "owner": self.owner,
                    **exception_payload(exc),
                },
            )
        finally:
            self.refresh_after_id = self.root.after(
                self.poll_ms, self.refresh_dashboard
            )

    def _run_control_once(self) -> dict[str, object]:
        if not self.control_lock.acquire(blocking=False):
            return {
                "status": "busy",
                "code": "CONTROL_BUSY",
                "result": "control loop busy",
            }
        try:
            return run_control_loop_once(owner=self.owner, config=self.config)
        finally:
            self.control_lock.release()

    def _refresh_dashboard_once(self) -> None:
        self._refresh_runtime_config()
        self._ensure_snapshot_contracts()
        self._update_operator_panel()
        self._update_browser_panel()
        self._update_gpu_panel()
        self._update_gpt_panel()
        self._update_queue_panel()
        self._update_artifact_panel()
        self._update_program_panel()
        self._update_warning_panel()
        self._update_log_panel()

    def _ensure_snapshot_contracts(self) -> None:
        ensure_runtime_bootstrap(
            self.config,
            workload=self._selected_gpu_workload(),
            run_id="gui-bootstrap",
            mode="local_gui",
        )

    def _update_operator_panel(self) -> None:
        readiness = load_runtime_readiness(self.config, completed=True)
        blocker_messages = _readiness_blocker_messages(readiness)
        if bool(readiness.get("ready", False)) and not blocker_messages:
            self.readiness_text.set("ready")
        elif blocker_messages:
            self.readiness_text.set(
                _truncate_ui_text(" | ".join(blocker_messages[:2]), 60)
            )
        else:
            self.readiness_text.set(
                str(readiness.get("code", "UNKNOWN")).strip() or "UNKNOWN"
            )

        result_payload = self._read_cached_json(self.config.result_router_file)
        gui_payload = self._read_cached_json(self.config.gui_status_file)
        self.evidence_summary_text.set(
            _format_terminal_evidence_summary(result_payload, gui_payload)
        )

        result_metadata = (
            None
            if result_payload is None
            else _coerce_mapping(result_payload.get("metadata"))
        )
        latest_run_id = ""
        if result_metadata is not None:
            latest_run_id = str(result_metadata.get("run_id", "")).strip()
        if not latest_run_id and gui_payload is not None:
            latest_run_id = str(gui_payload.get("run_id", "")).strip()
        if not latest_run_id:
            latest_run_id = str(readiness.get("snapshot_run_id", "")).strip()
        self.latest_run_text.set(f"run_id: {latest_run_id or '-'}")
        self.home_ready_text.set(_friendly_ready_text(readiness, blocker_messages))
        self.home_recent_text.set(
            "최근 실행이 없습니다"
            if not latest_run_id
            else _truncate_ui_text(f"최근 실행 ID {latest_run_id}", 40)
        )
        self.home_result_text.set(
            _truncate_ui_text(
                _friendly_evidence_text(self.evidence_summary_text.get()), 40
            )
        )

    def _normalize_gpu_health_snapshot(self) -> None:
        payload = self._read_cached_json(self.config.lease_file)
        if payload is None:
            _ = write_gpu_health_payload(
                build_gpu_health_payload(
                    self._selected_gpu_workload(),
                    lock_key=f"lock:{self._selected_gpu_workload()}",
                    lease=None,
                    event="idle",
                ),
                self.config.lease_file,
            )
            return
        has_schema = isinstance(payload.get("schema_version"), str) and isinstance(
            payload.get("checked_at"), (int, float)
        )
        if has_schema:
            return
        lease_payload = None
        lock_key = ""
        if all(key in payload for key in ("key", "owner", "expires_at")):
            lease_payload = payload
            lock_key = str(payload.get("key", "")).strip()
        workload = _workload_from_lock_key(lock_key) or self._selected_gpu_workload()
        normalized = build_gpu_health_payload(
            workload,
            lock_key=lock_key or f"lock:{workload}",
            lease=None
            if lease_payload is None
            else Lease.from_dict(_lease_like_payload(lease_payload)),
            event="normalized_legacy_snapshot" if lease_payload is not None else "idle",
        )
        _ = write_gpu_health_payload(normalized, self.config.lease_file)

    def _normalize_gui_status_snapshot(self) -> None:
        payload = self._read_cached_json(self.config.gui_status_file)
        if payload is None:
            return
        has_schema = isinstance(payload.get("schema_version"), str) and isinstance(
            payload.get("checked_at"), (int, float)
        )
        if has_schema:
            return
        status_payload = _coerce_mapping(payload.get("status", {}))
        normalized = build_gui_status_payload(
            {} if status_payload is None else status_payload,
            run_id=str(payload.get("run_id", "gui-bootstrap")),
            mode=str(payload.get("mode", "local_gui")),
            stage=str(payload.get("stage", "idle")),
            exit_code=_to_int(payload.get("exit_code", 0)),
        )
        _ = write_gui_status(normalized, self.config.gui_status_file)

    def _selected_gpu_workload(self) -> GpuWorkload:
        workload = self.selected_workload.get().strip()
        if workload == "rvc":
            return "rvc"
        if workload == "kenburns":
            return "kenburns"
        return "qwen3_tts"

    def _update_browser_panel(self) -> None:
        payload = self._read_cached_json(self.config.browser_health_file)
        registry = self._read_cached_json(self.config.browser_registry_file)
        if payload is None:
            self.browser_text.set("browser missing")
            self.browser_progress_value.set(0)
            self.browser_services_text.set("브라우저 상태 정보를 찾을 수 없습니다")
            self.home_browser_text.set("브라우저 상태 정보를 찾을 수 없습니다")
            return
        healthy = _to_int(payload.get("healthy_count", 0))
        total = _to_int(payload.get("session_count", 0))
        self.browser_text.set(f"browser {healthy}/{total} healthy")
        self.browser_progress_value.set(
            0 if total <= 0 else max(0, min(100, int((healthy / total) * 100)))
        )
        service_summaries: list[str] = []
        if registry is not None:
            sessions = registry.get("sessions", [])
            if isinstance(sessions, list):
                typed_sessions = cast(list[object], sessions)
                row_map = {
                    "chatgpt": "GPT",
                    "genspark": "Genspark",
                    "seaart": "SeaArt",
                    "geminigen": "GeminiGen",
                    "canva": "Browser",
                }
                for item in typed_sessions:
                    item_dict = _coerce_mapping(item)
                    if item_dict is None:
                        continue
                    row_name = row_map.get(
                        str(item_dict.get("service", "")).strip().lower()
                    )
                    if row_name is None:
                        continue
                    row_status = str(item_dict.get("status", "대기"))
                    port = _to_int(item_dict.get("port", 0))
                    failures = _to_int(item_dict.get("consecutive_failures", 0))
                    self._set_program_row(
                        row_name,
                        status=row_status,
                        detail=f"port={port} fail={failures}",
                        chips=(
                            f"port {port}",
                            f"fail {failures}",
                            self._recent_time_chip(item_dict.get("checked_at", 0)),
                        ),
                    )
                    service_summaries.append(f"{row_name}={row_status}")
        browser_summary = (
            ", ".join(service_summaries)
            if service_summaries
            else "실행 중인 브라우저가 없습니다"
        )
        self.browser_services_text.set(browser_summary)
        if healthy == total and total > 0:
            self.home_browser_text.set("브라우저가 모두 준비되었습니다")
        elif healthy == 0:
            self.home_browser_text.set("로그인 또는 브라우저 준비가 필요합니다")
        else:
            self.home_browser_text.set(_truncate_ui_text(browser_summary, 40))

    def _update_gpu_panel(self) -> None:
        payload = self._read_cached_json(self.config.lease_file)
        if payload is None:
            self.gpu_text.set("gpu: missing")
            return
        self.gpu_text.set(
            f"gpu: {payload.get('workload', '?')} / {payload.get('event', 'unknown')}"
        )
        lease_payload = _coerce_mapping(payload.get("lease"))
        owner = "-" if lease_payload is None else str(lease_payload.get("owner", "-"))
        workload = str(payload.get("workload", "qwen3_tts")).strip().lower()
        row_name = _program_row_for_workload(workload)
        self._set_program_row(
            row_name,
            status=str(payload.get("event", "대기")),
            detail=f"owner={owner}",
            chips=(
                f"owner {owner}",
                f"work {workload}",
                self._recent_time_chip(payload.get("checked_at", 0)),
            ),
        )

    def _update_gpt_panel(self) -> None:
        payload = load_gpt_status(self.config.gpt_status_file)
        if payload is None:
            self.gpt_text.set("gpt: missing")
            self.gpt_progress_value.set(0)
            return
        ok_count = _to_int(payload.get("ok_count", 0))
        pending_boot = _to_int(payload.get("pending_boot", 0))
        self.gpt_text.set(f"gpt: ok={ok_count}, pending_boot={pending_boot}")
        ready_score = ok_count * 25
        if pending_boot > 0:
            ready_score = min(95, ready_score)
        self.gpt_progress_value.set(max(0, min(100, ready_score)))
        self._set_program_row(
            "GPT",
            status="정상" if ok_count > 0 else ("실행" if pending_boot > 0 else "대기"),
            detail=f"ok={ok_count} pending={pending_boot}",
            chips=(
                f"ok {ok_count}",
                f"boot {pending_boot}",
                self._recent_time_chip(
                    payload.get("last_spawn_at", payload.get("checked_at", 0))
                ),
            ),
        )

    def _update_queue_panel(self) -> None:
        items = self._read_cached_json_list(self.config.queue_store_file)
        self.queue_text.set(f"queue: {len(items)}")
        self.queue_progress_value.set(max(0, min(100, len(items) * 10)))
        queue_list = self.queue_list
        if queue_list is None:
            return
        queue_list.delete(0, tk.END)
        for item in items[:100]:
            job_id = str(item.get("job_id", "unknown"))
            workload = str(item.get("workload", "?"))
            status = str(item.get("status", "queued"))
            attempts = item.get("attempts", 0)
            payload = _coerce_mapping(item.get("payload", {}))
            chain_depth = (
                0 if payload is None else _to_int(payload.get("chain_depth", 0))
            )
            routed_from = (
                "-" if payload is None else str(payload.get("routed_from", "-"))
            )
            queue_list.insert(
                tk.END,
                f"{status:10} {workload:10} d={chain_depth} tries={attempts} from={routed_from} {job_id}",
            )

    def _update_artifact_panel(self) -> None:
        payload = self._read_cached_json(self.config.result_router_file)
        if payload is None:
            inbox_root = self.config.input_root
            accepted = _archived_contract_count(inbox_root / "accepted")
            invalid = _archived_contract_count(inbox_root / "invalid")
            self.artifact_text.set(
                f"artifacts: 0 / accepted={accepted} invalid={invalid}"
            )
            self.artifact_progress_value.set(0)
            return
        artifacts_obj = payload.get("artifacts", [])
        artifacts = (
            cast(list[object], artifacts_obj) if isinstance(artifacts_obj, list) else []
        )
        count = len(artifacts)
        metadata = _coerce_mapping(payload.get("metadata", {}))
        result_job = "-" if metadata is None else str(metadata.get("job_id", "-"))
        result_run_id = "-" if metadata is None else str(metadata.get("run_id", "-"))
        completion_state = (
            "-" if metadata is None else str(metadata.get("completion_state", "-"))
        )
        final_output = (
            False if metadata is None else bool(metadata.get("final_output", False))
        )
        final_artifact = (
            "" if metadata is None else str(metadata.get("final_artifact", ""))
        )
        result_status = "-" if metadata is None else str(metadata.get("status", "-"))
        result_code = "-" if metadata is None else str(metadata.get("code", "-"))
        final_record = _latest_final_output_record(
            self.config.control_plane_events_file
        )
        last_final_artifact = (
            "" if final_record is None else str(final_record.get("final_artifact", ""))
        )
        inbox_root = self.config.input_root
        accepted = _archived_contract_count(inbox_root / "accepted")
        invalid = _archived_contract_count(inbox_root / "invalid")
        latest_text = f"artifacts: {count} / accepted={accepted} invalid={invalid} / job={result_job}"
        if last_final_artifact:
            latest_text = f"{latest_text} / last={last_final_artifact}"
        self.artifact_text.set(latest_text)
        self.artifact_progress_value.set(max(0, min(100, count * 20)))
        render_detail = f"art={count} acc={accepted} inv={invalid} run={result_run_id} code={result_code}"
        if final_output and final_artifact:
            render_detail = f"final={final_artifact}"
        elif last_final_artifact:
            render_detail = f"latest={result_code} last={last_final_artifact}"
        render_status = completion_state
        if final_output:
            render_status = "최종완료"
        elif result_status == "idle" or result_code == "NO_JOB":
            render_status = "유휴"
        elif not render_status or render_status == "-":
            render_status = "완료" if count > 0 else "대기"
        self._set_program_row(
            "Result",
            status=render_status,
            detail=render_detail[:40],
            chips=(
                f"run {result_run_id or '-'}",
                f"code {result_code}",
                self._recent_time_chip(payload.get("checked_at", 0)),
            ),
        )

    def _update_program_panel(self) -> None:
        gui_payload = self._read_cached_json(self.config.gui_status_file)
        if gui_payload is not None:
            status_payload = _coerce_mapping(gui_payload.get("status", {}))
            worker_stage = str(gui_payload.get("worker_stage", "")).strip()
            if not worker_stage and status_payload is not None:
                worker_stage = str(
                    status_payload.get("worker_stage", gui_payload.get("stage", "-"))
                ).strip()
            worker_error_code = str(gui_payload.get("worker_error_code", "")).strip()
            if not worker_error_code and status_payload is not None:
                worker_error_code = str(
                    status_payload.get("worker_error_code", "")
                ).strip()
            worker_error_display = _display_worker_error_code(worker_error_code)
            result_path = str(gui_payload.get("result_path", "")).strip()
            if not result_path and status_payload is not None:
                result_path = str(status_payload.get("result_path", "")).strip()
            manifest_path = str(gui_payload.get("manifest_path", "")).strip()
            if not manifest_path and status_payload is not None:
                manifest_path = str(status_payload.get("manifest_path", "")).strip()
            backoff_sec = (
                0
                if status_payload is None
                else _to_int(status_payload.get("backoff_sec", 0))
            )
            self._set_program_row(
                "GUI",
                status=str(gui_payload.get("stage", "대기")),
                detail=(
                    f"exit={gui_payload.get('exit_code', '?')} stage={worker_stage} err={worker_error_display or '-'} "
                    f"retry={backoff_sec}s"
                )[:40],
                chips=(
                    f"run {str(gui_payload.get('run_id', '-'))}",
                    f"retry {backoff_sec}s",
                    self._recent_time_chip(gui_payload.get("checked_at", 0)),
                ),
            )
            if worker_error_code or str(gui_payload.get("exit_code", 0)) != "0":
                path_ref = result_path or manifest_path or "-"
                self.last_result_text.set(
                    f"실패정보: stage={worker_stage} error={worker_error_display or '-'} path={path_ref}"
                )

        browser_health = self._read_cached_json(self.config.browser_health_file)
        if browser_health is not None:
            unhealthy = browser_health.get("unhealthy_count", 0)
            self._set_program_row(
                "Browser",
                status="경고" if str(unhealthy) != "0" else "정상",
                detail=f"availability={browser_health.get('availability_percent', 0)}% unhealthy={unhealthy}"[
                    :40
                ],
                chips=(
                    f"avail {browser_health.get('availability_percent', 0)}%",
                    f"bad {unhealthy}",
                    self._recent_time_chip(browser_health.get("checked_at", 0)),
                ),
            )

    def _update_log_panel(self) -> None:
        event_records = self._read_cached_jsonl_tail(
            self.config.control_plane_events_file, limit=20
        )
        new_lines: list[str] = []
        latest_seen_ts = self.last_event_ts
        for record in event_records:
            ts = _to_float(record.get("ts", 0.0), 0.0)
            if ts <= self.last_event_ts:
                continue
            event_name = str(record.get("event", "")).strip()
            if event_name == "next_job_rejected":
                new_lines.append(
                    f"event next_job_rejected parent={record.get('parent_job_id', '?')} job={record.get('job_id', '-')} reason={record.get('reason', '?')}"
                )
            elif event_name == "stale_running_recovered":
                new_lines.append(
                    f"event stale_running job={record.get('job_id', '?')} action={record.get('action', '?')} age={record.get('age_sec', 0)}"
                )
            elif event_name == "job_summary":
                done_state = str(record.get("completion_state", "-") or "-")
                final_output = bool(record.get("final_output", False))
                final_artifact = str(record.get("final_artifact", "") or "-")
                worker_error_display = _display_worker_error_code(
                    str(record.get("worker_error_code", "-") or "-")
                )
                new_lines.append(
                    f"summary job={record.get('job_id', '?')} ok={record.get('success', False)} stage={record.get('worker_stage', '-')} err={worker_error_display} tries={record.get('attempts', 0)} backoff={record.get('backoff_sec', 0)} art={record.get('artifact_count', 0)} next={record.get('next_jobs_count', 0)} routed={record.get('routed_count', 0)} depth={record.get('chain_depth', 0)} from={record.get('routed_from', '-') or '-'} done={done_state} final={final_output} file={final_artifact}"
                )
            else:
                new_lines.append(
                    f"event job={record.get('job_id', '?')} {record.get('previous_status', '?')} -> {record.get('status', '?')} from={record.get('routed_from', '-')} depth={record.get('chain_depth', 0)}"
                )
            latest_seen_ts = max(latest_seen_ts, ts)
        for line in new_lines:
            self._append_log(line)
        self.last_event_ts = latest_seen_ts

    def _update_warning_panel(self) -> None:
        warnings: list[str] = []
        gui_status = self._read_cached_json(self.config.gui_status_file)
        gui_warning = self._snapshot_warning("gui", gui_status, fresh_sec=120)
        if gui_warning is not None:
            warnings.append(gui_warning)
        browser_health = self._read_cached_json(self.config.browser_health_file)
        browser_warning = self._snapshot_warning(
            "browser", browser_health, fresh_sec=120
        )
        if browser_warning is not None:
            warnings.append(browser_warning)
        browser_registry = self._read_cached_json(self.config.browser_registry_file)
        browser_registry_warning = self._snapshot_warning(
            "browser_registry", browser_registry, fresh_sec=120
        )
        if browser_registry_warning is not None:
            warnings.append(browser_registry_warning)
        if browser_health is not None and bool(
            browser_health.get("unhealthy_count", 0)
        ):
            warnings.append("browser unhealthy detected")
        gpt_status = load_gpt_status(self.config.gpt_status_file)
        gpt_warning = self._snapshot_warning("gpt", gpt_status, fresh_sec=120)
        if gpt_warning is not None:
            warnings.append(gpt_warning)
        gpt_endpoints = (
            []
            if gpt_status is None
            or not isinstance(gpt_status.get("endpoints", []), list)
            else cast(list[object], gpt_status.get("endpoints", []))
        )
        gpt_pending_boot = (
            0 if gpt_status is None else _to_int(gpt_status.get("pending_boot", 0))
        )
        gpt_idle_bootstrap = len(gpt_endpoints) == 0 and gpt_pending_boot == 0
        if (
            gpt_status is not None
            and bool(gpt_status.get("floor_breached", False))
            and not gpt_idle_bootstrap
        ):
            warnings.append("gpt floor breached")
        gpu_health = _read_json(self.config.lease_file)
        gpu_warning = self._snapshot_warning("gpu", gpu_health, fresh_sec=120)
        if gpu_warning is not None:
            warnings.append(gpu_warning)
        if gpu_health is not None and str(gpu_health.get("event", "")) in {
            "lock_busy",
            "renew_failed",
        }:
            warnings.append(str(gpu_health.get("event", "gpu warning")))
        result_payload = self._read_cached_json(self.config.result_router_file)
        result_warning = self._snapshot_warning("result", result_payload, fresh_sec=300)
        if result_warning is not None and result_payload is not None:
            warnings.append(result_warning)
        mismatch_warning = _worker_error_code_mismatch_warning(result_payload)
        if mismatch_warning:
            warnings.append(f"worker error mismatch: {mismatch_warning}")
        gui_payload = self._read_cached_json(self.config.gui_status_file)
        status_payload = (
            None
            if gui_payload is None
            else _coerce_mapping(gui_payload.get("status", {}))
        )
        if gui_payload is not None:
            invalid_reason = (
                ""
                if status_payload is None
                else str(status_payload.get("invalid_reason", "")).strip()
            )
            if invalid_reason:
                warnings.append(f"invalid contract: {invalid_reason}")
            worker_stage = str(gui_payload.get("worker_stage", "")).strip()
            if not worker_stage and status_payload is not None:
                worker_stage = str(status_payload.get("worker_stage", "")).strip()
            worker_error_code = str(gui_payload.get("worker_error_code", "")).strip()
            if not worker_error_code and status_payload is not None:
                worker_error_code = str(
                    status_payload.get("worker_error_code", "")
                ).strip()
            worker_error_display = _display_worker_error_code(worker_error_code)
            result_path = str(gui_payload.get("result_path", "")).strip()
            if not result_path and status_payload is not None:
                result_path = str(status_payload.get("result_path", "")).strip()
            if worker_error_code or worker_stage:
                warnings.append(
                    f"job failure: stage={worker_stage or '-'} error={worker_error_display or '-'} path={result_path or '-'}"
                )
        warning_text = " | ".join(warnings) if warnings else "경고 없음"
        self.warning_text.set(warning_text)
        if not warnings:
            self.home_warning_text.set("문제가 없습니다")
        elif any("login" in warning.lower() for warning in warnings):
            self.home_warning_text.set("로그인 확인이 필요합니다")
        elif any("gpt" in warning.lower() for warning in warnings):
            self.home_warning_text.set("GPT 준비 상태를 확인해주세요")
        else:
            self.home_warning_text.set(_truncate_ui_text(warning_text, 40))

    def on_close(self) -> None:
        self.stop_loop()
        self._save_settings()
        if self.refresh_after_id is not None:
            self.root.after_cancel(self.refresh_after_id)
        self.root.destroy()

    def run(self) -> None:
        self.root.mainloop()


def main() -> int:
    app = RuntimeV2ManagerGUI()
    app.run()
    return 0


def _lease_like_payload(payload: dict[str, object]) -> dict[str, object]:
    return {
        "key": str(payload.get("key", "")),
        "owner": str(payload.get("owner", "")),
        "token": _to_int(payload.get("token", 0)),
        "expires_at": _to_float(payload.get("expires_at", 0.0), 0.0),
        "run_id": str(payload.get("run_id", "unknown")),
        "pid": _to_int(payload.get("pid", 0)),
        "started_at": _to_float(payload.get("started_at", 0.0), 0.0),
        "host": str(payload.get("host", "unknown")),
    }


def _workload_from_lock_key(lock_key: str) -> GpuWorkload | None:
    normalized = lock_key.strip().lower()
    if normalized.endswith("qwen3_tts"):
        return "qwen3_tts"
    if normalized.endswith("rvc"):
        return "rvc"
    if normalized.endswith("kenburns"):
        return "kenburns"
    return None


def _program_row_for_workload(workload: str) -> str:
    normalized = workload.strip().lower()
    if normalized == "rvc":
        return "RVC"
    if normalized == "kenburns":
        return "KenBurns"
    return "QWEN3_TTS"


if __name__ == "__main__":
    raise SystemExit(main())
