from __future__ import annotations

import json
import io
import os
import sys
import tempfile
import unittest
from pathlib import Path
from typing import cast
from unittest.mock import MagicMock, patch

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from runtime_v2 import exit_codes
from runtime_v2.browser import manager as browser_manager
from runtime_v2.cli import (
    _build_runtime_config,
    _copy_legacy_sessions,
    _finalize_probe_result_from_progress,
    _run_stage5_row1_probe,
    _write_probe_progress,
    _write_probe_result,
    exit_code_from_readiness,
    exit_code_from_status,
    main,
    _spawn_detached_probe,
    _write_detached_summary,
    CliArgs,
)
from runtime_v2.config import RuntimeConfig, runtime_state_root
from runtime_v2.contracts.job_contract import JobContract
from runtime_v2.workers.job_runtime import prepare_workspace


class RuntimeV2ChatSafeExecutionTests(unittest.TestCase):
    def test_cli_soak_report_writes_markdown_report(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = RuntimeConfig.from_root(root / "runtime")
            config.soak_events_file.parent.mkdir(parents=True, exist_ok=True)
            _ = config.soak_events_file.write_text(
                '{"ts":1.0,"run_id":"run-1","mode":"once","status":"ok","code":"OK","exit_code":0,"debug_log":"logs/run-1.jsonl","result_path":"","gui_status_path":"","browser_health_path":"","gpu_health_path":"","gpt_status_path":"","control_plane_events_path":"","manifest_path":"","final_artifact_path":"artifact.mp4","summary":{}}\n',
                encoding="utf-8",
            )
            with (
                patch("runtime_v2.cli._build_runtime_config", return_value=config),
                patch("sys.argv", ["runtime_v2.cli", "--soak-report"]),
            ):
                exit_code = main()
                report_exists = config.soak_report_file.exists()

        self.assertEqual(exit_code, exit_codes.SUCCESS)
        self.assertTrue(report_exists)

    def test_exit_code_from_status_maps_stage6_gpu_blockers(self) -> None:
        self.assertEqual(
            exit_code_from_status("GPU_HEALTH_STALE"), exit_codes.LEASE_BUSY
        )
        self.assertEqual(
            exit_code_from_status("GPU_LEASE_RENEW_FAILED"), exit_codes.LEASE_BUSY
        )

    def test_exit_code_from_readiness_maps_worker_stall_blocker(self) -> None:
        readiness: dict[str, object] = {
            "ready": False,
            "blockers": [
                {
                    "axis": "worker_registry",
                    "code": "WORKER_STALL_DETECTED",
                    "reason": "stalled_workloads_present",
                }
            ],
        }

        self.assertEqual(exit_code_from_readiness(readiness), exit_codes.SELFTEST_FAIL)

    def test_prepare_workspace_defaults_to_runtime_config_artifact_root(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            artifact_root = Path(tmp_dir) / "external-artifacts"
            job = JobContract(
                job_id="chat-safe-workspace",
                workload="qwen3_tts",
                checkpoint_key="seed:chat-safe-workspace",
                payload={},
            )

            with patch(
                "runtime_v2.workers.job_runtime.RuntimeConfig"
            ) as runtime_config:
                runtime_config.return_value = RuntimeConfig(artifact_root=artifact_root)
                workspace = prepare_workspace(job)

        self.assertEqual(workspace, artifact_root / job.workload / job.job_id)
        self.assertNotIn(r"D:\YOUTUBEAUTO\system\runtime_v2\artifacts", str(workspace))

    def test_prepare_workspace_namespaces_run_id_under_job_directory(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            artifact_root = Path(tmp_dir) / "external-artifacts"
            job = JobContract(
                job_id="chatgpt-sheet1-15",
                workload="chatgpt",
                checkpoint_key="seed:chatgpt-sheet1-15",
                payload={"run_id": "run-123"},
            )

            workspace = prepare_workspace(job, artifact_root=artifact_root)
            started_payload = json.loads(
                (workspace / "started.json").read_text(encoding="utf-8")
            )

        self.assertEqual(
            workspace,
            artifact_root / job.workload / job.job_id / "run-123",
        )
        self.assertEqual(started_payload["run_id"], "run-123")

    def test_default_runtime_config_uses_external_runtime_state_root(self) -> None:
        config = RuntimeConfig()
        state_root = runtime_state_root()

        self.assertEqual(
            config.queue_store_file, state_root / "state" / "job_queue.json"
        )
        self.assertEqual(
            config.gui_status_file, state_root / "health" / "gui_status.json"
        )
        self.assertEqual(config.artifact_root, state_root / "artifacts")
        self.assertEqual(config.debug_log_root, state_root / "logs")
        self.assertNotIn(
            r"D:\YOUTUBEAUTO\system\runtime_v2", str(config.queue_store_file)
        )

    def test_runtime_config_from_root_preserves_explicit_root_layout(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir) / "runtime"
            config = RuntimeConfig.from_root(root)

        self.assertEqual(
            config.queue_store_file, root.resolve() / "state" / "job_queue.json"
        )
        self.assertEqual(config.artifact_root, root.resolve() / "artifacts")
        self.assertEqual(config.debug_log_root, root.resolve() / "logs")

    def test_copy_legacy_sessions_copies_missing_directories(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            legacy_root = root / "legacy"
            external_root = root / "external"
            (legacy_root / "chatgpt-primary").mkdir(parents=True)
            (legacy_root / "chatgpt-primary" / "marker.txt").write_text(
                "ok", encoding="utf-8"
            )
            (external_root / "seaart-primary").mkdir(parents=True)

            report = _copy_legacy_sessions(legacy_root, external_root)

        self.assertTrue(bool(report["ok"]))
        self.assertEqual(report["migrated"], ["chatgpt-primary"])
        self.assertEqual(report["skipped_existing"], [])

    def test_write_detached_summary_writes_contract_fields(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            out_root = Path(tmp_dir) / "probe"
            summary_file = _write_detached_summary(
                out_root=out_root,
                kind="pytest",
                target="tests/test_runtime_v2_browser_plane.py",
                exit_code=0,
                payload={"status": "ok"},
            )
            payload = json.loads(summary_file.read_text(encoding="utf-8"))

        self.assertEqual(payload["kind"], "pytest")
        self.assertEqual(payload["target"], "tests/test_runtime_v2_browser_plane.py")
        self.assertEqual(payload["exit_code"], 0)
        self.assertEqual(payload["status"], "ok")
        self.assertEqual(payload["out_root"], str(out_root))

    def test_write_detached_summary_retries_winerror_5(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            original_replace = Path.replace
            calls = {"count": 0}

            def flaky_replace(self: Path, target: Path) -> Path:
                if self.suffix == ".tmp" and calls["count"] < 2:
                    calls["count"] += 1
                    error = PermissionError("locked")
                    error.winerror = 5
                    raise error
                return original_replace(self, target)

            with (
                patch("runtime_v2.cli.sleep", return_value=None),
                patch.object(Path, "replace", new=flaky_replace),
            ):
                summary_file = _write_detached_summary(
                    out_root=root / "probe",
                    kind="pytest",
                    target="tests/test_runtime_v2_chat_safe_execution.py",
                    exit_code=0,
                    payload={"status": "ok"},
                )

            payload = json.loads(summary_file.read_text(encoding="utf-8"))
        self.assertEqual(payload["status"], "ok")

    def test_stage5_probe_result_write_retries_winerror_5(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            original_replace = Path.replace
            calls = {"count": 0}

            def flaky_replace(self: Path, target: Path) -> Path:
                if self.suffix == ".tmp" and calls["count"] < 2:
                    calls["count"] += 1
                    error = PermissionError("locked")
                    error.winerror = 5
                    raise error
                return original_replace(self, target)

            with (
                patch("runtime_v2.cli.sleep", return_value=None),
                patch.object(Path, "replace", new=flaky_replace),
            ):
                result_path = _write_probe_result(
                    root / "probe",
                    {"status": "failed", "code": "BATCH_TIMEOUT"},
                )

            payload = json.loads(result_path.read_text(encoding="utf-8"))
        self.assertEqual(payload["code"], "BATCH_TIMEOUT")

    def test_spawn_detached_probe_writes_spawn_summary(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            args = CliArgs()
            args.probe_root = str(root / "probe")
            popen_result = MagicMock()
            popen_result.pid = 24680

            with patch("runtime_v2.cli.subprocess.Popen", return_value=popen_result):
                exit_code = _spawn_detached_probe(args, mode="selftest")

            payload = json.loads(
                (root / "probe" / "summary.json").read_text(encoding="utf-8")
            )

        self.assertEqual(exit_code, exit_codes.SUCCESS)
        self.assertEqual(payload["status"], "spawned")
        self.assertEqual(payload["kind"], "selftest")
        self.assertEqual(payload["exit_code"], exit_codes.SUCCESS)

    def test_spawn_detached_probe_for_stage5_forwards_excel_arguments(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            args = CliArgs()
            args.probe_root = str(root / "probe")
            args.excel_path = str(root / "topic.xlsx")
            args.sheet_name = "Sheet1"
            args.row_index = 2
            args.max_control_ticks = 7
            popen_result = MagicMock()
            popen_result.pid = 13579

            with patch(
                "runtime_v2.cli.subprocess.Popen", return_value=popen_result
            ) as popen:
                exit_code = _spawn_detached_probe(args, mode="stage5_row1")

        self.assertEqual(exit_code, exit_codes.SUCCESS)
        command = popen.call_args_list[0].args[0]
        self.assertIn("--stage5-row1-probe-child", command)
        self.assertIn("--excel-path", command)
        self.assertIn(str(root / "topic.xlsx"), command)
        self.assertIn("--row-index", command)
        self.assertIn("2", command)
        self.assertIn("--max-control-ticks", command)
        self.assertIn("7", command)

    def test_spawn_detached_probe_for_stage5_starts_finalize_watcher(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            args = CliArgs()
            args.probe_root = str(root / "probe")
            args.excel_path = str(root / "topic.xlsx")
            args.sheet_name = "Sheet1"
            args.row_index = 2
            args.max_control_ticks = 7
            probe_child = MagicMock()
            probe_child.pid = 13579
            watcher_child = MagicMock()
            watcher_child.pid = 24680

            with patch(
                "runtime_v2.cli.subprocess.Popen",
                side_effect=[probe_child, watcher_child],
            ) as popen:
                exit_code = _spawn_detached_probe(args, mode="stage5_row1")

        self.assertEqual(exit_code, exit_codes.SUCCESS)
        self.assertEqual(popen.call_count, 2)
        watcher_command = popen.call_args_list[1].args[0]
        self.assertEqual(
            watcher_command[:4], [sys.executable, "-u", "-m", "runtime_v2.cli"]
        )
        self.assertIn("--probe-finalize-child", watcher_command)
        self.assertIn("--child-pid", watcher_command)
        self.assertIn("13579", watcher_command)

    def test_probe_finalize_child_writes_probe_result_when_pid_dead(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            _write_probe_progress(
                probe_root,
                {
                    "run_id": "row15-run",
                    "mode": "stage5_row1",
                    "status": "running",
                    "code": "PROBE_RUNNING",
                    "probe_success": False,
                    "ticks": 4,
                },
            )
            args = CliArgs()
            args.probe_finalize_child = True
            args.probe_root = str(probe_root)
            args.child_pid = 0
            args.fallback_code = "PROBE_INCOMPLETE"
            args.fallback_status = "failed"

            with patch(
                "sys.argv",
                [
                    "runtime_v2.cli",
                    "--probe-finalize-child",
                    "--probe-root",
                    str(probe_root),
                    "--child-pid",
                    "0",
                    "--fallback-code",
                    "PROBE_INCOMPLETE",
                    "--fallback-status",
                    "failed",
                ],
            ):
                exit_code = main()
            probe_result = json.loads(
                (probe_root / "probe_result.json").read_text(encoding="utf-8")
            )

        self.assertEqual(exit_code, exit_codes.SUCCESS)
        self.assertEqual(probe_result["status"], "failed")
        self.assertEqual(probe_result["code"], "PROBE_INCOMPLETE")

    def test_probe_finalize_child_waits_until_pid_exits(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            _write_probe_progress(
                probe_root,
                {
                    "run_id": "row15-run",
                    "mode": "stage5_row1",
                    "status": "running",
                    "code": "PROBE_RUNNING",
                    "probe_success": False,
                    "ticks": 5,
                },
            )
            args = CliArgs()
            args.probe_finalize_child = True
            args.probe_root = str(probe_root)
            args.child_pid = 13579
            args.fallback_code = "PROBE_INCOMPLETE"
            args.fallback_status = "failed"

            with (
                patch("runtime_v2.cli._pid_is_alive", side_effect=[True, True, False]),
                patch("runtime_v2.cli.sleep", return_value=None) as sleep_mock,
                patch(
                    "sys.argv",
                    [
                        "runtime_v2.cli",
                        "--probe-finalize-child",
                        "--probe-root",
                        str(probe_root),
                        "--child-pid",
                        "13579",
                        "--fallback-code",
                        "PROBE_INCOMPLETE",
                        "--fallback-status",
                        "failed",
                    ],
                ),
            ):
                exit_code = main()

            probe_result = json.loads(
                (probe_root / "probe_result.json").read_text(encoding="utf-8")
            )

        self.assertEqual(exit_code, exit_codes.SUCCESS)
        self.assertEqual(sleep_mock.call_count, 2)
        self.assertEqual(probe_result["status"], "failed")
        self.assertEqual(probe_result["code"], "PROBE_INCOMPLETE")

    def test_job_contract_cli_surfaces_failed_worker_status(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = RuntimeConfig.from_root(root / "runtime")
            contract_path = root / "canva.job.json"
            contract_path.write_text("{}", encoding="utf-8")
            explicit_job = JobContract(
                job_id="canva-boundary-job",
                workload="canva",
                checkpoint_key="stage2:canva:Sheet1!row15:3",
                payload={"run_id": "canva-boundary-run", "row_ref": "Sheet1!row15"},
            )
            stdout_buffer = io.StringIO()

            with (
                patch("runtime_v2.cli._build_runtime_config", return_value=config),
                patch("runtime_v2.cli._load_job_contract", return_value=explicit_job),
                patch(
                    "runtime_v2.cli._run_explicit_job_contract",
                    return_value={
                        "status": "ok",
                        "code": "OK",
                        "job": explicit_job.to_dict(),
                        "worker_result": {
                            "status": "failed",
                            "stage": "canva_adapter",
                            "error_code": "CANVA_PRODUCT_BACKGROUND_CREDIT_EXHAUSTED",
                            "completion": {
                                "state": "failed",
                                "final_output": False,
                            },
                        },
                    },
                ),
                patch(
                    "sys.argv",
                    [
                        "runtime_v2.cli",
                        "--job-contract-path",
                        str(contract_path),
                    ],
                ),
                patch("sys.stdout", stdout_buffer),
            ):
                exit_code = main()

            final_payload = json.loads(
                stdout_buffer.getvalue().strip().splitlines()[-1]
            )

        self.assertEqual(exit_code, exit_codes.CLI_USAGE)
        self.assertEqual(final_payload["status"], "failed")
        self.assertEqual(
            final_payload["code"], "CANVA_PRODUCT_BACKGROUND_CREDIT_EXHAUSTED"
        )

    def test_job_contract_cli_report_uses_failed_summary_status(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = RuntimeConfig.from_root(root / "runtime")
            contract_path = root / "canva.job.json"
            contract_path.write_text("{}", encoding="utf-8")
            explicit_job = JobContract(
                job_id="canva-boundary-job",
                workload="canva",
                checkpoint_key="stage2:canva:Sheet1!row15:3",
                payload={"run_id": "canva-boundary-run", "row_ref": "Sheet1!row15"},
            )

            with (
                patch("runtime_v2.cli._build_runtime_config", return_value=config),
                patch("runtime_v2.cli._load_job_contract", return_value=explicit_job),
                patch(
                    "runtime_v2.cli._run_explicit_job_contract",
                    return_value={
                        "status": "ok",
                        "code": "OK",
                        "job": explicit_job.to_dict(),
                        "worker_result": {
                            "status": "failed",
                            "stage": "canva_adapter",
                            "error_code": "CANVA_PRODUCT_BACKGROUND_CREDIT_EXHAUSTED",
                            "completion": {
                                "state": "failed",
                                "final_output": False,
                            },
                        },
                    },
                ),
                patch(
                    "sys.argv",
                    [
                        "runtime_v2.cli",
                        "--job-contract-path",
                        str(contract_path),
                    ],
                ),
            ):
                exit_code = main()

            debug_events = [
                json.loads(line)
                for line in config.debug_log_root.joinpath("canva-boundary-run.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
                if line.strip()
            ]
            final_event = debug_events[-1]

        self.assertEqual(exit_code, exit_codes.CLI_USAGE)
        self.assertEqual(final_event["event"], "run_finished")
        self.assertEqual(final_event["status"], "failed")
        self.assertEqual(
            final_event["code"], "CANVA_PRODUCT_BACKGROUND_CREDIT_EXHAUSTED"
        )

    def test_job_contract_cli_result_event_uses_failed_summary_status(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = RuntimeConfig.from_root(root / "runtime")
            contract_path = root / "canva.job.json"
            contract_path.write_text("{}", encoding="utf-8")
            explicit_job = JobContract(
                job_id="canva-boundary-job",
                workload="canva",
                checkpoint_key="stage2:canva:Sheet1!row15:3",
                payload={"run_id": "canva-boundary-run", "row_ref": "Sheet1!row15"},
            )

            with (
                patch("runtime_v2.cli._build_runtime_config", return_value=config),
                patch("runtime_v2.cli._load_job_contract", return_value=explicit_job),
                patch(
                    "runtime_v2.cli._run_explicit_job_contract",
                    return_value={
                        "status": "ok",
                        "code": "OK",
                        "job": explicit_job.to_dict(),
                        "worker_result": {
                            "status": "failed",
                            "stage": "canva_adapter",
                            "error_code": "CANVA_PRODUCT_BACKGROUND_CREDIT_EXHAUSTED",
                            "completion": {
                                "state": "failed",
                                "final_output": False,
                            },
                        },
                    },
                ),
                patch(
                    "sys.argv",
                    [
                        "runtime_v2.cli",
                        "--job-contract-path",
                        str(contract_path),
                    ],
                ),
            ):
                exit_code = main()

            debug_events = [
                json.loads(line)
                for line in config.debug_log_root.joinpath("canva-boundary-run.jsonl")
                .read_text(encoding="utf-8")
                .splitlines()
                if line.strip()
            ]
            cli_result_event = next(
                event for event in debug_events if event.get("event") == "cli_result"
            )

        self.assertEqual(exit_code, exit_codes.CLI_USAGE)
        self.assertEqual(cli_result_event["status"], "failed")
        self.assertEqual(
            cli_result_event["code"], "CANVA_PRODUCT_BACKGROUND_CREDIT_EXHAUSTED"
        )

    def test_build_runtime_config_uses_probe_root_for_stage5_child(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            probe_root = Path(tmp_dir) / "probe"
            args = CliArgs()
            args.probe_root = str(probe_root)
            args.stage5_row1_probe_child = True

            config = _build_runtime_config(args)

        self.assertEqual(
            config.result_router_file, probe_root / "evidence" / "result.json"
        )

    def test_stage5_probe_can_seed_ok_status_row_for_closeout(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()

            config = RuntimeConfig.from_root(root / "runtime")
            probe_root = root / "probe"
            final_artifact = root / "render_final.mp4"
            final_artifact.write_bytes(b"mp4")

            def fake_control_loop_once(
                *, owner: str, config: RuntimeConfig, run_id: str
            ) -> dict[str, object]:
                _ = owner
                self.assertEqual(config.stable_file_age_sec, 0)
                config.result_router_file.parent.mkdir(parents=True, exist_ok=True)
                config.result_router_file.write_text(
                    json.dumps(
                        {
                            "metadata": {
                                "run_id": run_id,
                                "workload": "render",
                                "final_output": True,
                                "final_artifact_path": str(final_artifact.resolve()),
                            }
                        },
                        ensure_ascii=True,
                    ),
                    encoding="utf-8",
                )
                return {"status": "ok", "code": "OK"}

            with (
                patch(
                    "runtime_v2.cli.run_control_loop_once",
                    side_effect=fake_control_loop_once,
                ),
                patch(
                    "runtime_v2.cli.load_runtime_readiness",
                    return_value={"ready": True, "code": "OK", "blockers": []},
                ),
            ):
                report = _run_stage5_row1_probe(
                    owner="runtime_v2",
                    config=config,
                    probe_root=probe_root,
                    run_id="stage5-ok-closeout",
                    excel_path=str(excel_path),
                    sheet_name="Sheet1",
                    row_index=0,
                    max_control_ticks=2,
                )

        self.assertTrue(bool(report["probe_success"]))
        self.assertEqual(report["code"], "OK")

    def test_stage5_probe_continues_when_control_result_requests_retry(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()
            config = RuntimeConfig.from_root(root / "runtime")
            config.result_router_file.parent.mkdir(parents=True, exist_ok=True)
            call_counter = {"count": 0}

            def fake_control_loop_once(
                *, owner: str, config: RuntimeConfig, run_id: str
            ):
                _ = (owner, config, run_id)
                call_counter["count"] += 1
                if call_counter["count"] == 1:
                    return {
                        "status": "failed",
                        "code": "BROWSER_UNHEALTHY",
                        "queue_status": "retry",
                    }
                render_payload = {
                    "metadata": {
                        "run_id": "stage5-retry-closeout",
                        "workload": "render",
                        "final_output": True,
                        "final_artifact_path": str((root / "final.mp4").resolve()),
                    }
                }
                _ = config.result_router_file.write_text(
                    json.dumps(render_payload, ensure_ascii=True),
                    encoding="utf-8",
                )
                return {"status": "ok", "code": "OK", "queue_status": "completed"}

            with (
                patch(
                    "runtime_v2.cli.run_control_loop_once",
                    side_effect=fake_control_loop_once,
                ),
                patch(
                    "runtime_v2.cli.load_runtime_readiness",
                    return_value={"ready": True, "code": "OK", "blockers": []},
                ),
            ):
                report = _run_stage5_row1_probe(
                    owner="runtime_v2",
                    config=config,
                    probe_root=probe_root,
                    run_id="stage5-retry-closeout",
                    excel_path=str(excel_path),
                    sheet_name="Sheet1",
                    row_index=0,
                    max_control_ticks=3,
                )

            progress_payload = json.loads(
                (probe_root / "probe_progress.json").read_text(encoding="utf-8")
            )
            probe_payload = json.loads(
                (probe_root / "probe_result.json").read_text(encoding="utf-8")
            )

        self.assertTrue(bool(report["probe_success"]))
        self.assertEqual(report["code"], "OK")
        self.assertEqual(call_counter["count"], 2)
        self.assertEqual(progress_payload["mode"], "stage5_row1")
        self.assertEqual(progress_payload["status"], "running")
        self.assertEqual(progress_payload["code"], "PROBE_RUNNING")
        self.assertEqual(progress_payload["ticks"], 2)
        self.assertEqual(probe_payload["status"], "ok")
        self.assertEqual(probe_payload["code"], "OK")

    def test_write_probe_progress_retries_permission_error_on_replace(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            probe_root = Path(tmp_dir) / "probe"
            original_replace = Path.replace
            call_count = {"count": 0}

            def flaky_replace(self: Path, target: Path) -> Path:
                call_count["count"] += 1
                if call_count["count"] < 3:
                    error = PermissionError("locked")
                    error.winerror = 5
                    raise error
                return original_replace(self, target)

            with patch.object(Path, "replace", new=flaky_replace):
                output_path = _write_probe_progress(
                    probe_root,
                    {"status": "running", "code": "PROBE_RUNNING", "ticks": 1},
                )

        self.assertEqual(output_path, probe_root / "probe_progress.json")
        self.assertEqual(call_count["count"], 3)

    def test_write_probe_progress_falls_back_after_winerror_5_retry_exhausted(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            probe_root = Path(tmp_dir) / "probe"

            def always_locked_replace(self: Path, target: Path) -> Path:
                if self.suffix == ".tmp":
                    error = PermissionError("locked")
                    error.winerror = 5
                    raise error
                raise AssertionError("unexpected replace target")

            with (
                patch("runtime_v2.cli.sleep", return_value=None),
                patch.object(Path, "replace", new=always_locked_replace),
            ):
                output_path = _write_probe_progress(
                    probe_root,
                    {"status": "running", "code": "PROBE_RUNNING", "ticks": 1},
                )
            payload = json.loads(output_path.read_text(encoding="utf-8"))

        self.assertEqual(output_path, probe_root / "probe_progress.json")
        self.assertEqual(payload["schema_version"], "1.0")
        self.assertEqual(payload["runtime"], "runtime_v2")
        self.assertEqual(payload["status"], "running")
        self.assertEqual(payload["code"], "PROBE_RUNNING")

    def test_write_probe_result_falls_back_after_winerror_5_retry_exhausted(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            probe_root = Path(tmp_dir) / "probe"

            def always_locked_replace(self: Path, target: Path) -> Path:
                if self.suffix == ".tmp":
                    error = PermissionError("locked")
                    error.winerror = 5
                    raise error
                raise AssertionError("unexpected replace target")

            with (
                patch("runtime_v2.cli.sleep", return_value=None),
                patch.object(Path, "replace", new=always_locked_replace),
            ):
                output_path = _write_probe_result(
                    probe_root,
                    {"status": "failed", "code": "BATCH_TIMEOUT"},
                )
            payload = json.loads(output_path.read_text(encoding="utf-8"))

        self.assertEqual(output_path, probe_root / "probe_result.json")
        self.assertEqual(payload["schema_version"], "1.0")
        self.assertEqual(payload["runtime"], "runtime_v2")
        self.assertEqual(payload["status"], "failed")
        self.assertEqual(payload["code"], "BATCH_TIMEOUT")

    def test_finalize_probe_result_from_progress_writes_fail_closed_artifact(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            probe_root = Path(tmp_dir) / "probe"
            probe_root.mkdir(parents=True, exist_ok=True)
            progress = _write_probe_progress(
                probe_root,
                {
                    "run_id": "row15-run",
                    "mode": "stage5_row1",
                    "status": "ok",
                    "code": "OK",
                    "probe_success": False,
                    "ticks": 8,
                    "latest_result": {
                        "job": {"workload": "seaart"},
                        "code": "OK",
                    },
                },
            )

            output_path = _finalize_probe_result_from_progress(
                probe_root,
                fallback_code="PROBE_INCOMPLETE",
                fallback_status="failed",
            )
            payload = json.loads(output_path.read_text(encoding="utf-8"))

        self.assertEqual(progress, probe_root / "probe_progress.json")
        self.assertEqual(output_path, probe_root / "probe_result.json")
        self.assertEqual(payload["status"], "failed")
        self.assertEqual(payload["code"], "PROBE_INCOMPLETE")
        self.assertEqual(payload["ticks"], 8)
        self.assertEqual(payload["latest_result"]["job"]["workload"], "seaart")

    def test_finalize_probe_result_from_progress_overwrites_running_placeholder(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            probe_root = Path(tmp_dir) / "probe"
            probe_root.mkdir(parents=True, exist_ok=True)
            _write_probe_result(
                probe_root,
                {
                    "run_id": "row15-run",
                    "mode": "stage5_row1",
                    "status": "running",
                    "code": "PROBE_RUNNING",
                    "probe_success": False,
                    "ticks": 1,
                },
            )
            _write_probe_progress(
                probe_root,
                {
                    "run_id": "row15-run",
                    "mode": "stage5_row1",
                    "status": "running",
                    "code": "PROBE_RUNNING",
                    "probe_success": False,
                    "ticks": 8,
                    "latest_result": {
                        "job": {"workload": "seaart"},
                        "code": "OK",
                    },
                },
            )

            output_path = _finalize_probe_result_from_progress(
                probe_root,
                fallback_code="PROBE_INCOMPLETE",
                fallback_status="failed",
            )
            payload = json.loads(output_path.read_text(encoding="utf-8"))

        self.assertEqual(output_path, probe_root / "probe_result.json")
        self.assertEqual(payload["status"], "failed")
        self.assertEqual(payload["code"], "PROBE_INCOMPLETE")
        self.assertEqual(payload["ticks"], 8)

    def test_stage5_probe_uses_probe_local_runtime_root_for_control_loop(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()
            config = RuntimeConfig.from_root(root / "runtime")

            def fake_control_loop_once(
                *, owner: str, config: RuntimeConfig, run_id: str
            ):
                _ = owner
                _ = run_id
                self.assertEqual(
                    config.queue_store_file, probe_root / "state" / "job_queue.json"
                )
                return {"status": "failed", "code": "BROWSER_UNHEALTHY"}

            with patch(
                "runtime_v2.cli.run_control_loop_once",
                side_effect=fake_control_loop_once,
            ):
                _run_stage5_row1_probe(
                    owner="runtime_v2",
                    config=config,
                    probe_root=probe_root,
                    run_id="stage5-probe-local-root",
                    excel_path=str(excel_path),
                    sheet_name="Sheet1",
                    row_index=0,
                    max_control_ticks=1,
                )

    def test_stage5_probe_immediately_follows_up_after_seeded_result(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()
            config = RuntimeConfig.from_root(root / "runtime")
            config.result_router_file.parent.mkdir(parents=True, exist_ok=True)

            call_counter = {"count": 0}

            def fake_control_loop_once(
                *, owner: str, config: RuntimeConfig, run_id: str
            ):
                _ = (owner, config, run_id)
                call_counter["count"] += 1
                if call_counter["count"] == 1:
                    return {
                        "status": "seeded",
                        "code": "SEEDED_JOB",
                        "queue_status": "seeded",
                    }
                render_payload = {
                    "metadata": {
                        "run_id": "stage5-seeded-followup",
                        "workload": "render",
                        "final_output": True,
                        "final_artifact_path": str((root / "final.mp4").resolve()),
                    }
                }
                _ = config.result_router_file.write_text(
                    json.dumps(render_payload, ensure_ascii=True),
                    encoding="utf-8",
                )
                return {"status": "ok", "code": "OK", "queue_status": "completed"}

            with (
                patch(
                    "runtime_v2.cli.run_control_loop_once",
                    side_effect=fake_control_loop_once,
                ),
                patch(
                    "runtime_v2.cli.load_runtime_readiness",
                    return_value={"ready": True, "code": "OK", "blockers": []},
                ),
            ):
                report = _run_stage5_row1_probe(
                    owner="runtime_v2",
                    config=config,
                    probe_root=probe_root,
                    run_id="stage5-seeded-followup",
                    excel_path=str(excel_path),
                    sheet_name="Sheet1",
                    row_index=0,
                    max_control_ticks=3,
                )

        self.assertTrue(bool(report["probe_success"]))
        self.assertEqual(report["code"], "OK")
        self.assertEqual(call_counter["count"], 2)

    def test_stage5_probe_updates_latest_after_each_tick(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()
            config = RuntimeConfig.from_root(root / "runtime")
            config.result_router_file.parent.mkdir(parents=True, exist_ok=True)

            call_counter = {"count": 0}

            def fake_control_loop_once(
                *, owner: str, config: RuntimeConfig, run_id: str
            ):
                _ = (owner, config, run_id)
                call_counter["count"] += 1
                if call_counter["count"] == 1:
                    return {
                        "status": "seeded",
                        "code": "SEEDED_JOB",
                        "queue_status": "seeded",
                        "debug_log": str((probe_root / "seeded.jsonl").resolve()),
                    }
                render_payload = {
                    "metadata": {
                        "run_id": "stage5-latest-progress",
                        "workload": "render",
                        "final_output": True,
                        "final_artifact_path": str((root / "final.mp4").resolve()),
                    }
                }
                _ = config.result_router_file.write_text(
                    json.dumps(render_payload, ensure_ascii=True),
                    encoding="utf-8",
                )
                return {
                    "status": "ok",
                    "code": "OK",
                    "queue_status": "completed",
                    "workload": "render",
                    "debug_log": str((probe_root / "ok.jsonl").resolve()),
                }

            with (
                patch(
                    "runtime_v2.cli.run_control_loop_once",
                    side_effect=fake_control_loop_once,
                ),
                patch(
                    "runtime_v2.cli.load_runtime_readiness",
                    return_value={"ready": True, "code": "OK", "blockers": []},
                ),
            ):
                report = _run_stage5_row1_probe(
                    owner="runtime_v2",
                    config=config,
                    probe_root=probe_root,
                    run_id="stage5-latest-progress",
                    excel_path=str(excel_path),
                    sheet_name="Sheet1",
                    row_index=0,
                    max_control_ticks=3,
                )

            latest_probe = json.loads(
                (probe_root / "probe_result.json").read_text(encoding="utf-8")
            )

        self.assertEqual(report["code"], "OK")
        self.assertEqual(latest_probe["code"], "OK")

    def test_stage5_probe_writes_progress_probe_result_before_terminal_state(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()
            config = RuntimeConfig.from_root(root / "runtime")

            call_counter = {"count": 0}

            def fake_control_loop_once(
                *, owner: str, config: RuntimeConfig, run_id: str
            ):
                _ = (owner, config, run_id)
                call_counter["count"] += 1
                if call_counter["count"] == 1:
                    return {
                        "status": "seeded",
                        "code": "SEEDED_JOB",
                        "queue_status": "seeded",
                    }
                return {
                    "status": "failed",
                    "code": "missing_scene_prompts",
                    "queue_status": "failed",
                }

            with patch(
                "runtime_v2.cli.run_control_loop_once",
                side_effect=fake_control_loop_once,
            ):
                report = _run_stage5_row1_probe(
                    owner="runtime_v2",
                    config=config,
                    probe_root=probe_root,
                    run_id="stage5-progress-probe",
                    excel_path=str(excel_path),
                    sheet_name="Sheet1",
                    row_index=0,
                    max_control_ticks=2,
                )

            probe_payload = json.loads(
                (probe_root / "probe_result.json").read_text(encoding="utf-8")
            )

        self.assertEqual(report["status"], "failed")
        self.assertEqual(probe_payload["mode"], "stage5_row1")
        self.assertEqual(probe_payload["code"], "missing_scene_prompts")
        self.assertEqual(probe_payload["ticks"], 2)
        self.assertEqual(
            probe_payload["control_results"][-1]["code"], "missing_scene_prompts"
        )

    def test_stage5_probe_writes_failure_summary_for_terminal_failure(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()
            config = RuntimeConfig.from_root(root / "runtime")

            with patch(
                "runtime_v2.cli.run_control_loop_once",
                return_value={
                    "status": "failed",
                    "code": "missing_scene_prompts",
                    "queue_status": "failed",
                },
            ):
                report = _run_stage5_row1_probe(
                    owner="runtime_v2",
                    config=config,
                    probe_root=probe_root,
                    run_id="stage5-terminal-failure",
                    excel_path=str(excel_path),
                    sheet_name="Sheet1",
                    row_index=0,
                    max_control_ticks=1,
                )

            failure_summary_path = str(report.get("failure_summary_path", "")).strip()
            failure_summary_exists = Path(failure_summary_path).exists()

        self.assertEqual(report["status"], "failed")
        self.assertEqual(report["code"], "missing_scene_prompts")
        self.assertTrue(bool(failure_summary_path))
        self.assertTrue(failure_summary_exists)

    def test_stage5_probe_child_writes_probe_result_file(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()

            config = RuntimeConfig.from_root(root / "runtime")

            with (
                patch("runtime_v2.cli._build_runtime_config", return_value=config),
                patch(
                    "runtime_v2.cli._run_stage5_row1_probe",
                    return_value={"status": "failed", "code": "BATCH_TIMEOUT"},
                ),
                patch(
                    "sys.argv",
                    [
                        "runtime_v2.cli",
                        "--stage5-row1-probe-child",
                        "--owner",
                        "runtime_v2",
                        "--probe-root",
                        str(probe_root),
                        "--excel-path",
                        str(excel_path),
                        "--sheet-name",
                        "Sheet1",
                        "--row-index",
                        "0",
                        "--max-control-ticks",
                        "1",
                    ],
                ),
            ):
                exit_code = main()

            probe_result = json.loads(
                (probe_root / "probe_result.json").read_text(encoding="utf-8")
            )

        self.assertEqual(exit_code, exit_codes.CLI_USAGE)
        self.assertEqual(probe_result["status"], "failed")
        self.assertEqual(probe_result["code"], "BATCH_TIMEOUT")

    def test_stage5_probe_child_preserves_probe_success_fields(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()

            config = RuntimeConfig.from_root(root / "runtime")

            with (
                patch("runtime_v2.cli._build_runtime_config", return_value=config),
                patch(
                    "runtime_v2.cli._run_stage5_row1_probe",
                    return_value={
                        "status": "ok",
                        "code": "OK",
                        "probe_success": False,
                        "ticks": 3,
                        "control_results": [
                            {"status": "ok", "code": "OK", "workload": "chatgpt"},
                            {
                                "status": "ok",
                                "code": "OK",
                                "workload": "qwen3_tts",
                            },
                        ],
                        "latest_result": {
                            "status": "ok",
                            "code": "OK",
                            "job": {"workload": "qwen3_tts"},
                        },
                    },
                ),
                patch(
                    "sys.argv",
                    [
                        "runtime_v2.cli",
                        "--stage5-row1-probe-child",
                        "--owner",
                        "runtime_v2",
                        "--probe-root",
                        str(probe_root),
                        "--excel-path",
                        str(excel_path),
                        "--sheet-name",
                        "Sheet1",
                        "--row-index",
                        "0",
                        "--max-control-ticks",
                        "1",
                    ],
                ),
            ):
                exit_code = main()

            probe_result = json.loads(
                (probe_root / "probe_result.json").read_text(encoding="utf-8")
            )

        self.assertEqual(exit_code, exit_codes.SUCCESS)
        self.assertEqual(probe_result["status"], "ok")
        self.assertFalse(bool(probe_result["probe_success"]))
        self.assertEqual(probe_result["ticks"], 3)
        self.assertEqual(probe_result["latest_result"]["job"]["workload"], "qwen3_tts")

    def test_stage5_probe_child_writes_probe_result_on_unhandled_exception(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            probe_root = root / "probe"
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(["Topic", "Status"])
            sheet.append(["Semantic row", "OK"])
            workbook.save(excel_path)
            workbook.close()

            config = RuntimeConfig.from_root(root / "runtime")

            with (
                patch("runtime_v2.cli._build_runtime_config", return_value=config),
                patch(
                    "runtime_v2.cli._run_stage5_row1_probe",
                    side_effect=RuntimeError("chatgpt timeout exploded"),
                ),
                patch(
                    "sys.argv",
                    [
                        "runtime_v2.cli",
                        "--stage5-row1-probe-child",
                        "--owner",
                        "runtime_v2",
                        "--probe-root",
                        str(probe_root),
                        "--excel-path",
                        str(excel_path),
                        "--sheet-name",
                        "Sheet1",
                        "--row-index",
                        "0",
                        "--max-control-ticks",
                        "1",
                    ],
                ),
            ):
                exit_code = main()

            probe_result = json.loads(
                (probe_root / "probe_result.json").read_text(encoding="utf-8")
            )
            detached_summary = json.loads(
                (probe_root / "summary.json").read_text(encoding="utf-8")
            )

        self.assertEqual(exit_code, exit_codes.CLI_USAGE)
        self.assertEqual(probe_result["status"], "failed")
        self.assertEqual(probe_result["code"], "UNHANDLED_EXCEPTION")
        self.assertEqual(detached_summary["status"], "failed")
        self.assertEqual(detached_summary["code"], "UNHANDLED_EXCEPTION")

    def test_spawn_detached_probe_for_stage5b_forwards_batch_arguments(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            args = CliArgs()
            args.probe_root = str(root / "probe")
            args.excel_path = str(root / "topic.xlsx")
            args.sheet_name = "Sheet1"
            args.batch_count = 5
            args.max_control_ticks = 9
            popen_result = MagicMock()
            popen_result.pid = 24681

            with patch(
                "runtime_v2.cli.subprocess.Popen", return_value=popen_result
            ) as popen:
                exit_code = _spawn_detached_probe(args, mode="stage5b_5row")

        self.assertEqual(exit_code, exit_codes.SUCCESS)
        command = popen.call_args.args[0]
        self.assertIn("--stage5b-5row-probe-child", command)
        self.assertIn("--excel-path", command)
        self.assertIn(str(root / "topic.xlsx"), command)
        self.assertIn("--batch-count", command)
        self.assertIn("5", command)
        self.assertIn("--max-control-ticks", command)
        self.assertIn("9", command)

    def test_default_session_profile_dir_uses_legacy_only_when_explicitly_allowed(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            legacy_root = root / "legacy"
            external_root = root / "external"
            (legacy_root / "chatgpt-primary").mkdir(parents=True)

            with (
                patch.object(browser_manager, "LEGACY_SESSION_ROOT", legacy_root),
                patch(
                    "runtime_v2.browser.manager.browser_session_root",
                    return_value=external_root,
                ),
                patch.dict(os.environ, {}, clear=False),
            ):
                blocked = browser_manager._default_session_profile_dir(
                    "chatgpt-primary"
                )

            with (
                patch.object(browser_manager, "LEGACY_SESSION_ROOT", legacy_root),
                patch(
                    "runtime_v2.browser.manager.browser_session_root",
                    return_value=external_root,
                ),
                patch.dict(
                    os.environ,
                    {"RUNTIME_V2_ALLOW_LEGACY_SESSION_ROOT": "1"},
                    clear=False,
                ),
            ):
                allowed = browser_manager._default_session_profile_dir(
                    "chatgpt-primary"
                )

        self.assertEqual(blocked, str((external_root / "chatgpt-primary").resolve()))
        self.assertEqual(allowed, str((legacy_root / "chatgpt-primary").resolve()))


if __name__ == "__main__":
    _ = unittest.main()
