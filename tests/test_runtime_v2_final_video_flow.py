from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from typing import cast
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from runtime_v2.config import RuntimeConfig
from runtime_v2.contracts.job_contract import JobContract
from runtime_v2.control_plane import _run_worker
from runtime_v2.manager import sync_final_video_result
from runtime_v2.stage3.render_worker import run_render_job


def _write_excel_fixture(path: Path, *, status: str = "Voice OK") -> Path:
    workbook = Workbook()
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "Sheet1"
    sheet.append(["Topic", "Status", "Video Plan", "Reason Code"])
    sheet.append(["Bridge topic", status, "", ""])
    workbook.save(path)
    workbook.close()
    return path


def _read_status_row(path: Path) -> dict[str, object]:
    workbook = load_workbook(path)
    try:
        sheet = workbook["Sheet1"]
        return {
            "status": sheet.cell(row=2, column=2).value,
            "summary": sheet.cell(row=2, column=3).value,
            "reason_code": sheet.cell(row=2, column=4).value,
        }
    finally:
        workbook.close()


def _write_render_fixture(
    root: Path, *, final_name: str = "render_final.mp4"
) -> tuple[Path, Path, Path]:
    render_folder = root / "render_workspace"
    video_dir = render_folder / "video"
    output_dir = render_folder / "output"
    video_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    clip_path = video_dir / "#01_RVC.mp4"
    clip_path.write_bytes(b"mp4")
    final_output = output_dir / final_name
    final_output.write_bytes(b"final-mp4")
    voice_json = root / "voice.json"
    voice_json.write_text(
        json.dumps(
            {
                "voice_texts": [
                    {"col": "#01", "text": "bridge line", "original_voices": [1]}
                ]
            },
            ensure_ascii=True,
        ),
        encoding="utf-8",
    )
    render_spec = root / "render_spec.json"
    render_spec.write_text(
        json.dumps(
            {
                "contract": "render_spec",
                "locked": True,
                "asset_refs": [str(clip_path.resolve())],
                "timeline": [
                    {"scene_index": 1, "asset_path": str(clip_path.resolve())}
                ],
                "audio_refs": [str(voice_json.resolve())],
                "thumbnail_refs": [],
            },
            ensure_ascii=True,
        ),
        encoding="utf-8",
    )
    return render_folder, voice_json, render_spec


class RuntimeV2FinalVideoFlowTests(unittest.TestCase):
    def test_shorts_render_worker_generates_vertical_video(self) -> None:
        from runtime_v2.workers.shorts_render_worker import run_shorts_render_job

        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            source_video = root / "source.mp4"
            source_video.write_bytes(b"mp4")
            output_path = artifact_root / "shorts.mp4"
            job = JobContract(
                job_id="shorts-render-job",
                workload="shorts_render",
                payload={
                    "run_id": "shorts-run-1",
                    "source_video_path": str(source_video.resolve()),
                    "voice_json_path": str((root / "voice.json").resolve()),
                    "service_artifact_path": str(output_path.resolve()),
                },
            )
            (root / "voice.json").write_text(
                json.dumps(
                    {"voice_texts": [{"col": "#01", "text": "hello"}]},
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            commands: list[list[str]] = []

            def fake_process(
                command: list[str],
                *,
                cwd: Path,
                extra_env: dict[str, str] | None = None,
                timeout_sec: int = 3600,
            ) -> dict[str, object]:
                _ = cwd
                _ = extra_env
                _ = timeout_sec
                commands.append(command)
                output = Path(str(command[-1]))
                output.parent.mkdir(parents=True, exist_ok=True)
                output.write_bytes(b"mp4")
                return {
                    "command": command,
                    "cwd": str(cwd),
                    "exit_code": 0,
                    "stdout": "",
                    "stderr": "",
                    "timed_out": False,
                    "timeout_sec": 3600,
                    "duration_sec": 0.01,
                }

            with patch(
                "runtime_v2.workers.shorts_render_worker.run_external_process",
                side_effect=fake_process,
            ):
                result = run_shorts_render_job(job, artifact_root=artifact_root)

            self.assertEqual(result["status"], "ok")
            self.assertTrue(output_path.exists())
            self.assertTrue(
                any(
                    "-filter_complex" in part
                    for command in commands
                    for part in command
                )
            )
            self.assertTrue(
                any(
                    "overlay=(W-w)/2:(H-h)/2" in part
                    for command in commands
                    for part in command
                )
            )

    def test_shorts_render_worker_fails_closed_without_voice_json(self) -> None:
        from runtime_v2.workers.shorts_render_worker import run_shorts_render_job

        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            source_video = root / "source.mp4"
            source_video.write_bytes(b"mp4")
            output_path = artifact_root / "shorts.mp4"
            job = JobContract(
                job_id="shorts-render-job-missing-voice-json",
                workload="shorts_render",
                payload={
                    "run_id": "shorts-run-2",
                    "source_video_path": str(source_video.resolve()),
                    "service_artifact_path": str(output_path.resolve()),
                },
            )

            result = run_shorts_render_job(job, artifact_root=artifact_root)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "missing_shorts_inputs")

    def test_shorts_render_worker_fails_closed_when_voice_json_staging_raises_oserror(
        self,
    ) -> None:
        from runtime_v2.workers.shorts_render_worker import run_shorts_render_job

        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            source_video = root / "source.mp4"
            voice_json = root / "voice.json"
            output_path = artifact_root / "shorts.mp4"
            source_video.write_bytes(b"mp4")
            voice_json.write_text(
                json.dumps(
                    {"voice_texts": [{"col": "#01", "text": "hello"}]},
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="shorts-render-job-stage-io-fail",
                workload="shorts_render",
                payload={
                    "run_id": "shorts-run-3",
                    "source_video_path": str(source_video.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "service_artifact_path": str(output_path.resolve()),
                },
            )

            with patch(
                "runtime_v2.workers.shorts_render_worker.stage_local_input",
                side_effect=OSError("copy failed"),
            ):
                result = run_shorts_render_job(job, artifact_root=artifact_root)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "shorts_input_io_failed")

    def test_render_worker_fails_closed_without_render_inputs(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            artifact_root = Path(tmp_dir) / "artifacts"
            job = JobContract(
                job_id="render-job-fail", workload="render", payload={"timeline": []}
            )

            result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "missing_render_inputs")

    def test_render_worker_blocks_without_native_render_implementation(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder, voice_json, render_spec = _write_render_fixture(root)
            job = JobContract(
                job_id="render-job-ok",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )
            result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["stage"], "render")
        completion = cast(dict[object, object], result["completion"])
        details = cast(dict[object, object], result["details"])
        self.assertTrue(bool(completion["final_output"]))
        self.assertEqual(completion["reused"], True)
        self.assertEqual(str(details["render_mode"]), "reused")
        self.assertTrue(
            str(completion["final_artifact_path"]).endswith("render_final.mp4")
        )

    def test_render_worker_emits_srt_next_job_when_render_succeeds(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder, voice_json, render_spec = _write_render_fixture(root)
            job = JobContract(
                job_id="render-job-srt",
                workload="render",
                payload={
                    "run_id": "render-run-1",
                    "row_ref": "Sheet1!row1",
                    "chain_depth": 2,
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "ok")
        next_jobs = cast(list[object], result.get("next_jobs", []))
        self.assertEqual(len(next_jobs), 2)
        next_job_contracts = [cast(dict[str, object], job) for job in next_jobs]
        next_job_contract = cast(
            dict[str, object],
            next(
                job
                for job in next_job_contracts
                if str(cast(dict[str, object], job["job"])["worker"]) == "srt"
            ),
        )
        chain = cast(dict[str, object], next_job_contract["chain"])
        next_job = cast(dict[str, object], next_job_contract["job"])
        next_payload = cast(dict[str, object], next_job["payload"])
        self.assertEqual(str(next_job["worker"]), "srt")
        self.assertEqual(str(next_job["job_id"]), "srt-render-job-srt")
        self.assertEqual(cast(int, chain["step"]), 3)
        self.assertEqual(cast(int, next_payload["chain_depth"]), 3)
        self.assertTrue(
            str(next_payload["voice_json_path"])
            .replace("\\", "/")
            .endswith("/voice.json")
        )
        self.assertTrue(
            str(next_payload["render_spec_path"])
            .replace("\\", "/")
            .endswith("/render_spec.json")
        )
        self.assertTrue(
            str(next_payload["service_artifact_path"])
            .replace("\\", "/")
            .endswith("/render_final.srt")
        )
        shorts_contract = cast(
            dict[str, object],
            next(
                job
                for job in next_job_contracts
                if str(cast(dict[str, object], job["job"])["worker"]) == "shorts_render"
            ),
        )
        shorts_chain = cast(dict[str, object], shorts_contract["chain"])
        shorts_job = cast(dict[str, object], shorts_contract["job"])
        shorts_payload = cast(dict[str, object], shorts_job["payload"])
        self.assertEqual(str(shorts_job["job_id"]), "shorts-render-job-srt")
        self.assertEqual(cast(int, shorts_chain["step"]), 3)
        self.assertEqual(cast(int, shorts_payload["chain_depth"]), 3)
        self.assertEqual(
            str(shorts_payload["source_video_path"]),
            str((render_folder / "output" / "render_final.mp4").resolve()),
        )
        self.assertTrue(
            str(shorts_payload["voice_json_path"])
            .replace("\\", "/")
            .endswith("/voice.json")
        )
        self.assertTrue(
            str(shorts_payload["service_artifact_path"])
            .replace("\\", "/")
            .endswith("/shorts/shorts_final.mp4")
        )

    def test_render_worker_emits_n8n_upload_when_callback_url_present(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder, voice_json, render_spec = _write_render_fixture(root)
            job = JobContract(
                job_id="render-job-upload",
                workload="render",
                payload={
                    "run_id": "render-run-upload",
                    "row_ref": "Sheet1!row2",
                    "chain_depth": 1,
                    "callback_url": "https://example.test/webhook",
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "ok")
        next_jobs = [
            cast(dict[str, object], item)
            for item in cast(list[object], result.get("next_jobs", []))
        ]
        upload_contract = next(
            item
            for item in next_jobs
            if str(cast(dict[str, object], item["job"])["worker"]) == "n8n_upload"
        )
        upload_chain = cast(dict[str, object], upload_contract["chain"])
        upload_job = cast(dict[str, object], upload_contract["job"])
        upload_payload = cast(dict[str, object], upload_job["payload"])
        self.assertEqual(str(upload_job["job_id"]), "n8n-render-job-upload")
        self.assertEqual(cast(int, upload_chain["step"]), 2)
        self.assertEqual(
            str(upload_payload["callback_url"]), "https://example.test/webhook"
        )
        self.assertEqual(
            str(upload_payload["artifact_path"]),
            str((render_folder / "output" / "render_final.mp4").resolve()),
        )

    def test_render_worker_reports_native_only_boundary(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder, voice_json, render_spec = _write_render_fixture(root)
            job = JobContract(
                job_id="render-job-invalid-json",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )
            final_output = render_folder / "output" / "render_final.mp4"
            final_output.unlink()
            result = run_render_job(job, artifact_root)
            self.assertTrue(final_output.exists())

        self.assertEqual(result["status"], "ok")
        completion = cast(dict[object, object], result["completion"])
        details = cast(dict[object, object], result["details"])
        self.assertTrue(bool(completion["final_output"]))
        self.assertEqual(completion["reused"], False)
        self.assertEqual(str(details["render_mode"]), "video_copy")
        self.assertTrue(str(details["source_asset_path"]).endswith("#01_RVC.mp4"))
        self.assertTrue(
            str(completion["final_artifact_path"]).endswith("render_final.mp4")
        )

    def test_render_worker_fails_closed_on_invalid_render_spec_json(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder, voice_json, render_spec = _write_render_fixture(root)
            render_spec.write_text("{not-json", encoding="utf-8")
            job = JobContract(
                job_id="render-job-bad-spec",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "invalid_render_spec")
        completion = cast(dict[object, object], result["completion"])
        self.assertEqual(str(completion["state"]), "failed")

    def test_render_worker_fails_closed_when_render_spec_staging_raises_oserror(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder, voice_json, render_spec = _write_render_fixture(root)
            job = JobContract(
                job_id="render-job-stage-io-fail",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            with patch(
                "runtime_v2.stage3.render_worker.stage_local_input",
                side_effect=OSError("staging failed"),
            ):
                result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "render_io_failed")
        completion = cast(dict[object, object], result["completion"])
        self.assertEqual(str(completion["state"]), "failed")

    def test_render_worker_blocks_when_voice_texts_exist_but_audio_is_not_ready(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder = root / "render_workspace"
            image_dir = render_folder / "images"
            output_dir = render_folder / "output"
            image_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)
            image_path = image_dir / "#01_GENS.png"
            image_path.write_bytes(b"png")
            voice_json = root / "voice.json"
            voice_json.write_text(
                json.dumps(
                    {"voice_texts": [{"col": "#01", "text": "bridge"}]},
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            render_spec = root / "render_spec.json"
            render_spec.write_text(
                json.dumps(
                    {
                        "contract": "render_spec",
                        "locked": True,
                        "asset_refs": [str(image_path.resolve())],
                        "timeline": [
                            {"scene_index": 1, "asset_path": str(image_path.resolve())}
                        ],
                        "audio_refs": [str(voice_json.resolve())],
                        "thumbnail_refs": [],
                        "reason_code": "ok",
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="render-job-image",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "render_audio_not_ready")
        completion = cast(dict[object, object], result["completion"])
        self.assertEqual(str(completion["state"]), "blocked")
        self.assertFalse(bool(completion["final_output"]))

    def test_render_worker_builds_timeline_and_muxes_audio(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder = root / "render_workspace"
            image_dir = render_folder / "images"
            output_dir = render_folder / "output"
            image_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)
            image_a = image_dir / "#01_GENS.png"
            image_b = image_dir / "#02_SEAA.png"
            image_a.write_bytes(b"png")
            image_b.write_bytes(b"png")
            audio_path = root / "speech_rvc.wav"
            audio_path.write_bytes(b"wav")
            voice_json = root / "voice.json"
            voice_json.write_text(
                json.dumps({"voice_texts": []}, ensure_ascii=True), encoding="utf-8"
            )
            render_spec = root / "render_spec.json"
            render_spec.write_text(
                json.dumps(
                    {
                        "contract": "render_spec",
                        "contract_version": "1.1",
                        "run_id": "render-run-1",
                        "row_ref": "Sheet1!row1",
                        "asset_refs": [str(image_a.resolve()), str(image_b.resolve())],
                        "audio_refs": [str(audio_path.resolve())],
                        "thumbnail_refs": [],
                        "timeline": [
                            {
                                "scene_index": 1,
                                "asset_path": str(image_a.resolve()),
                                "asset_kind": "image",
                                "duration_sec": 4,
                            },
                            {
                                "scene_index": 2,
                                "asset_path": str(image_b.resolve()),
                                "asset_kind": "image",
                                "duration_sec": 5,
                            },
                        ],
                        "reason_code": "ok",
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="render-job-timeline",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            def fake_process(
                command: list[str],
                *,
                cwd: Path,
                extra_env: dict[str, str] | None = None,
                timeout_sec: int = 3600,
            ) -> dict[str, object]:
                _ = extra_env
                _ = timeout_sec
                output_path = Path(str(command[-1]))
                output_path.parent.mkdir(parents=True, exist_ok=True)
                output_path.write_bytes(b"mp4")
                return {
                    "command": command,
                    "cwd": str(cwd),
                    "exit_code": 0,
                    "stdout": "",
                    "stderr": "",
                    "timed_out": False,
                    "timeout_sec": 3600,
                    "duration_sec": 0.01,
                }

            with patch(
                "runtime_v2.stage3.render_worker.run_external_process",
                side_effect=fake_process,
            ):
                result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "ok")
        details = cast(dict[object, object], result["details"])
        completion = cast(dict[object, object], result["completion"])
        next_jobs = cast(list[object], result.get("next_jobs", []))
        self.assertEqual(str(details["render_mode"]), "timeline_ffmpeg_audio")
        self.assertEqual(str(details["audio_source_path"]), str(audio_path.resolve()))
        self.assertEqual(str(completion["state"]), "succeeded")
        self.assertTrue(bool(completion["final_output"]))
        self.assertEqual(len(next_jobs), 2)

    def test_render_worker_mixes_optional_bgm_track(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder = root / "render_workspace"
            video_dir = render_folder / "video"
            output_dir = render_folder / "output"
            video_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)
            clip_path = video_dir / "#01_RVC.mp4"
            clip_path.write_bytes(b"mp4")
            voice_json = root / "voice.json"
            voice_json.write_text(
                json.dumps({"voice_texts": []}, ensure_ascii=True), encoding="utf-8"
            )
            voice_audio = root / "speech.wav"
            voice_audio.write_bytes(b"wav")
            bgm_audio = root / "bgm.mp3"
            bgm_audio.write_bytes(b"mp3")
            render_spec = root / "render_spec.json"
            render_spec.write_text(
                json.dumps(
                    {
                        "contract": "render_spec",
                        "contract_version": "1.1",
                        "run_id": "render-run-bgm",
                        "row_ref": "Sheet1!row1",
                        "asset_refs": [str(clip_path.resolve())],
                        "audio_refs": [str(voice_audio.resolve())],
                        "bgm_path": str(bgm_audio.resolve()),
                        "bgm_volume": 0.15,
                        "thumbnail_refs": [],
                        "timeline": [
                            {
                                "scene_index": 1,
                                "asset_path": str(clip_path.resolve()),
                                "asset_kind": "video",
                                "duration_sec": 6,
                            }
                        ],
                        "reason_code": "ok",
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="render-job-bgm",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )
            commands: list[list[str]] = []

            def fake_process(
                command: list[str],
                *,
                cwd: Path,
                extra_env: dict[str, str] | None = None,
                timeout_sec: int = 3600,
            ) -> dict[str, object]:
                _ = cwd
                _ = extra_env
                _ = timeout_sec
                commands.append(command)
                output_path = Path(str(command[-1]))
                output_path.parent.mkdir(parents=True, exist_ok=True)
                output_path.write_bytes(b"mp4")
                return {
                    "command": command,
                    "cwd": str(cwd),
                    "exit_code": 0,
                    "stdout": "",
                    "stderr": "",
                    "timed_out": False,
                    "timeout_sec": 3600,
                    "duration_sec": 0.01,
                }

            with patch(
                "runtime_v2.stage3.render_worker.run_external_process",
                side_effect=fake_process,
            ):
                result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "ok")
        details = cast(dict[object, object], result["details"])
        self.assertEqual(str(details["render_mode"]), "timeline_ffmpeg_audio")
        self.assertTrue(
            any("-filter_complex" in command for command in commands for _ in [0])
        )
        self.assertTrue(
            any(
                "amix=inputs=2:duration=first:normalize=0" in part
                for command in commands
                for part in command
            )
        )

    def test_render_worker_concats_multiple_audio_refs_before_mux(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder = root / "render_workspace"
            image_dir = render_folder / "images"
            output_dir = render_folder / "output"
            image_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)
            image_path = image_dir / "#01_GENS.png"
            image_path.write_bytes(b"png")
            audio_a = root / "speech_a.flac"
            audio_b = root / "speech_b.flac"
            audio_a.write_bytes(b"flac")
            audio_b.write_bytes(b"flac")
            voice_json = root / "voice.json"
            voice_json.write_text(
                json.dumps({"voice_texts": []}, ensure_ascii=True), encoding="utf-8"
            )
            render_spec = root / "render_spec.json"
            render_spec.write_text(
                json.dumps(
                    {
                        "contract": "render_spec",
                        "contract_version": "1.1",
                        "run_id": "render-run-multi-audio",
                        "row_ref": "Sheet1!row1",
                        "asset_refs": [str(image_path.resolve())],
                        "audio_refs": [str(audio_a.resolve()), str(audio_b.resolve())],
                        "thumbnail_refs": [],
                        "timeline": [
                            {
                                "scene_index": 1,
                                "asset_path": str(image_path.resolve()),
                                "asset_kind": "image",
                                "duration_sec": 4,
                            }
                        ],
                        "reason_code": "ok",
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="render-job-multi-audio",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )
            commands: list[list[str]] = []

            def fake_process(
                command: list[str],
                *,
                cwd: Path,
                extra_env: dict[str, str] | None = None,
                timeout_sec: int = 3600,
            ) -> dict[str, object]:
                _ = extra_env
                _ = timeout_sec
                commands.append(command)
                output_path = Path(str(command[-1]))
                output_path.parent.mkdir(parents=True, exist_ok=True)
                output_path.write_bytes(b"bin")
                return {
                    "command": command,
                    "cwd": str(cwd),
                    "exit_code": 0,
                    "stdout": "",
                    "stderr": "",
                    "timed_out": False,
                    "timeout_sec": 3600,
                    "duration_sec": 0.01,
                }

            with patch(
                "runtime_v2.stage3.render_worker.run_external_process",
                side_effect=fake_process,
            ):
                result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "ok")
        details = cast(dict[object, object], result["details"])
        self.assertEqual(str(details["render_mode"]), "timeline_ffmpeg_audio")
        self.assertTrue(
            str(details["audio_source_path"]).endswith("_audio_concat.flac")
        )
        self.assertTrue(any("concat" in command for command in commands for _ in [0]))

    def test_render_worker_falls_back_to_canonical_rvc_audio(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder = root / "render_workspace"
            video_dir = render_folder / "video"
            output_dir = render_folder / "output"
            video_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)
            clip_path = video_dir / "#01_GEMI.mp4"
            clip_path.write_bytes(b"mp4")
            canonical_audio = (
                artifact_root / "rvc" / "rvc-qwen3-render-run-2" / "speech_rvc.wav"
            )
            canonical_audio.parent.mkdir(parents=True, exist_ok=True)
            canonical_audio.write_bytes(b"wav")
            voice_json = root / "voice.json"
            voice_json.write_text(
                json.dumps({"voice_texts": []}, ensure_ascii=True), encoding="utf-8"
            )
            render_spec = root / "render_spec.json"
            render_spec.write_text(
                json.dumps(
                    {
                        "contract": "render_spec",
                        "contract_version": "1.1",
                        "run_id": "render-run-2",
                        "row_ref": "Sheet1!row1",
                        "asset_refs": [str(clip_path.resolve())],
                        "audio_refs": [],
                        "thumbnail_refs": [],
                        "timeline": [
                            {
                                "scene_index": 1,
                                "asset_path": str(clip_path.resolve()),
                                "asset_kind": "video",
                                "duration_sec": 6,
                            }
                        ],
                        "reason_code": "ok",
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="render-job-canonical-audio",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            def fake_process(
                command: list[str],
                *,
                cwd: Path,
                extra_env: dict[str, str] | None = None,
                timeout_sec: int = 3600,
            ) -> dict[str, object]:
                _ = extra_env
                _ = timeout_sec
                output_path = Path(str(command[-1]))
                output_path.parent.mkdir(parents=True, exist_ok=True)
                output_path.write_bytes(b"mp4")
                return {
                    "command": command,
                    "cwd": str(cwd),
                    "exit_code": 0,
                    "stdout": "",
                    "stderr": "",
                    "timed_out": False,
                    "timeout_sec": 3600,
                    "duration_sec": 0.01,
                }

            with patch(
                "runtime_v2.stage3.render_worker.run_external_process",
                side_effect=fake_process,
            ):
                result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "ok")
        details = cast(dict[object, object], result["details"])
        self.assertEqual(
            str(details["audio_source_path"]), str(canonical_audio.resolve())
        )

    def test_render_worker_falls_back_to_canonical_geminigen_rvc_audio(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder = root / "render_workspace"
            video_dir = render_folder / "video"
            output_dir = render_folder / "output"
            video_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)
            clip_path = video_dir / "#01_GEMI.mp4"
            clip_path.write_bytes(b"mp4")
            canonical_audio = (
                artifact_root / "rvc" / "rvc-geminigen-render-run-3" / "speech_rvc.wav"
            )
            canonical_audio.parent.mkdir(parents=True, exist_ok=True)
            canonical_audio.write_bytes(b"wav")
            voice_json = root / "voice.json"
            voice_json.write_text(
                json.dumps({"voice_texts": []}, ensure_ascii=True), encoding="utf-8"
            )
            render_spec = root / "render_spec.json"
            render_spec.write_text(
                json.dumps(
                    {
                        "contract": "render_spec",
                        "contract_version": "1.1",
                        "run_id": "render-run-3",
                        "row_ref": "Sheet1!row1",
                        "asset_refs": [str(clip_path.resolve())],
                        "audio_refs": [],
                        "thumbnail_refs": [],
                        "timeline": [
                            {
                                "scene_index": 1,
                                "asset_path": str(clip_path.resolve()),
                                "asset_kind": "video",
                                "duration_sec": 6,
                            }
                        ],
                        "reason_code": "ok",
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="render-job-canonical-gemi-audio",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            def fake_process(
                command: list[str],
                *,
                cwd: Path,
                extra_env: dict[str, str] | None = None,
                timeout_sec: int = 3600,
            ) -> dict[str, object]:
                _ = extra_env
                _ = timeout_sec
                output_path = Path(str(command[-1]))
                output_path.parent.mkdir(parents=True, exist_ok=True)
                output_path.write_bytes(b"mp4")
                return {
                    "command": command,
                    "cwd": str(cwd),
                    "exit_code": 0,
                    "stdout": "",
                    "stderr": "",
                    "timed_out": False,
                    "timeout_sec": 3600,
                    "duration_sec": 0.01,
                }

            with patch(
                "runtime_v2.stage3.render_worker.run_external_process",
                side_effect=fake_process,
            ):
                result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "ok")
        details = cast(dict[object, object], result["details"])
        self.assertEqual(
            str(details["audio_source_path"]), str(canonical_audio.resolve())
        )

    def test_render_worker_blocks_when_voice_texts_exist_but_audio_is_missing(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder = root / "render_workspace"
            video_dir = render_folder / "video"
            output_dir = render_folder / "output"
            video_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)
            clip_path = video_dir / "#01_GEMI.mp4"
            clip_path.write_bytes(b"mp4")
            voice_json = root / "voice.json"
            voice_json.write_text(
                json.dumps(
                    {"voice_texts": [{"col": "#01", "text": "bridge line"}]},
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            render_spec = root / "render_spec.json"
            render_spec.write_text(
                json.dumps(
                    {
                        "contract": "render_spec",
                        "contract_version": "1.1",
                        "run_id": "render-run-audio-blocked",
                        "row_ref": "Sheet1!row1",
                        "asset_refs": [str(clip_path.resolve())],
                        "audio_refs": [],
                        "thumbnail_refs": [],
                        "timeline": [
                            {
                                "scene_index": 1,
                                "asset_path": str(clip_path.resolve()),
                                "asset_kind": "video",
                                "duration_sec": 6,
                            }
                        ],
                        "reason_code": "ok",
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="render-job-audio-blocked",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "render_audio_not_ready")
        completion = cast(dict[object, object], result["completion"])
        self.assertEqual(str(completion["state"]), "blocked")

    def test_render_worker_fails_closed_on_render_spec_run_id_mismatch(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder = root / "render_workspace"
            video_dir = render_folder / "video"
            output_dir = render_folder / "output"
            video_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)
            clip_path = video_dir / "#01_RVC.mp4"
            clip_path.write_bytes(b"mp4")
            voice_json = root / "voice.json"
            voice_json.write_text(
                json.dumps({"voice_texts": []}, ensure_ascii=True), encoding="utf-8"
            )
            render_spec = root / "render_spec.json"
            render_spec.write_text(
                json.dumps(
                    {
                        "contract": "render_spec",
                        "run_id": "spec-run-1",
                        "row_ref": "Sheet1!row1",
                        "asset_refs": [str(clip_path.resolve())],
                        "audio_refs": [],
                        "thumbnail_refs": [],
                        "timeline": [
                            {
                                "scene_index": 1,
                                "asset_path": str(clip_path.resolve()),
                                "asset_kind": "video",
                            }
                        ],
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="render-job-run-mismatch",
                workload="render",
                payload={
                    "run_id": "job-run-1",
                    "row_ref": "Sheet1!row1",
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "render_run_id_mismatch")

    def test_render_worker_fails_closed_on_asset_manifest_run_id_mismatch(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder = root / "render_workspace"
            video_dir = render_folder / "video"
            output_dir = render_folder / "output"
            video_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)
            clip_path = video_dir / "#01_RVC.mp4"
            clip_path.write_bytes(b"mp4")
            voice_json = root / "voice.json"
            voice_json.write_text(
                json.dumps({"voice_texts": []}, ensure_ascii=True), encoding="utf-8"
            )
            render_spec = root / "render_spec.json"
            render_spec.write_text(
                json.dumps(
                    {
                        "contract": "render_spec",
                        "run_id": "job-run-1",
                        "row_ref": "Sheet1!row1",
                        "asset_refs": [str(clip_path.resolve())],
                        "audio_refs": [],
                        "thumbnail_refs": [],
                        "timeline": [
                            {
                                "scene_index": 1,
                                "asset_path": str(clip_path.resolve()),
                                "asset_kind": "video",
                            }
                        ],
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            asset_manifest = root / "asset_manifest.json"
            asset_manifest.write_text(
                json.dumps(
                    {
                        "run_id": "other-run",
                        "row_ref": "Sheet1!row1",
                        "roles": {"voice_json": "D:/voice.json"},
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            job = JobContract(
                job_id="render-job-manifest-mismatch",
                workload="render",
                payload={
                    "run_id": "job-run-1",
                    "row_ref": "Sheet1!row1",
                    "asset_manifest_path": str(asset_manifest.resolve()),
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            result = run_render_job(job, artifact_root)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "render_manifest_run_id_mismatch")

    def test_render_worker_blocks_retry_when_render_assets_are_not_ready(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder, voice_json, render_spec = _write_render_fixture(root)
            missing_clip = render_folder / "video" / "#01_RVC.mp4"
            missing_clip.unlink()
            job = JobContract(
                job_id="render-job-blocked",
                workload="render",
                payload={
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                    "render_spec_path": str(render_spec.resolve()),
                },
            )

            result = run_render_job(job, artifact_root)
            self.assertEqual(result["status"], "failed")
            self.assertEqual(result["error_code"], "render_inputs_not_ready")
            self.assertTrue(bool(result["retryable"]))
            completion = cast(dict[object, object], result["completion"])
            self.assertEqual(str(completion["state"]), "blocked")
            details = cast(dict[object, object], result["details"])
            self.assertIn(
                str(missing_clip.resolve()),
                cast(list[object], details["missing_paths"]),
            )

    def test_final_video_success_marks_excel_done_and_updates_latest_run(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx", status="Voice OK")
            config = RuntimeConfig(result_router_file=root / "evidence" / "result.json")
            final_video = root / "final_video.mp4"
            final_video.write_bytes(b"mp4")

            updated = sync_final_video_result(
                config=config,
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                worker_result={
                    "completion": {
                        "state": "completed",
                        "final_output": True,
                        "final_artifact": "final_video.mp4",
                        "final_artifact_path": str(final_video.resolve()),
                    },
                    "details": {"reason_code": "ok"},
                },
                run_id="final-run-1",
                artifact_root=root,
                debug_log=str((root / "logs" / "final-run-1.jsonl").resolve()),
            )

            status_row = _read_status_row(excel_path)
            latest_result = json.loads(
                config.result_router_file.read_text(encoding="utf-8")
            )

        self.assertTrue(updated)
        self.assertEqual(status_row["status"], "Done")
        self.assertEqual(len(latest_result["artifacts"]), 1)
        self.assertTrue(latest_result["metadata"]["final_output"])
        self.assertTrue(
            str(latest_result["metadata"]["final_artifact_path"]).endswith(
                "final_video.mp4"
            )
        )

    def test_partial_failure_marks_excel_partial_with_reason(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx", status="Video OK")
            config = RuntimeConfig(result_router_file=root / "evidence" / "result.json")

            updated = sync_final_video_result(
                config=config,
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                worker_result={
                    "completion": {"state": "failed", "final_output": False},
                    "error_code": "ffmpeg_failed",
                },
                run_id="final-run-2",
                artifact_root=root,
                debug_log=str((root / "logs" / "final-run-2.jsonl").resolve()),
            )

            status_row = _read_status_row(excel_path)

        self.assertTrue(updated)
        self.assertEqual(status_row["status"], "partial")
        self.assertEqual(status_row["reason_code"], "ffmpeg_failed")

    def test_partial_failure_with_none_reason_code_falls_back_to_ok(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx", status="Video OK")
            config = RuntimeConfig(result_router_file=root / "evidence" / "result.json")

            updated = sync_final_video_result(
                config=config,
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                worker_result={
                    "completion": {"state": "failed", "final_output": False},
                    "details": {"reason_code": None},
                },
                run_id="final-run-3",
                artifact_root=root,
                debug_log=str((root / "logs" / "final-run-3.jsonl").resolve()),
            )

            status_row = _read_status_row(excel_path)

        self.assertTrue(updated)
        self.assertEqual(status_row["reason_code"], "ok")

    def test_completed_final_artifact_path_without_flag_still_syncs_done(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx", status="Voice OK")
            config = RuntimeConfig(result_router_file=root / "evidence" / "result.json")
            final_video = root / "final_video.mp4"
            final_video.write_bytes(b"mp4")

            updated = sync_final_video_result(
                config=config,
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                worker_result={
                    "completion": {
                        "state": "completed",
                        "final_output": False,
                        "final_artifact": "final_video.mp4",
                        "final_artifact_path": str(final_video.resolve()),
                    },
                    "details": {"reason_code": "ok"},
                },
                run_id="final-run-4",
                artifact_root=root,
                debug_log=str((root / "logs" / "final-run-4.jsonl").resolve()),
            )

            status_row = _read_status_row(excel_path)

        self.assertTrue(updated)
        self.assertEqual(status_row["status"], "Done")

    def test_done_row_closeout_is_idempotent(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx", status="Done")
            config = RuntimeConfig(result_router_file=root / "evidence" / "result.json")
            final_video = root / "final_video.mp4"
            final_video.write_bytes(b"mp4")

            updated = sync_final_video_result(
                config=config,
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                worker_result={
                    "completion": {
                        "state": "completed",
                        "final_output": True,
                        "final_artifact": "final_video.mp4",
                        "final_artifact_path": str(final_video.resolve()),
                    },
                    "details": {"reason_code": "ok"},
                },
                run_id="final-run-idempotent",
                artifact_root=root,
                debug_log=str((root / "logs" / "final-run-idempotent.jsonl").resolve()),
            )

            status_row = _read_status_row(excel_path)
            latest_result = json.loads(
                config.result_router_file.read_text(encoding="utf-8")
            )

        self.assertTrue(updated)
        self.assertEqual(status_row["status"], "Done")
        self.assertTrue(bool(latest_result["metadata"]["excel_synced"]))

    def test_render_spec_is_merged_only_by_manager(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            render_folder, voice_json, canonical_render_spec = _write_render_fixture(
                root
            )
            job = JobContract(
                job_id="render-job-1",
                workload="render",
                payload={
                    "render_spec_path": str(canonical_render_spec.resolve()),
                    "render_folder_path": str(render_folder.resolve()),
                    "voice_json_path": str(voice_json.resolve()),
                },
            )
            result = run_render_job(job, artifact_root)

            workspace_spec = (
                artifact_root / "render" / "render-job-1" / "render_spec.json"
            )
            canonical_payload = canonical_render_spec.read_text(encoding="utf-8")
            workspace_spec_exists = workspace_spec.exists()

        self.assertEqual(result["status"], "ok")
        self.assertIn('"locked": true', canonical_payload)
        self.assertTrue(workspace_spec_exists)

    def test_final_stage_workers_remain_resident_while_processing_multiple_jobs(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = RuntimeConfig(
                worker_registry_file=root / "health" / "worker_registry.json",
                artifact_root=root / "artifacts",
            )
            render_folder, voice_json, _ = _write_render_fixture(root)
            render_payload = {
                "render_folder_path": str(render_folder.resolve()),
                "voice_json_path": str(voice_json.resolve()),
            }
            render_job_1 = JobContract(
                job_id="render-job-1", workload="render", payload=dict(render_payload)
            )
            render_job_2 = JobContract(
                job_id="render-job-2", workload="render", payload=dict(render_payload)
            )
            result_1 = _run_worker(
                render_job_1,
                config.artifact_root,
                registry_file=config.worker_registry_file,
            )
            result_2 = _run_worker(
                render_job_2,
                config.artifact_root,
                registry_file=config.worker_registry_file,
            )
            registry_payload = json.loads(
                config.worker_registry_file.read_text(encoding="utf-8")
            )

        self.assertEqual(result_1["status"], "ok")
        self.assertEqual(result_2["status"], "ok")
        self.assertEqual(registry_payload["render"]["state"], "idle")


if __name__ == "__main__":
    _ = unittest.main()
