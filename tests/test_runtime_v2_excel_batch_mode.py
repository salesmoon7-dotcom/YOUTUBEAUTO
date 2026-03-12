from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from typing import cast
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from runtime_v2.cli import _run_excel_batch_mode
from runtime_v2.config import RuntimeConfig
from runtime_v2.excel.selector import select_pending_row_indexes


def _write_batch_excel(path: Path) -> Path:
    workbook = Workbook()
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "Sheet1"
    sheet.append(["Topic", "Status", "Video Plan", "Reason Code"])
    sheet.append(["Topic A", "", "", ""])
    sheet.append(["Topic B", "failed", "", ""])
    sheet.append(["Topic C", "Done", "", ""])
    sheet.append(["Topic D", "partial", "", ""])
    workbook.save(path)
    workbook.close()
    return path


class RuntimeV2ExcelBatchModeTests(unittest.TestCase):
    def test_select_pending_row_indexes_returns_first_pending_rows(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            excel_path = _write_batch_excel(Path(tmp_dir) / "topic.xlsx")

            selected = select_pending_row_indexes(
                excel_path,
                sheet_name="Sheet1",
                limit=3,
            )

        self.assertEqual(selected, [0, 1, 3])

    def test_run_excel_batch_mode_seeds_rows_and_drives_control_loop(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_batch_excel(root / "topic.xlsx")
            config = RuntimeConfig.from_root(root)

            def fake_control_loop_once(
                *, owner: str, config: RuntimeConfig, run_id: str
            ) -> dict[str, object]:
                _ = owner
                _ = config
                _ = run_id
                workbook = load_workbook(excel_path)
                try:
                    sheet = workbook["Sheet1"]
                    self.assertEqual(
                        cast(Cell, sheet.cell(row=2, column=2)).value, "Running"
                    )
                    self.assertEqual(
                        cast(Cell, sheet.cell(row=3, column=2)).value, "Running"
                    )
                    self.assertEqual(
                        cast(Cell, sheet.cell(row=5, column=2)).value, "Running"
                    )
                    cast(Cell, sheet.cell(row=2, column=2)).value = "Done"
                    cast(Cell, sheet.cell(row=3, column=2)).value = "failed"
                    cast(Cell, sheet.cell(row=5, column=2)).value = "Done"
                    workbook.save(excel_path)
                finally:
                    workbook.close()
                return {"status": "ok", "code": "OK"}

            with patch(
                "runtime_v2.cli.run_control_loop_once",
                side_effect=fake_control_loop_once,
            ):
                result = _run_excel_batch_mode(
                    owner="runtime_v2",
                    config=config,
                    run_id="batch-run-1",
                    excel_path=str(excel_path),
                    sheet_name="Sheet1",
                    batch_count=3,
                    max_control_ticks=5,
                )

            workbook = load_workbook(excel_path)
            try:
                statuses = [
                    workbook["Sheet1"].cell(row=2, column=2).value,
                    workbook["Sheet1"].cell(row=3, column=2).value,
                    workbook["Sheet1"].cell(row=5, column=2).value,
                ]
            finally:
                workbook.close()

        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["code"], "OK")
        self.assertEqual(statuses, ["Done", "failed", "Done"])
        selected_rows = cast(list[object], result["selected_rows"])
        self.assertEqual(len(selected_rows), 3)


if __name__ == "__main__":
    _ = unittest.main()
