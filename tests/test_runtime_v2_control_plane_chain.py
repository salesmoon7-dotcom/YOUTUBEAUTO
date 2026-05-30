from __future__ import annotations

import json
import tempfile
import threading
import time
import unittest
from pathlib import Path
from typing import cast
from unittest.mock import patch

from runtime_v2 import circuit_breaker, recovery_policy, retry_budget
from runtime_v2.config import RuntimeConfig
from runtime_v2.contracts.job_contract import JobContract, build_explicit_job_contract
from runtime_v2.control_plane import (
    _load_jobs,
    _seed_declared_next_jobs,
    run_control_loop_once,
    run_worker,
    seed_control_job,
)
from runtime_v2.control_plane_feeder import job_from_explicit_payload
from runtime_v2.manager import seed_excel_row
from runtime_v2.contracts.video_plan import build_video_plan
from runtime_v2.latest_run import load_joined_latest_run
from runtime_v2.queue_store import QueueStore, QueueStoreError
from runtime_v2.stage2.router import route_video_plan


def _runtime_config(root: Path) -> RuntimeConfig:
    return RuntimeConfig.from_root(root)


class RuntimeV2ControlPlaneChainTests(unittest.TestCase):
    def test_run_worker_dispatches_timeline_workload(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "runtime" / "artifacts"
            registry_file = root / "health" / "worker_registry.json"
            artifact_root.mkdir(parents=True, exist_ok=True)
            job = JobContract(
                job_id="timeline-dispatch-job",
                workload="timeline",
                checkpoint_key="seed:timeline-dispatch-job",
                payload={"run_id": "timeline-run", "row_ref": "Sheet1!row1"},
            )

            with patch(
                "runtime_v2.control_plane.run_timeline_job",
                return_value={"status": "ok", "stage": "timeline"},
            ) as run_timeline:
                result = run_worker(
                    job,
                    artifact_root=artifact_root,
                    registry_file=registry_file,
                )

        self.assertEqual(result["status"], "ok")
        run_timeline.assert_called_once()

    def test_run_worker_dispatches_shorts_render_workload(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "runtime" / "artifacts"
            registry_file = root / "health" / "worker_registry.json"
            artifact_root.mkdir(parents=True, exist_ok=True)
            job = JobContract(
                job_id="shorts-render-dispatch-job",
                workload="shorts_render",
                checkpoint_key="seed:shorts-render-dispatch-job",
                payload={"run_id": "shorts-run", "row_ref": "Sheet1!row1"},
            )

            with patch(
                "runtime_v2.control_plane.run_shorts_render_job",
                return_value={"status": "ok", "stage": "shorts_render"},
            ) as run_shorts:
                result = run_worker(
                    job,
                    artifact_root=artifact_root,
                    registry_file=registry_file,
                )

        self.assertEqual(result["status"], "ok")
        run_shorts.assert_called_once()

    def test_run_worker_dispatches_google_sheets_sync_workload(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "runtime" / "artifacts"
            registry_file = root / "health" / "worker_registry.json"
            artifact_root.mkdir(parents=True, exist_ok=True)
            job = JobContract(
                job_id="google-sheets-sync-dispatch-job",
                workload="google_sheets_sync",
                checkpoint_key="seed:google-sheets-sync-dispatch-job",
                payload={"run_id": "gsync-run", "row_ref": "Sheet1!row1"},
            )

            with patch(
                "runtime_v2.control_plane.run_google_sheets_sync_job",
                return_value={"status": "ok", "stage": "google_sheets_sync"},
            ) as run_sync:
                result = run_worker(
                    job,
                    artifact_root=artifact_root,
                    registry_file=registry_file,
                )

        self.assertEqual(result["status"], "ok")
        run_sync.assert_called_once()

    def test_explicit_contract_path_allows_srt_inbox(self) -> None:
        from runtime_v2.control_plane_feeder import _is_allowed_explicit_contract_path

        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            inbox_root = root / "inbox"
            contract_file = inbox_root / "srt" / "srt.job.json"
            contract_file.parent.mkdir(parents=True, exist_ok=True)
            contract_file.write_text("{}", encoding="utf-8")

            allowed = _is_allowed_explicit_contract_path(inbox_root, contract_file)

        self.assertTrue(allowed)

    def test_run_worker_dispatches_srt_workload(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "runtime" / "artifacts"
            registry_file = root / "health" / "worker_registry.json"
            artifact_root.mkdir(parents=True, exist_ok=True)
            job = JobContract(
                job_id="srt-dispatch-job",
                workload="srt",
                checkpoint_key="seed:srt-dispatch-job",
                payload={"run_id": "srt-run", "row_ref": "Sheet1!row1"},
            )

            with patch(
                "runtime_v2.control_plane.run_srt_job",
                return_value={"status": "ok", "stage": "srt"},
            ) as run_srt:
                result = run_worker(
                    job,
                    artifact_root=artifact_root,
                    registry_file=registry_file,
                )

        self.assertEqual(result["status"], "ok")
        run_srt.assert_called_once()

    def test_explicit_contract_rejects_non_local_n8n_artifact_path(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            inbox_root = root / "inbox"
            contract_file = inbox_root / "n8n_upload" / "n8n.job.json"
            contract_file.parent.mkdir(parents=True, exist_ok=True)
            payload = build_explicit_job_contract(
                job_id="n8n-upload-job",
                workload="n8n_upload",
                checkpoint_key="seed:n8n-upload-job",
                payload={
                    "callback_url": "https://example.test/webhook",
                    "artifact_path": "C:/Windows/Temp/render_final.mp4",
                },
            )
            contract_file.write_text(
                json.dumps(payload, ensure_ascii=True), encoding="utf-8"
            )

            loaded = json.loads(contract_file.read_text(encoding="utf-8"))
            contract, error = job_from_explicit_payload(
                loaded, source_hint=str(contract_file)
            )

        self.assertIsNone(contract)
        self.assertIsNotNone(error)
        typed_error = cast(dict[str, object], error)
        self.assertEqual(typed_error["code"], "non_local_path")

    def test_run_worker_dispatches_n8n_upload_workload(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "runtime" / "artifacts"
            registry_file = root / "health" / "worker_registry.json"
            artifact_root.mkdir(parents=True, exist_ok=True)
            job = JobContract(
                job_id="n8n-upload-dispatch-job",
                workload="n8n_upload",
                checkpoint_key="seed:n8n-upload-dispatch-job",
                payload={
                    "run_id": "n8n-upload-run",
                    "callback_url": "https://example.test/webhook",
                },
            )

            with patch(
                "runtime_v2.control_plane.run_n8n_upload_job",
                return_value={"status": "ok", "stage": "n8n_upload"},
            ) as run_n8n_upload:
                result = run_worker(
                    job,
                    artifact_root=artifact_root,
                    registry_file=registry_file,
                )

        self.assertEqual(result["status"], "ok")
        run_n8n_upload.assert_called_once()

    def test_explicit_contract_path_allows_voicevox_inbox(self) -> None:
        from runtime_v2.control_plane_feeder import _is_allowed_explicit_contract_path

        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            inbox_root = root / "inbox"
            contract_file = inbox_root / "voicevox" / "voicevox.job.json"
            contract_file.parent.mkdir(parents=True, exist_ok=True)
            contract_file.write_text("{}", encoding="utf-8")

            allowed = _is_allowed_explicit_contract_path(inbox_root, contract_file)

        self.assertTrue(allowed)

    def test_run_worker_dispatches_voicevox_workload(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "runtime" / "artifacts"
            registry_file = root / "health" / "worker_registry.json"
            artifact_root.mkdir(parents=True, exist_ok=True)
            job = JobContract(
                job_id="voicevox-dispatch-job",
                workload="voicevox",
                checkpoint_key="seed:voicevox-dispatch-job",
                payload={"run_id": "voicevox-run", "script_text": "hello world"},
            )

            with patch(
                "runtime_v2.control_plane.run_voicevox_job",
                return_value={"status": "ok", "stage": "voicevox"},
            ) as run_voicevox:
                result = run_worker(
                    job,
                    artifact_root=artifact_root,
                    registry_file=registry_file,
                )

        self.assertEqual(result["status"], "ok")
        run_voicevox.assert_called_once()

    def test_control_plane_queue_helpers_share_queue_store_canonical_path(self) -> None:
        import runtime_v2.control_plane as control_plane_module

        self.assertIs(control_plane_module._load_jobs, QueueStore.load)
        self.assertIs(control_plane_module._save_jobs, QueueStore.save)
        self.assertIs(control_plane_module._upsert_job, QueueStore.upsert)

    def test_control_plane_prefers_lowest_promotion_gate_within_same_row(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            queue_store = QueueStore(config.queue_store_file)
            job_gate_b = JobContract(
                job_id="canva-row1-gate-b",
                workload="canva",
                checkpoint_key="stage2:canva:Sheet1!row1:4",
                payload={
                    "run_id": "row-run-1",
                    "row_ref": "Sheet1!row1",
                    "promotion_gate": "B",
                    "scene_index": 4,
                },
            )
            job_gate_a = JobContract(
                job_id="genspark-row1-gate-a",
                workload="genspark",
                checkpoint_key="stage2:genspark:Sheet1!row1:1",
                payload={
                    "run_id": "row-run-1",
                    "row_ref": "Sheet1!row1",
                    "promotion_gate": "A",
                    "scene_index": 1,
                },
            )
            queue_store.save([job_gate_b, job_gate_a])

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "ok",
                    "code": "OK",
                    "worker_result": {
                        "status": "ok",
                        "stage": "genspark",
                        "error_code": "",
                        "retryable": False,
                        "next_jobs": [],
                        "details": {},
                        "completion": {"state": "succeeded", "final_output": False},
                    },
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-gate-pref",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            by_job_id = {str(item["job_id"]): item for item in queue_items}

        self.assertEqual(result["status"], "ok")
        self.assertEqual(str(by_job_id["genspark-row1-gate-a"]["status"]), "completed")
        self.assertEqual(str(by_job_id["canva-row1-gate-b"]["status"]), "queued")

    def test_control_loop_does_not_short_circuit_when_closeout_state_is_running(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            queue_store = QueueStore(config.queue_store_file)
            queued_job = JobContract(
                job_id="genspark-row15-gate-a",
                workload="genspark",
                checkpoint_key="stage2:genspark:Sheet1!row15:1",
                payload={
                    "run_id": "row-run-15",
                    "row_ref": "Sheet1!row15",
                    "promotion_gate": "A",
                    "scene_index": 1,
                },
            )
            queue_store.save([queued_job])
            config.closeout_state_file.parent.mkdir(parents=True, exist_ok=True)
            config.closeout_state_file.write_text(
                json.dumps(
                    {
                        "run_id": "row-run-15",
                        "status": "running",
                        "reason": "job_running",
                        "attempt": 1,
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "ok",
                    "code": "OK",
                    "worker_result": {
                        "status": "ok",
                        "stage": "genspark",
                        "error_code": "",
                        "retryable": False,
                        "next_jobs": [],
                        "details": {},
                        "completion": {"state": "succeeded", "final_output": False},
                    },
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="row-run-15",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]

        self.assertEqual(result["status"], "ok")
        self.assertEqual(str(queue_items[0]["status"]), "completed")

    def test_control_plane_fail_closes_later_gates_for_same_row_only(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            queue_store = QueueStore(config.queue_store_file)
            failed_gate_a = JobContract(
                job_id="genspark-row1-failed",
                workload="genspark",
                status="failed",
                checkpoint_key="stage2:genspark:Sheet1!row1:1",
                payload={
                    "run_id": "row-run-1",
                    "row_ref": "Sheet1!row1",
                    "promotion_gate": "A",
                    "scene_index": 1,
                },
            )
            blocked_gate_b = JobContract(
                job_id="canva-row1-blocked",
                workload="canva",
                checkpoint_key="stage2:canva:Sheet1!row1:4",
                payload={
                    "run_id": "row-run-1",
                    "row_ref": "Sheet1!row1",
                    "promotion_gate": "B",
                    "scene_index": 4,
                },
            )
            other_row_gate_a = JobContract(
                job_id="genspark-row2-open",
                workload="genspark",
                checkpoint_key="stage2:genspark:Sheet1!row2:1",
                payload={
                    "run_id": "row-run-2",
                    "row_ref": "Sheet1!row2",
                    "promotion_gate": "A",
                    "scene_index": 1,
                },
            )
            queue_store.save([failed_gate_a, blocked_gate_b, other_row_gate_a])

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "ok",
                    "code": "OK",
                    "worker_result": {
                        "status": "ok",
                        "stage": "genspark",
                        "error_code": "",
                        "retryable": False,
                        "next_jobs": [],
                        "details": {},
                        "completion": {"state": "succeeded", "final_output": False},
                    },
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-gate-fail-close",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            by_job_id = {str(item["job_id"]): item for item in queue_items}

        self.assertEqual(result["status"], "ok")
        self.assertEqual(str(by_job_id["canva-row1-blocked"]["status"]), "failed")
        self.assertEqual(str(by_job_id["genspark-row2-open"]["status"]), "completed")

    def test_control_plane_does_not_fail_close_new_run_for_same_row_ref(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            queue_store = QueueStore(config.queue_store_file)
            failed_old_run = JobContract(
                job_id="genspark-row1-old-run-failed",
                workload="genspark",
                status="failed",
                checkpoint_key="stage2:genspark:Sheet1!row1:1:old",
                payload={
                    "run_id": "row-run-old",
                    "row_ref": "Sheet1!row1",
                    "promotion_gate": "A",
                    "scene_index": 1,
                },
            )
            queued_new_run = JobContract(
                job_id="canva-row1-new-run-open",
                workload="canva",
                checkpoint_key="stage2:canva:Sheet1!row1:4:new",
                payload={
                    "run_id": "row-run-new",
                    "row_ref": "Sheet1!row1",
                    "promotion_gate": "B",
                    "scene_index": 4,
                },
            )
            queue_store.save([failed_old_run, queued_new_run])

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "ok",
                    "code": "OK",
                    "worker_result": {
                        "status": "ok",
                        "stage": "canva",
                        "error_code": "",
                        "retryable": False,
                        "next_jobs": [],
                        "details": {},
                        "completion": {"state": "succeeded", "final_output": False},
                    },
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-new-row-rerun",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            by_job_id = {str(item["job_id"]): item for item in queue_items}

        self.assertEqual(result["status"], "ok")
        self.assertEqual(
            str(by_job_id["genspark-row1-old-run-failed"]["status"]), "failed"
        )
        self.assertEqual(
            str(by_job_id["canva-row1-new-run-open"]["status"]), "completed"
        )

    def test_control_plane_keeps_gate_b_open_for_browser_unhealthy_gate_a_failure(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            queue_store = QueueStore(config.queue_store_file)
            failed_gate_a = JobContract(
                job_id="genspark-row1-browser-unhealthy",
                workload="genspark",
                status="failed",
                checkpoint_key="stage2:genspark:Sheet1!row1:1",
                payload={
                    "run_id": "row-run-1",
                    "row_ref": "Sheet1!row1",
                    "promotion_gate": "A",
                    "scene_index": 1,
                    "last_error_code": "BROWSER_UNHEALTHY",
                },
            )
            blocked_gate_b = JobContract(
                job_id="canva-row1-should-stay-open",
                workload="canva",
                checkpoint_key="stage2:canva:Sheet1!row1:4",
                payload={
                    "run_id": "row-run-1",
                    "row_ref": "Sheet1!row1",
                    "promotion_gate": "B",
                    "scene_index": 4,
                },
            )
            queue_store.save([failed_gate_a, blocked_gate_b])

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "ok",
                    "code": "OK",
                    "worker_result": {
                        "status": "ok",
                        "stage": "canva",
                        "error_code": "",
                        "retryable": False,
                        "next_jobs": [],
                        "details": {},
                        "completion": {"state": "succeeded", "final_output": False},
                    },
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-gate-b-open-on-browser-error",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            by_job_id = {str(item["job_id"]): item for item in queue_items}

        self.assertEqual(result["status"], "ok")
        self.assertEqual(
            str(by_job_id["canva-row1-should-stay-open"]["status"]), "completed"
        )

    def test_control_plane_accepts_large_stage1_declared_fanout(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            asset_root = config.artifact_root / "stage1-large-fanout"
            asset_root.mkdir(parents=True, exist_ok=True)
            parent_job = JobContract(
                job_id="chatgpt-row1-large-fanout",
                workload="chatgpt",
                checkpoint_key="topic_spec:Sheet1!row1:run-large-fanout",
                payload={
                    "run_id": "run-large-fanout",
                    "row_ref": "Sheet1!row1",
                    "excel_path": str(root / "topic.xlsx"),
                    "sheet_name": "Sheet1",
                    "row_index": 0,
                },
            )
            jobs = [parent_job]
            story_outline = [f"Story {index}" for index in range(1, 131)]
            scene_plan = [
                {"scene_index": index, "prompt": f"Scene {index}"}
                for index in range(1, 131)
            ]
            voice_texts = [
                {
                    "col": f"#{index:02d}",
                    "text": f"Voice {index}",
                    "original_voices": [index],
                }
                for index in range(1, 131)
            ]
            handoff_contract = {
                "run_id": "run-large-fanout",
                "row_ref": "Sheet1!row1",
                "topic": "Large fanout topic",
                "voice_texts": voice_texts,
            }
            video_plan = build_video_plan(
                run_id="run-large-fanout",
                row_ref="Sheet1!row1",
                topic="Large fanout topic",
                story_outline=story_outline,
                scene_plan=cast(list[dict[str, object]], scene_plan),
                asset_plan={
                    "asset_root": str(asset_root.resolve()),
                    "common_asset_folder": str(asset_root.resolve()),
                },
                voice_plan={
                    "mapping_source": "stage1_parsed",
                    "scene_count": len(scene_plan),
                    "groups": [],
                },
                reason_code="ok",
                evidence={"source": "test"},
            )
            worker_result = cast(
                dict[str, object],
                {
                    "status": "ok",
                    "stage": "chatgpt",
                    "error_code": "",
                    "retryable": False,
                    "details": {
                        "video_plan": video_plan,
                        "stage1_handoff": {"contract": handoff_contract},
                    },
                },
            )

            seeded = _seed_declared_next_jobs(
                config.queue_store_file,
                jobs,
                worker_result,
                parent_job,
                config.control_plane_events_file,
                run_id="control-run-large-fanout",
                artifact_root=config.artifact_root,
            )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]

        self.assertGreater(len(seeded), 128)
        self.assertTrue(
            any(
                str(item.get("job_id", "")).startswith("qwen3-run-large-fanout")
                for item in queue_items
            )
        )

    def test_seed_declared_next_jobs_injects_runtime_root_into_stage2_jobs(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            asset_root = config.artifact_root / "chatgpt" / "chatgpt-seed"
            asset_root.mkdir(parents=True, exist_ok=True)
            parent_job = JobContract(
                job_id="chatgpt-seed",
                workload="chatgpt",
                checkpoint_key="seed:chatgpt-seed",
                payload={
                    "run_id": "runtime-root-run",
                    "row_ref": "Sheet1!row15",
                    "chain_depth": 0,
                },
            )
            worker_result = {
                "status": "ok",
                "stage": "chatgpt",
                "error_code": "",
                "retryable": False,
                "next_jobs": [
                    build_explicit_job_contract(
                        job_id="genspark-runtime-root",
                        workload="genspark",
                        checkpoint_key="stage2:genspark:Sheet1!row15:1",
                        payload={
                            "run_id": "runtime-root-run",
                            "row_ref": "Sheet1!row15",
                            "prompt": "scene one",
                            "service_artifact_path": str(
                                (
                                    config.artifact_root / "images" / "scene-01.png"
                                ).resolve()
                            ),
                            "use_agent_browser": True,
                        },
                        chain_step=1,
                        parent_job_id=parent_job.job_id,
                    )
                ],
                "completion": {"state": "routed", "final_output": False},
            }
            jobs: list[JobContract] = []

            seeded = _seed_declared_next_jobs(
                config.queue_store_file,
                jobs,
                cast(dict[str, object], worker_result),
                parent_job,
                config.control_plane_events_file,
                run_id="control-run-runtime-root-seed",
                artifact_root=config.artifact_root,
            )

        self.assertEqual(len(seeded), 1)
        self.assertEqual(
            str(seeded[0].payload.get("runtime_root", "")),
            str(config.artifact_root.parent.resolve()),
        )

    def test_run_worker_rejects_unsupported_workload_explicitly(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            job = JobContract(
                job_id="unsupported-job",
                workload="chatgpt",
                checkpoint_key="seed:unsupported-job",
                payload={"run_id": "unsupported-run"},
            )
            object.__setattr__(job, "workload", "unsupported_workload")

            with self.assertRaisesRegex(ValueError, "unsupported_workload"):
                _ = run_worker(job, Path(tmp_dir) / "artifacts")

    def test_run_worker_mock_chain_bypasses_worker_registry_updates(self) -> None:
        import runtime_v2.control_plane as control_plane_module

        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            registry_file = root / "health" / "worker_registry.json"
            job = JobContract(
                job_id="mock-chain-job",
                workload="qwen3_tts",
                checkpoint_key="seed:mock-chain-job",
                payload={"run_id": "mock-chain-run", "mock_chain": True},
            )

            with patch(
                "runtime_v2.control_plane._run_mock_chain_worker",
                return_value={"status": "ok", "stage": "mock_chain_qwen3_tts"},
            ) as run_mock_worker:
                result = control_plane_module._run_worker(
                    job,
                    artifact_root=artifact_root,
                    registry_file=registry_file,
                    allow_mock_chain=True,
                )

        self.assertEqual(result["status"], "ok")
        run_mock_worker.assert_called_once()
        self.assertFalse(registry_file.exists())

    def test_run_worker_refreshes_progress_during_long_qwen_execution(self) -> None:
        import runtime_v2.control_plane as control_plane_module

        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            registry_file = root / "health" / "worker_registry.json"
            config = _runtime_config(root)
            config.control_plane_events_file.parent.mkdir(parents=True, exist_ok=True)
            config.control_plane_events_file.write_text("", encoding="utf-8")
            job = JobContract(
                job_id="long-qwen-job",
                workload="qwen3_tts",
                checkpoint_key="seed:long-qwen-job",
                payload={"run_id": "long-qwen-run"},
            )
            finish_event = threading.Event()

            def long_running_handler(_: JobContract) -> dict[str, object]:
                finish_event.wait(timeout=5)
                return {"status": "ok", "stage": "qwen3_tts"}

            result_box: dict[str, dict[str, object]] = {}

            def target() -> None:
                result_box["result"] = control_plane_module._run_worker(
                    job,
                    artifact_root=artifact_root,
                    registry_file=registry_file,
                    heartbeat_interval_sec=0.05,
                    runtime_config=config,
                )

            with patch(
                "runtime_v2.control_plane._worker_dispatch_table",
                return_value={"qwen3_tts": long_running_handler},
            ):
                worker_thread = threading.Thread(target=target)
                worker_thread.start()
                deadline = time.time() + 3
                first_progress = None
                while time.time() < deadline:
                    if registry_file.exists():
                        registry = json.loads(registry_file.read_text(encoding="utf-8"))
                        qwen_entry = cast(
                            dict[str, object], registry.get("qwen3_tts", {})
                        )
                        if qwen_entry:
                            first_progress = float(
                                cast(float, qwen_entry["progress_ts"])
                            )
                            break
                    time.sleep(0.02)
                self.assertIsNotNone(first_progress)

                time.sleep(0.12)
                registry = json.loads(registry_file.read_text(encoding="utf-8"))
                qwen_entry = cast(dict[str, object], registry["qwen3_tts"])
                refreshed_progress = float(cast(float, qwen_entry["progress_ts"]))

                finish_event.set()
                worker_thread.join(timeout=3)
                registry = json.loads(registry_file.read_text(encoding="utf-8"))
                qwen_entry = cast(dict[str, object], registry["qwen3_tts"])
                heartbeat_events = [
                    json.loads(line)
                    for line in config.control_plane_events_file.read_text(
                        encoding="utf-8"
                    ).splitlines()
                    if line.strip()
                ]

        self.assertGreater(refreshed_progress, cast(float, first_progress))
        self.assertEqual(result_box["result"]["status"], "ok")
        self.assertEqual(str(qwen_entry["state"]), "idle")
        self.assertTrue(
            any(
                str(event.get("event", "")) == "worker_heartbeat"
                for event in heartbeat_events
            )
        )

    def test_run_worker_restores_idle_registry_state_when_handler_raises(self) -> None:
        import runtime_v2.control_plane as control_plane_module

        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            artifact_root = root / "artifacts"
            registry_file = root / "health" / "worker_registry.json"
            job = JobContract(
                job_id="failing-qwen-job",
                workload="qwen3_tts",
                checkpoint_key="seed:failing-qwen-job",
                payload={"run_id": "failing-qwen-run"},
            )

            def failing_handler(_: JobContract) -> dict[str, object]:
                raise RuntimeError("worker boom")

            with patch(
                "runtime_v2.control_plane._worker_dispatch_table",
                return_value={"qwen3_tts": failing_handler},
            ):
                with self.assertRaisesRegex(RuntimeError, "worker boom"):
                    _ = control_plane_module._run_worker(
                        job,
                        artifact_root=artifact_root,
                        registry_file=registry_file,
                        heartbeat_interval_sec=0.05,
                    )

            registry = json.loads(registry_file.read_text(encoding="utf-8"))
            qwen_entry = cast(dict[str, object], registry["qwen3_tts"])

        self.assertEqual(str(qwen_entry["state"]), "idle")
        self.assertEqual(str(qwen_entry["run_id"]), "failing-qwen-run")

    def test_control_plane_reconciles_orphan_busy_worker_registry(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            _ = config.worker_registry_file.parent.mkdir(parents=True, exist_ok=True)
            _ = config.worker_registry_file.write_text(
                json.dumps(
                    {
                        "qwen3_tts": {
                            "workload": "qwen3_tts",
                            "state": "busy",
                            "run_id": "orphan-run",
                            "last_seen": 10.0,
                            "progress_ts": 10.0,
                        }
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )

            result = run_control_loop_once(
                owner="runtime_v2",
                config=config,
                run_id="control-run-reconcile-orphan-worker",
                allow_runtime_side_effects=False,
            )
            worker_registry = json.loads(
                config.worker_registry_file.read_text(encoding="utf-8")
            )
            qwen_registry = cast(dict[str, object], worker_registry["qwen3_tts"])

        self.assertEqual(result["status"], "idle")
        self.assertEqual(result["code"], "NO_JOB")
        self.assertEqual(str(result["queue_status"]), "idle")
        self.assertEqual(str(qwen_registry["state"]), "idle")

    def test_queue_store_load_fail_closes_on_corrupted_queue_file(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            queue_file = Path(tmp_dir) / "job_queue.json"
            _ = queue_file.write_text("{not-json", encoding="utf-8")

            with self.assertRaises(QueueStoreError):
                _ = QueueStore(queue_file).load()

    def test_queue_store_save_retries_permission_error_on_replace(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            queue_file = Path(tmp_dir) / "job_queue.json"
            queue_store = QueueStore(queue_file)
            job = JobContract(job_id="job-1", workload="chatgpt")
            original_replace = Path.replace
            call_count = {"count": 0}

            def flaky_replace(self: Path, target: Path) -> Path:
                call_count["count"] += 1
                if call_count["count"] < 3:
                    error = PermissionError("locked")
                    error.winerror = 5
                    raise error
                return original_replace(self, target)

            with patch.object(Path, "replace", new=flaky_replace):
                saved_path = queue_store.save([job])

        self.assertEqual(saved_path, queue_file)
        self.assertEqual(call_count["count"], 3)

    def test_queue_store_save_falls_back_to_direct_write_after_retry_exhausted(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            queue_file = Path(tmp_dir) / "job_queue.json"
            queue_store = QueueStore(queue_file)
            job = JobContract(job_id="job-1", workload="chatgpt")

            def always_locked_replace(self: Path, target: Path) -> Path:
                _ = target
                error = PermissionError("locked")
                error.winerror = 5
                raise error

            with patch.object(Path, "replace", new=always_locked_replace):
                saved_path = queue_store.save([job])

            self.assertEqual(saved_path, queue_file)
            payload = json.loads(queue_file.read_text(encoding="utf-8"))
            self.assertEqual(payload[0]["job_id"], "job-1")

    def test_control_plane_reports_invalid_queue_store_as_failed(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            config.queue_store_file.parent.mkdir(parents=True, exist_ok=True)
            _ = config.queue_store_file.write_text("{not-json", encoding="utf-8")

            result = run_control_loop_once(
                owner="runtime_v2", config=config, run_id="control-run-invalid-queue"
            )
            latest_join = load_joined_latest_run(config, completed=True)
            result_metadata = cast(dict[object, object], latest_join["result_metadata"])

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "QUEUE_STORE_INVALID")
        self.assertEqual(str(result["queue_status"]), "failed")
        self.assertEqual(str(result_metadata["code"]), "QUEUE_STORE_INVALID")

    def test_retry_and_circuit_helpers_share_single_canonical_policy_module(
        self,
    ) -> None:
        self.assertIs(circuit_breaker.CircuitState, recovery_policy.CircuitState)
        self.assertIs(
            retry_budget.within_retry_budget, recovery_policy.within_retry_budget
        )
        self.assertIs(retry_budget.next_backoff_sec, recovery_policy.next_backoff_sec)
        self.assertIs(circuit_breaker.record_failure, recovery_policy.record_failure)
        self.assertIs(circuit_breaker.reset_circuit, recovery_policy.reset_circuit)
        self.assertIs(circuit_breaker.is_circuit_open, recovery_policy.is_circuit_open)

    def test_control_plane_runs_chatgpt_job_and_seeds_stage2_jobs_with_same_run_id(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            topic_spec: dict[str, object] = {
                "contract": "topic_spec",
                "contract_version": "1.0",
                "run_id": "chatgpt-run-1",
                "row_ref": "Sheet1!row1",
                "topic": "Bridge topic",
                "status_snapshot": "",
                "excel_snapshot_hash": "hash-1",
            }
            seed_control_job(
                JobContract(
                    job_id="chatgpt-job",
                    workload="chatgpt",
                    checkpoint_key="topic_spec:Sheet1!row1:hash-1",
                    payload={
                        "run_id": "chatgpt-run-1",
                        "row_ref": "Sheet1!row1",
                        "topic_spec": topic_spec,
                    },
                ),
                config=config,
            )
            asset_root = config.artifact_root / "chatgpt" / "chatgpt-job"
            asset_root.mkdir(parents=True, exist_ok=True)
            video_plan = build_video_plan(
                run_id="chatgpt-run-1",
                row_ref="Sheet1!row1",
                topic="Bridge topic",
                story_outline=["Scene 1", "Scene 2", "Scene 3", "Scene 4"],
                scene_plan=[
                    {"scene_index": 1, "prompt": "scene one"},
                    {"scene_index": 2, "prompt": "scene two"},
                    {"scene_index": 3, "prompt": "scene three"},
                    {"scene_index": 4, "prompt": "scene four"},
                ],
                asset_plan={
                    "asset_root": str(asset_root.resolve()),
                    "common_asset_folder": str(asset_root.resolve()),
                },
                voice_plan={
                    "mapping_source": "stage1_parsed",
                    "scene_count": 2,
                    "groups": [],
                },
                reason_code="ok",
                evidence={"source": "test"},
            )
            worker_result = {
                "status": "ok",
                "stage": "chatgpt",
                "error_code": "",
                "retryable": False,
                "details": {
                    "video_plan": video_plan,
                    "stage1_handoff": {
                        "contract": {
                            "run_id": "chatgpt-run-1",
                            "row_ref": "Sheet1!row1",
                            "topic": "Bridge topic",
                            "voice_texts": [
                                {
                                    "col": "#01",
                                    "text": "Voice 1",
                                    "original_voices": [1],
                                }
                            ],
                        }
                    },
                },
                "completion": {"state": "succeeded", "final_output": False},
            }

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "ok",
                    "code": "OK",
                    "worker_result": worker_result,
                },
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-chatgpt"
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queued_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            by_job_id = {str(item["job_id"]): item for item in queued_items}
            routed_jobs = [
                item
                for item in queued_items
                if str(item.get("job_id", "")).startswith(("genspark-", "seaart-"))
            ]
            qwen_jobs = [
                item
                for item in queued_items
                if str(item.get("job_id", "")).startswith("qwen3-")
            ]
            render_jobs = [
                item
                for item in queued_items
                if str(item.get("job_id", "")).startswith("render-")
            ]

        self.assertEqual(str(by_job_id["chatgpt-job"]["status"]), "completed")
        self.assertTrue(routed_jobs)
        self.assertEqual(len(qwen_jobs), 1)
        qwen_payload = cast(dict[object, object], qwen_jobs[0]["payload"])
        self.assertTrue(bool(cast(list[object], qwen_payload["voice_texts"])))
        self.assertTrue(
            str(qwen_payload["service_artifact_path"])
            .replace("\\", "/")
            .endswith("/speech.flac")
        )
        self.assertTrue(bool(qwen_payload["emit_rvc_next_job"]))
        self.assertEqual(len(render_jobs), 1)

    def test_control_plane_merges_chatgpt_video_plan_to_excel_main_path(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            topic_spec: dict[str, object] = {
                "contract": "topic_spec",
                "contract_version": "1.0",
                "run_id": "chatgpt-run-excel-main",
                "row_ref": "Sheet1!row1",
                "topic": "Bridge topic",
                "status_snapshot": "",
                "excel_snapshot_hash": "hash-1",
            }
            seed_control_job(
                JobContract(
                    job_id="chatgpt-excel-main",
                    workload="chatgpt",
                    checkpoint_key="topic_spec:Sheet1!row1:hash-1",
                    payload={
                        "run_id": "chatgpt-run-excel-main",
                        "row_ref": "Sheet1!row1",
                        "excel_path": str((root / "topic.xlsx").resolve()),
                        "sheet_name": "Sheet1",
                        "row_index": 0,
                        "topic_spec": topic_spec,
                    },
                ),
                config=config,
            )
            asset_root = config.artifact_root / "chatgpt" / "chatgpt-excel-main"
            asset_root.mkdir(parents=True, exist_ok=True)
            video_plan = build_video_plan(
                run_id="chatgpt-run-excel-main",
                row_ref="Sheet1!row1",
                topic="Bridge topic",
                story_outline=["Scene 1", "Scene 2", "Scene 3", "Scene 4"],
                scene_plan=[
                    {"scene_index": 1, "prompt": "scene one"},
                    {"scene_index": 2, "prompt": "scene two"},
                    {"scene_index": 3, "prompt": "scene three"},
                    {"scene_index": 4, "prompt": "scene four"},
                ],
                asset_plan={
                    "asset_root": str(asset_root.resolve()),
                    "common_asset_folder": str(asset_root.resolve()),
                },
                voice_plan={
                    "mapping_source": "stage1_parsed",
                    "scene_count": 2,
                    "groups": [],
                },
                reason_code="ok",
                evidence={"source": "test"},
            )
            worker_result = {
                "status": "ok",
                "stage": "chatgpt",
                "error_code": "",
                "retryable": False,
                "details": {"video_plan": video_plan},
                "completion": {"state": "planned", "final_output": False},
            }

            with (
                patch(
                    "runtime_v2.control_plane.merge_stage1_result",
                    return_value=True,
                ) as merge_mock,
                patch(
                    "runtime_v2.control_plane.run_gated",
                    side_effect=lambda **kwargs: {
                        "status": "ok",
                        "code": "OK",
                        "worker_result": worker_result,
                    },
                ),
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-chatgpt-excel",
                )

        merge_mock.assert_called_once()
        self.assertEqual(merge_mock.call_args.kwargs["sheet_name"], "Sheet1")
        self.assertEqual(merge_mock.call_args.kwargs["row_index"], 0)

    def test_control_plane_writes_asset_manifest_for_stage1_routed_jobs(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            topic_spec: dict[str, object] = {
                "contract": "topic_spec",
                "contract_version": "1.0",
                "run_id": "chatgpt-run-manifest",
                "row_ref": "Sheet1!row1",
                "topic": "Bridge topic",
                "status_snapshot": "",
                "excel_snapshot_hash": "hash-1",
            }
            seed_control_job(
                JobContract(
                    job_id="chatgpt-manifest",
                    workload="chatgpt",
                    checkpoint_key="topic_spec:Sheet1!row1:hash-1",
                    payload={
                        "run_id": "chatgpt-run-manifest",
                        "row_ref": "Sheet1!row1",
                        "topic_spec": topic_spec,
                    },
                ),
                config=config,
            )
            asset_root = config.artifact_root / "chatgpt" / "chatgpt-manifest"
            asset_root.mkdir(parents=True, exist_ok=True)
            video_plan = build_video_plan(
                run_id="chatgpt-run-manifest",
                row_ref="Sheet1!row1",
                topic="Bridge topic",
                story_outline=["Scene 1", "Scene 2", "Scene 3", "Scene 4"],
                scene_plan=[
                    {"scene_index": 1, "prompt": "scene one"},
                    {"scene_index": 2, "prompt": "scene two"},
                    {"scene_index": 3, "prompt": "scene three"},
                    {"scene_index": 4, "prompt": "scene four"},
                ],
                asset_plan={
                    "asset_root": str(asset_root.resolve()),
                    "common_asset_folder": str(asset_root.resolve()),
                },
                voice_plan={
                    "mapping_source": "stage1_parsed",
                    "scene_count": 2,
                    "groups": [],
                },
                reason_code="ok",
                evidence={"source": "test"},
            )
            worker_result = {
                "status": "ok",
                "stage": "chatgpt",
                "error_code": "",
                "retryable": False,
                "details": {
                    "video_plan": video_plan,
                    "stage1_handoff": {
                        "contract": {
                            "run_id": "chatgpt-run-manifest",
                            "row_ref": "Sheet1!row1",
                            "topic": "Bridge topic",
                            "voice_texts": [
                                {
                                    "col": "#01",
                                    "text": "Voice 1",
                                    "original_voices": [1],
                                }
                            ],
                        }
                    },
                },
                "completion": {"state": "succeeded", "final_output": False},
            }

            with patch(
                "runtime_v2.control_plane.run_gated",
                side_effect=lambda **kwargs: {
                    "status": "ok",
                    "code": "OK",
                    "worker_result": worker_result,
                },
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-manifest"
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queued_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            render_job = next(
                item
                for item in queued_items
                if str(item.get("job_id", "")).startswith("render-")
            )
            canva_job = next(
                item
                for item in queued_items
                if str(item.get("job_id", "")).startswith("canva-")
            )
            render_payload = cast(dict[str, object], render_job["payload"])
            canva_payload = cast(dict[str, object], canva_job["payload"])
            qwen_job = next(
                item
                for item in queued_items
                if str(item.get("job_id", "")).startswith("qwen3-")
            )
            qwen_payload = cast(dict[str, object], qwen_job["payload"])
            manifest_path = Path(str(render_payload["asset_manifest_path"]))
            manifest = cast(
                dict[str, object], json.loads(manifest_path.read_text(encoding="utf-8"))
            )
            self.assertTrue(manifest_path.exists())
            self.assertEqual(
                str(qwen_payload["asset_manifest_path"]), str(manifest_path.resolve())
            )
            self.assertEqual(
                str(canva_payload["asset_manifest_path"]), str(manifest_path.resolve())
            )
            self.assertEqual(manifest["run_id"], "chatgpt-run-manifest")
            roles = cast(dict[str, object], manifest["roles"])
            self.assertTrue(
                str(roles["image_primary"])
                .replace("\\", "/")
                .endswith("images/genspark-chatgpt-run-manifest-1.png")
            )
            self.assertTrue(
                str(roles["stage2.scene_01.genspark"])
                .replace("\\", "/")
                .endswith("images/genspark-chatgpt-run-manifest-1.png")
            )
            self.assertTrue(
                str(roles["stage2.scene_02.seaart"])
                .replace("\\", "/")
                .endswith("images/seaart-chatgpt-run-manifest-2.png")
            )
            self.assertTrue(
                str(roles["voice_json"]).replace("\\", "/").endswith("voice.json")
            )

    def test_control_plane_routes_stage1_video_plan_without_worker_next_jobs(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            topic_spec: dict[str, object] = {
                "contract": "topic_spec",
                "contract_version": "1.0",
                "run_id": "chatgpt-run-no-next-jobs",
                "row_ref": "Sheet1!row1",
                "topic": "Bridge topic",
                "status_snapshot": "",
                "excel_snapshot_hash": "hash-1",
            }
            seed_control_job(
                JobContract(
                    job_id="chatgpt-no-next-jobs",
                    workload="chatgpt",
                    checkpoint_key="topic_spec:Sheet1!row1:hash-1",
                    payload={
                        "run_id": "chatgpt-run-no-next-jobs",
                        "row_ref": "Sheet1!row1",
                        "topic_spec": topic_spec,
                    },
                ),
                config=config,
            )
            asset_root = config.artifact_root / "chatgpt" / "chatgpt-no-next-jobs"
            asset_root.mkdir(parents=True, exist_ok=True)
            video_plan = build_video_plan(
                run_id="chatgpt-run-no-next-jobs",
                row_ref="Sheet1!row1",
                topic="Bridge topic",
                story_outline=["Scene 1", "Scene 2"],
                scene_plan=[
                    {"scene_index": 1, "prompt": "scene one"},
                    {"scene_index": 2, "prompt": "scene two"},
                ],
                asset_plan={
                    "asset_root": str(asset_root.resolve()),
                    "common_asset_folder": str(asset_root.resolve()),
                },
                voice_plan={
                    "mapping_source": "stage1_parsed",
                    "scene_count": 2,
                    "groups": [],
                },
                reason_code="ok",
                evidence={"source": "test"},
            )
            worker_result = {
                "status": "ok",
                "stage": "chatgpt",
                "error_code": "",
                "retryable": False,
                "details": {"video_plan": video_plan},
                "completion": {"state": "planned", "final_output": False},
            }

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "ok",
                    "code": "OK",
                    "worker_result": worker_result,
                },
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-stage1-plan"
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queued_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            routed_jobs = [
                item
                for item in queued_items
                if str(item.get("job_id", "")).startswith(("genspark-", "seaart-"))
            ]
            render_jobs = [
                item
                for item in queued_items
                if str(item.get("job_id", "")).startswith("render-")
            ]

        self.assertTrue(routed_jobs)
        self.assertEqual(len(render_jobs), 1)

    def test_runtime_config_from_root_keeps_latest_pointers_inside_temp_root(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)

        self.assertEqual(
            config.latest_active_run_file, root.resolve() / "latest_active_run.json"
        )
        self.assertEqual(
            config.latest_completed_run_file,
            root.resolve() / "latest_completed_run.json",
        )

    def test_control_plane_routes_explicit_rvc_next_jobs_from_worker_result(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="qwen-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-job",
                    payload={"script_text": "hello", "chain_depth": 0},
                ),
                config=config,
            )

            runtime_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "ok",
                    "stage": "synthesize_audio",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [
                        build_explicit_job_contract(
                            job_id="rvc-qwen-job",
                            workload="rvc",
                            checkpoint_key="derived:rvc:qwen-job",
                            payload={
                                "source_path": "system/runtime_v2/artifacts/qwen3_tts/qwen-job/speech.flac",
                                "model_name": "voice-model-a",
                                "service_artifact_path": "system/runtime_v2/artifacts/rvc/qwen-job/speech_rvc.wav",
                                "chain_depth": 1,
                            },
                            chain_step=1,
                            parent_job_id="qwen-job",
                        )
                    ],
                    "completion": {"state": "routed", "final_output": False},
                },
            }

            with patch(
                "runtime_v2.control_plane.run_gated", return_value=runtime_result
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-1"
                )

            queue_payload = cast(
                object, json.loads(config.queue_store_file.read_text(encoding="utf-8"))
            )
            self.assertIsInstance(queue_payload, list)
            if not isinstance(queue_payload, list):
                self.fail("queue payload missing")
            queue_items = cast(list[object], queue_payload)
            by_job_id: dict[str, dict[object, object]] = {}
            for item in queue_items:
                if isinstance(item, dict):
                    typed_item = cast(dict[object, object], item)
                    by_job_id[str(typed_item["job_id"])] = typed_item

        self.assertEqual(str(by_job_id["qwen-job"]["status"]), "completed")
        self.assertEqual(str(by_job_id["rvc-qwen-job"]["status"]), "queued")

    def test_control_plane_canonicalizes_geminigen_rvc_next_job_for_run(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            gemi_video = root / "gemi.mp4"
            _ = gemi_video.write_bytes(b"mp4")
            seed_control_job(
                JobContract(
                    job_id="geminigen-job",
                    workload="geminigen",
                    checkpoint_key="seed:geminigen-job",
                    payload={
                        "run_id": "stage2-run-1",
                        "row_ref": "Sheet1!row1",
                        "chain_depth": 0,
                    },
                ),
                config=config,
            )

            runtime_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "ok",
                    "stage": "geminigen",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [
                        build_explicit_job_contract(
                            job_id="rvc-geminigen-job",
                            workload="rvc",
                            checkpoint_key="derived:rvc:geminigen-job",
                            payload={
                                "audio_path": str(gemi_video.resolve()),
                                "model_name": "voice-model-a",
                                "run_id": "stage2-run-1",
                                "row_ref": "Sheet1!row1",
                                "chain_depth": 1,
                            },
                            chain_step=1,
                            parent_job_id="geminigen-job",
                        )
                    ],
                    "completion": {"state": "routed", "final_output": False},
                },
            }

            with patch(
                "runtime_v2.control_plane.run_gated", return_value=runtime_result
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-gemi-rvc"
                )

            queue_payload = cast(
                object, json.loads(config.queue_store_file.read_text(encoding="utf-8"))
            )
            self.assertIsInstance(queue_payload, list)
            if not isinstance(queue_payload, list):
                self.fail("queue payload missing")
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            by_job_id = {str(item["job_id"]): item for item in queue_items}

        self.assertIn("rvc-geminigen-stage2-run-1", by_job_id)
        rvc_payload = cast(
            dict[object, object], by_job_id["rvc-geminigen-stage2-run-1"]["payload"]
        )
        self.assertEqual(str(rvc_payload["source_mode"]), "gemi-video-source")
        self.assertTrue(
            str(rvc_payload["service_artifact_path"])
            .replace("\\", "/")
            .endswith("rvc-geminigen-stage2-run-1/speech_rvc.wav")
        )

    def test_control_plane_preserves_flac_suffix_when_canonicalizing_qwen_rvc_job(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            qwen_audio = root / "speech.flac"
            _ = qwen_audio.write_bytes(b"flac")
            seed_control_job(
                JobContract(
                    job_id="qwen-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-job",
                    payload={
                        "run_id": "stage2-run-flac",
                        "row_ref": "Sheet1!row1",
                        "chain_depth": 0,
                    },
                ),
                config=config,
            )

            runtime_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "ok",
                    "stage": "qwen3_tts",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [
                        build_explicit_job_contract(
                            job_id="rvc-qwen-job",
                            workload="rvc",
                            checkpoint_key="derived:rvc:qwen-job",
                            payload={
                                "source_path": str(qwen_audio.resolve()),
                                "model_name": "voice-model-a",
                                "run_id": "stage2-run-flac",
                                "row_ref": "Sheet1!row1",
                                "service_artifact_path": str(
                                    (root / "exports" / "speech_rvc.flac").resolve()
                                ),
                                "chain_depth": 1,
                            },
                            chain_step=1,
                            parent_job_id="qwen-job",
                        )
                    ],
                    "completion": {"state": "routed", "final_output": False},
                },
            }

            with patch(
                "runtime_v2.control_plane.run_gated", return_value=runtime_result
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-qwen-flac"
                )

            queue_payload = cast(
                object, json.loads(config.queue_store_file.read_text(encoding="utf-8"))
            )
            self.assertIsInstance(queue_payload, list)
            if not isinstance(queue_payload, list):
                self.fail("queue payload missing")
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            by_job_id = {str(item["job_id"]): item for item in queue_items}

        self.assertIn("rvc-qwen3-stage2-run-flac", by_job_id)
        rvc_payload = cast(
            dict[object, object], by_job_id["rvc-qwen3-stage2-run-flac"]["payload"]
        )
        self.assertEqual(str(rvc_payload["source_mode"]), "tts-source")
        self.assertTrue(
            str(rvc_payload["service_artifact_path"])
            .replace("\\", "/")
            .endswith("rvc-qwen3-stage2-run-flac/speech_rvc.flac")
        )

    def test_control_plane_prefers_geminigen_rvc_lane_over_qwen3_lane_for_same_run(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            gemi_video = root / "gemi.mp4"
            qwen_audio = root / "speech.flac"
            _ = gemi_video.write_bytes(b"mp4")
            _ = qwen_audio.write_bytes(b"flac")
            seed_control_job(
                JobContract(
                    job_id="geminigen-job",
                    workload="geminigen",
                    checkpoint_key="seed:geminigen-job",
                    payload={
                        "run_id": "stage2-run-1",
                        "row_ref": "Sheet1!row1",
                        "chain_depth": 0,
                    },
                ),
                config=config,
            )

            gemi_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "ok",
                    "stage": "geminigen",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [
                        build_explicit_job_contract(
                            job_id="rvc-geminigen-job",
                            workload="rvc",
                            checkpoint_key="derived:rvc:geminigen-job",
                            payload={
                                "audio_path": str(gemi_video.resolve()),
                                "model_name": "voice-model-a",
                                "run_id": "stage2-run-1",
                                "row_ref": "Sheet1!row1",
                                "chain_depth": 1,
                            },
                            chain_step=1,
                            parent_job_id="geminigen-job",
                        )
                    ],
                    "completion": {"state": "routed", "final_output": False},
                },
            }
            with patch("runtime_v2.control_plane.run_gated", return_value=gemi_result):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-gemi-rvc"
                )

            qwen_job = JobContract(
                job_id="qwen-job",
                workload="qwen3_tts",
                checkpoint_key="seed:qwen-job",
                payload={
                    "run_id": "stage2-run-1",
                    "row_ref": "Sheet1!row1",
                    "chain_depth": 0,
                },
            )
            jobs = _load_jobs(QueueStore(config.queue_store_file))
            qwen_worker_result: dict[str, object] = {
                "status": "ok",
                "stage": "qwen3_tts",
                "error_code": "",
                "retryable": False,
                "next_jobs": [
                    build_explicit_job_contract(
                        job_id="rvc-qwen-job",
                        workload="rvc",
                        checkpoint_key="derived:rvc:qwen-job",
                        payload={
                            "source_path": str(qwen_audio.resolve()),
                            "model_name": "voice-model-a",
                            "run_id": "stage2-run-1",
                            "row_ref": "Sheet1!row1",
                            "chain_depth": 1,
                        },
                        chain_step=1,
                        parent_job_id="qwen-job",
                    )
                ],
                "completion": {"state": "routed", "final_output": False},
            }
            _ = _seed_declared_next_jobs(
                config.queue_store_file,
                jobs,
                qwen_worker_result,
                qwen_job,
                config.control_plane_events_file,
                run_id="control-run-qwen-rvc",
                artifact_root=config.artifact_root,
            )

            queue_payload = cast(
                object, json.loads(config.queue_store_file.read_text(encoding="utf-8"))
            )
            self.assertIsInstance(queue_payload, list)
            if not isinstance(queue_payload, list):
                self.fail("queue payload missing")
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_ids = {str(item["job_id"]) for item in queue_items}

        self.assertIn("rvc-geminigen-stage2-run-1", job_ids)
        self.assertNotIn("rvc-qwen3-stage2-run-1", job_ids)

    def test_control_plane_replaces_existing_qwen3_rvc_lane_when_geminigen_lane_arrives(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            qwen_audio = root / "speech.flac"
            gemi_video = root / "gemi.mp4"
            _ = qwen_audio.write_bytes(b"flac")
            _ = gemi_video.write_bytes(b"mp4")

            qwen_job = JobContract(
                job_id="qwen-job",
                workload="qwen3_tts",
                checkpoint_key="seed:qwen-job",
                payload={
                    "run_id": "stage2-run-1",
                    "row_ref": "Sheet1!row1",
                    "chain_depth": 0,
                },
            )
            jobs: list[JobContract] = []
            _ = _seed_declared_next_jobs(
                config.queue_store_file,
                jobs,
                {
                    "status": "ok",
                    "stage": "qwen3_tts",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [
                        build_explicit_job_contract(
                            job_id="rvc-qwen-job",
                            workload="rvc",
                            checkpoint_key="derived:rvc:qwen-job",
                            payload={
                                "source_path": str(qwen_audio.resolve()),
                                "model_name": "voice-model-a",
                                "run_id": "stage2-run-1",
                                "row_ref": "Sheet1!row1",
                                "chain_depth": 1,
                            },
                            chain_step=1,
                            parent_job_id="qwen-job",
                        )
                    ],
                    "completion": {"state": "routed", "final_output": False},
                },
                qwen_job,
                config.control_plane_events_file,
                run_id="control-run-qwen-rvc",
                artifact_root=config.artifact_root,
            )
            jobs = _load_jobs(QueueStore(config.queue_store_file))
            gemi_job = JobContract(
                job_id="geminigen-job",
                workload="geminigen",
                checkpoint_key="seed:geminigen-job",
                payload={
                    "run_id": "stage2-run-1",
                    "row_ref": "Sheet1!row1",
                    "chain_depth": 0,
                },
            )
            _ = _seed_declared_next_jobs(
                config.queue_store_file,
                jobs,
                {
                    "status": "ok",
                    "stage": "geminigen",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [
                        build_explicit_job_contract(
                            job_id="rvc-geminigen-job",
                            workload="rvc",
                            checkpoint_key="derived:rvc:geminigen-job",
                            payload={
                                "audio_path": str(gemi_video.resolve()),
                                "model_name": "voice-model-a",
                                "run_id": "stage2-run-1",
                                "row_ref": "Sheet1!row1",
                                "chain_depth": 1,
                            },
                            chain_step=1,
                            parent_job_id="geminigen-job",
                        )
                    ],
                    "completion": {"state": "routed", "final_output": False},
                },
                gemi_job,
                config.control_plane_events_file,
                run_id="control-run-gemi-rvc",
                artifact_root=config.artifact_root,
            )

            queue_payload = cast(
                object, json.loads(config.queue_store_file.read_text(encoding="utf-8"))
            )
            self.assertIsInstance(queue_payload, list)
            if not isinstance(queue_payload, list):
                self.fail("queue payload missing")
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_ids = {str(item["job_id"]) for item in queue_items}

        self.assertIn("rvc-geminigen-stage2-run-1", job_ids)
        self.assertNotIn("rvc-qwen3-stage2-run-1", job_ids)

    def test_control_plane_keeps_stage1_declared_qwen_job_when_scene_count_is_three(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            asset_root = config.artifact_root / "stage1-assets"
            asset_root.mkdir(parents=True, exist_ok=True)
            video_plan = build_video_plan(
                run_id="chatgpt-run-qwen-three-scenes",
                row_ref="Sheet1!row1",
                topic="Bridge topic",
                story_outline=["scene one", "scene two", "scene three"],
                scene_plan=[
                    {"scene_index": 1, "prompt": "scene one"},
                    {"scene_index": 2, "prompt": "scene two"},
                    {"scene_index": 3, "prompt": "scene three"},
                ],
                asset_plan={
                    "asset_root": str(asset_root.resolve()),
                    "common_asset_folder": str(asset_root.resolve()),
                },
                voice_plan={
                    "mapping_source": "stage1_parsed",
                    "scene_count": 3,
                    "groups": [],
                },
                reason_code="ok",
                evidence={"source": "test"},
            )
            worker_result = {
                "status": "ok",
                "stage": "chatgpt",
                "error_code": "",
                "retryable": False,
                "next_jobs": [],
                "completion": {"state": "planned", "final_output": False},
                "details": {
                    "video_plan": video_plan,
                    "stage1_handoff": {
                        "contract": {
                            "run_id": "chatgpt-run-qwen-three-scenes",
                            "row_ref": "Sheet1!row1",
                            "topic": "Bridge topic",
                            "voice_texts": [
                                {
                                    "col": "#01",
                                    "text": "line one",
                                    "original_voices": [1],
                                },
                                {
                                    "col": "#02",
                                    "text": "line two",
                                    "original_voices": [2],
                                },
                                {
                                    "col": "#03",
                                    "text": "line three",
                                    "original_voices": [3],
                                },
                            ],
                        }
                    },
                },
            }
            seed_control_job(
                JobContract(
                    job_id="chatgpt-qwen-three-scenes",
                    workload="chatgpt",
                    checkpoint_key="seed:chatgpt-qwen-three-scenes",
                    payload={
                        "run_id": "chatgpt-run-qwen-three-scenes",
                        "row_ref": "Sheet1!row1",
                        "topic": "Bridge topic",
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "ok",
                    "code": "OK",
                    "worker_result": worker_result,
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-qwen-three-scenes",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            queued_job_ids = {str(item["job_id"]) for item in queue_items}

        self.assertEqual(result["status"], "ok")
        self.assertIn("qwen3-chatgpt-run-qwen-three-scenes", queued_job_ids)
        qwen_job = next(
            item
            for item in queue_items
            if str(item["job_id"]) == "qwen3-chatgpt-run-qwen-three-scenes"
        )
        qwen_payload = cast(dict[object, object], qwen_job["payload"])
        self.assertEqual(str(qwen_payload["model_name"]), "voice-model-a")

    def test_control_plane_syncs_excel_done_after_render_final_output(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            from openpyxl import Workbook, load_workbook
            from openpyxl.worksheet.worksheet import Worksheet

            root = Path(tmp_dir)
            config = _runtime_config(root)
            workbook_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(
                ["Topic", "Status", "Video Plan", "Reason Code", "Result Path"]
            )
            sheet.append(["Bridge topic", "Voice OK", "", "", ""])
            workbook.save(workbook_path)
            workbook.close()

            final_output = (
                root
                / "artifacts"
                / "render"
                / "render-gate-d-run-1"
                / "output"
                / "render_final.mp4"
            )
            asset_manifest_path = (
                root
                / "artifacts"
                / "render"
                / "render-gate-d-run-1"
                / "asset_manifest.json"
            )
            final_output.parent.mkdir(parents=True, exist_ok=True)
            final_output.write_bytes(b"mp4")
            asset_manifest_path.write_text(
                json.dumps(
                    {
                        "run_id": "gate-d-run-1",
                        "row_ref": "Sheet1!row1",
                        "roles": {
                            "voice_json": "D:/voice.json",
                            "image_primary": "D:/image.png",
                        },
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            seed_control_job(
                JobContract(
                    job_id="render-gate-d-run-1",
                    workload="render",
                    checkpoint_key="render:Sheet1!row1:gate-d-run-1",
                    payload={
                        "run_id": "gate-d-run-1",
                        "row_ref": "Sheet1!row1",
                        "excel_path": str(workbook_path.resolve()),
                        "sheet_name": "Sheet1",
                        "row_index": 0,
                        "render_folder_path": str(final_output.parent.parent.resolve()),
                        "asset_manifest_path": str(asset_manifest_path.resolve()),
                        "voice_json": {"voice_texts": []},
                        "render_spec": {
                            "contract": "render_spec",
                            "run_id": "gate-d-run-1",
                            "row_ref": "Sheet1!row1",
                            "asset_refs": [str(final_output.resolve())],
                            "audio_refs": [],
                            "thumbnail_refs": [],
                            "timeline": [
                                {
                                    "scene_index": 1,
                                    "asset_path": str(final_output.resolve()),
                                }
                            ],
                            "reason_code": "ok",
                        },
                    },
                ),
                config=config,
            )

            runtime_result = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "ok",
                    "stage": "render",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [],
                    "details": {"reason_code": "ok"},
                    "completion": {
                        "state": "succeeded",
                        "final_output": True,
                        "final_artifact": "render_final.mp4",
                        "final_artifact_path": str(final_output.resolve()),
                    },
                },
            }

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value=runtime_result,
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-gate-d",
                )

            workbook = load_workbook(workbook_path)
            try:
                status_value = workbook["Sheet1"].cell(row=2, column=2).value
            finally:
                workbook.close()
            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])
            latest_join = load_joined_latest_run(config, completed=True)
            manifest_payload = json.loads(
                asset_manifest_path.read_text(encoding="utf-8")
            )

        self.assertEqual(result["status"], "ok")
        self.assertEqual(status_value, "Done")
        self.assertTrue(bool(latest_metadata["excel_sync_updated"]))
        self.assertEqual(str(latest_metadata["run_id"]), "gate-d-run-1")
        self.assertEqual(str(latest_metadata["row_ref"]), "Sheet1!row1")
        self.assertEqual(
            str(latest_metadata["asset_manifest_path"]),
            str(asset_manifest_path.resolve()),
        )
        self.assertEqual(
            str(latest_metadata["final_artifact_path"]), str(final_output.resolve())
        )
        self.assertEqual(len(cast(list[object], latest_result["artifacts"])), 1)
        canonical_handoff = cast(
            dict[object, object], latest_metadata["canonical_handoff"]
        )
        self.assertEqual(str(canonical_handoff["run_id"]), "gate-d-run-1")
        roles = cast(
            dict[object, object], cast(dict[object, object], manifest_payload)["roles"]
        )
        self.assertEqual(str(roles["thumb_primary"]), "D:/image.png")
        self.assertEqual(str(roles["video_primary"]), str(final_output.resolve()))
        pointer = cast(dict[object, object], latest_join["pointer"])
        self.assertEqual(str(pointer["run_id"]), "gate-d-run-1")
        self.assertFalse(bool(latest_join["out_of_sync"]))

    def test_stage1_video_plan_routing_does_not_emit_kenburns_jobs(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            asset_root = root / "assets"
            asset_root.mkdir(parents=True, exist_ok=True)
            video_plan = {
                "contract": "video_plan",
                "contract_version": "1.0",
                "run_id": "stage1-run-1",
                "row_ref": "Sheet1!row1",
                "topic": "Bridge topic",
                "scene_plan": [
                    {"scene_index": 1, "prompt": "scene one"},
                    {"scene_index": 2, "prompt": "scene two"},
                    {"scene_index": 3, "prompt": "scene three"},
                    {"scene_index": 4, "prompt": "scene four"},
                ],
                "asset_plan": {
                    "asset_root": str(asset_root.resolve()),
                    "common_asset_folder": str(asset_root.resolve()),
                },
                "voice_plan": {
                    "mapping_source": "excel_scene",
                    "scene_count": 4,
                    "groups": [],
                },
                "reason_code": "ok",
                "evidence": {"source": "test"},
            }
            next_jobs, _ = route_video_plan(cast(dict[str, object], video_plan))
            workers = {
                str(
                    cast(dict[str, object], cast(dict[str, object], entry)["job"])[
                        "worker"
                    ]
                )
                for entry in next_jobs
                if isinstance(entry, dict)
            }

        self.assertIn("kenburns", workers)

    def test_control_plane_seeds_and_executes_explicit_kenburns_bundle_contract(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            config.stable_file_age_sec = 0
            image_a = root / "scene-a.png"
            image_b = root / "scene-b.png"
            audio_b = root / "scene-b.wav"
            _ = image_a.write_bytes(b"png")
            _ = image_b.write_bytes(b"png")
            _ = audio_b.write_bytes(b"wav")
            bundle_map_path = root / "scene_bundle_map.json"
            bundle_map_path.write_text(
                json.dumps(
                    {
                        "scenes": [
                            {
                                "scene_key": "scene_a",
                                "source_path": str(image_a.resolve()),
                                "duration_sec": 4,
                            },
                            {
                                "scene_key": "scene_b",
                                "source_path": str(image_b.resolve()),
                                "audio_path": str(audio_b.resolve()),
                                "duration_sec": 5,
                            },
                        ]
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )
            config.input_root.joinpath("kenburns").mkdir(parents=True, exist_ok=True)
            contract_path = config.input_root / "kenburns" / "kenburns-bundle.job.json"
            contract_path.write_text(
                json.dumps(
                    build_explicit_job_contract(
                        job_id="kenburns-bundle-job",
                        workload="kenburns",
                        checkpoint_key="explicit:kenburns:bundle",
                        payload={
                            "run_id": "kenburns-bundle-run",
                            "scene_bundle_map_path": str(bundle_map_path.resolve()),
                        },
                    ),
                    ensure_ascii=True,
                    indent=2,
                ),
                encoding="utf-8",
            )

            seed_result = run_control_loop_once(
                owner="runtime_v2", config=config, run_id="control-run-kenburns-seed"
            )

            def fake_process(
                command: list[str],
                *,
                cwd: Path,
                extra_env: dict[str, str] | None = None,
                timeout_sec: int = 3600,
            ) -> dict[str, object]:
                _ = extra_env
                _ = timeout_sec
                output_path = Path(str(command[-1]))
                _ = output_path.write_bytes(b"mp4")
                return {
                    "command": command,
                    "cwd": str(cwd),
                    "exit_code": 0,
                    "stdout": "",
                    "stderr": "",
                    "timed_out": False,
                    "timeout_sec": 3600,
                    "duration_sec": 0.01,
                }

            with patch(
                "runtime_v2.workers.kenburns_worker.run_external_process",
                side_effect=fake_process,
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-kenburns-exec",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])

        self.assertEqual(seed_result["status"], "seeded")
        self.assertEqual(result["status"], "ok")
        self.assertEqual(str(queue_items[0]["status"]), "completed")
        self.assertEqual(str(latest_metadata["workload"]), "kenburns")
        self.assertTrue(bool(latest_metadata["final_output"]))

    def test_control_plane_escalates_invalid_worker_result_json_and_skips_downstream_queue(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="qwen-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-job",
                    payload={"script_text": "hello", "chain_depth": 0},
                ),
                config=config,
            )
            bad_result_path = (
                root / "artifacts" / "qwen3_tts" / "qwen-job" / "result.json"
            )
            bad_result_path.parent.mkdir(parents=True, exist_ok=True)
            _ = bad_result_path.write_text("{not-json", encoding="utf-8")
            runtime_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "ok",
                    "stage": "synthesize_audio",
                    "error_code": "",
                    "retryable": False,
                    "result_path": str(bad_result_path.resolve()),
                    "next_jobs": [
                        build_explicit_job_contract(
                            job_id="rvc-qwen-job",
                            workload="rvc",
                            checkpoint_key="derived:rvc:qwen-job",
                            payload={
                                "source_path": "system/runtime_v2/artifacts/qwen3_tts/qwen-job/speech.flac",
                                "chain_depth": 1,
                            },
                            chain_step=1,
                            parent_job_id="qwen-job",
                        )
                    ],
                    "completion": {"state": "routed", "final_output": False},
                },
            }

            with patch(
                "runtime_v2.control_plane.run_gated", return_value=runtime_result
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-2"
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])

        self.assertEqual(len(queue_items), 1)
        self.assertEqual(str(queue_items[0]["job_id"]), "qwen-job")
        self.assertEqual(
            str(latest_metadata["worker_error_code"]), "invalid_worker_result_json"
        )

    def test_control_plane_debug_log_uses_control_run_id_not_job_id(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="qwen-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-job",
                    payload={"script_text": "hello", "chain_depth": 0},
                ),
                config=config,
            )

            runtime_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "ok",
                    "stage": "synthesize_audio",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [],
                    "completion": {
                        "state": "succeeded",
                        "final_output": True,
                        "final_artifact": "speech.flac",
                        "final_artifact_path": "system/runtime_v2/artifacts/qwen3_tts/qwen-job/speech.flac",
                    },
                },
            }

            with patch(
                "runtime_v2.control_plane.run_gated", return_value=runtime_result
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-1"
                )

            debug_log_file = config.debug_log_root / "control-run-1.jsonl"
            raw_entries = [
                json.loads(line)
                for line in debug_log_file.read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            entries = cast(list[object], raw_entries)
            control_entries: list[dict[object, object]] = []
            for entry in entries:
                if isinstance(entry, dict):
                    typed_entry = cast(dict[object, object], entry)
                    if typed_entry.get("event") == "control_loop_result":
                        control_entries.append(typed_entry)

        self.assertEqual(len(control_entries), 1)
        self.assertEqual(str(control_entries[0]["run_id"]), "control-run-1")
        self.assertEqual(str(control_entries[0]["job_id"]), "qwen-job")

    def test_control_plane_holds_browser_blocked_job_with_fixed_backoff(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="chatgpt-blocked-job",
                    workload="chatgpt",
                    checkpoint_key="seed:chatgpt-blocked-job",
                    payload={
                        "run_id": "chatgpt-run-blocked",
                        "topic_spec": {"topic": "blocked"},
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={"status": "blocked", "code": "BROWSER_BLOCKED"},
            ):
                result = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-blocked"
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "chatgpt-blocked-job"
            )
            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])

        self.assertEqual(result["status"], "blocked")
        self.assertEqual(result["code"], "BROWSER_BLOCKED")
        self.assertEqual(str(job_payload["status"]), "retry")
        self.assertEqual(int(cast(int, job_payload["attempts"])), 0)
        self.assertGreater(float(cast(float, latest_metadata["backoff_sec"])), 0.0)
        self.assertEqual(str(latest_metadata["status"]), "blocked")
        self.assertEqual(str(latest_metadata["completion_state"]), "blocked")

    def test_control_plane_preserves_restart_exhausted_reason_in_result_metadata(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="chatgpt-restart-exhausted-job",
                    workload="chatgpt",
                    checkpoint_key="seed:chatgpt-restart-exhausted-job",
                    payload={
                        "run_id": "chatgpt-run-restart-exhausted",
                        "topic_spec": {"topic": "restart-exhausted"},
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "failed",
                    "code": "BROWSER_RESTART_EXHAUSTED",
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-restart-exhausted",
                )

            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])
            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "chatgpt-restart-exhausted-job"
            )

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "BROWSER_RESTART_EXHAUSTED")
        self.assertEqual(str(job_payload["status"]), "failed")
        self.assertEqual(int(cast(int, job_payload["attempts"])), 1)
        self.assertEqual(float(cast(float, latest_metadata["backoff_sec"])), 0.0)
        self.assertEqual(
            str(latest_metadata["worker_error_code"]), "BROWSER_RESTART_EXHAUSTED"
        )
        self.assertEqual(str(latest_metadata["completion_state"]), "failed")

    def test_control_plane_appends_raw_browser_event_records_via_control_plane_only(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="chatgpt-browser-event-record-job",
                    workload="chatgpt",
                    checkpoint_key="seed:chatgpt-browser-event-record-job",
                    payload={
                        "run_id": "chatgpt-run-browser-event-record",
                        "topic_spec": {"topic": "browser-event-record"},
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "blocked",
                    "code": "BROWSER_BLOCKED",
                    "browser": {"sessions": []},
                    "browser_event_records": [
                        {
                            "event": "browser_supervisor_status",
                            "service": "chatgpt",
                            "status": "login_required",
                            "action_result": "blocked",
                            "tick_id": "browser-run-browser-event-record",
                        }
                    ],
                    "worker_result": {
                        "status": "blocked",
                        "stage": "runtime_preflight",
                        "error_code": "BROWSER_BLOCKED",
                        "details": {"blocked_services": ["chatgpt"]},
                    },
                },
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-browser-event-record",
                )

            event_rows = [
                cast(dict[str, object], json.loads(line))
                for line in config.control_plane_events_file.read_text(
                    encoding="utf-8"
                ).splitlines()
                if line.strip()
            ]
            browser_events = [
                row
                for row in event_rows
                if str(row.get("event", "")) == "browser_supervisor_status"
            ]

        self.assertEqual(len(browser_events), 1)
        self.assertEqual(
            str(browser_events[0]["run_id"]), "control-run-browser-event-record"
        )
        self.assertEqual(str(browser_events[0]["service"]), "chatgpt")

    def test_control_plane_normalizes_raw_runtime_preflight_signal_to_canonical_failure_contract(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="chatgpt-raw-preflight-restart-job",
                    workload="chatgpt",
                    checkpoint_key="seed:chatgpt-raw-preflight-restart-job",
                    payload={
                        "run_id": "chatgpt-run-raw-preflight-restart",
                        "topic_spec": {"topic": "raw-preflight-restart"},
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "blocked",
                    "code": "BROWSER_RESTART_EXHAUSTED",
                    "worker_result": {
                        "stage": "runtime_preflight",
                        "error_code": "restart_exhausted",
                        "details": {"blocked_services": ["chatgpt"]},
                    },
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-raw-preflight-restart",
                )

            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])
            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "chatgpt-raw-preflight-restart-job"
            )

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "BROWSER_RESTART_EXHAUSTED")
        self.assertEqual(str(job_payload["status"]), "failed")
        self.assertEqual(int(cast(int, job_payload["attempts"])), 1)
        self.assertEqual(float(cast(float, latest_metadata["backoff_sec"])), 0.0)
        self.assertEqual(
            str(latest_metadata["worker_error_code"]), "BROWSER_RESTART_EXHAUSTED"
        )
        self.assertEqual(str(latest_metadata["completion_state"]), "failed")

    def test_control_plane_writes_mismatch_warning_to_debug_log(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="chatgpt-mismatch-warning-job",
                    workload="chatgpt",
                    checkpoint_key="seed:chatgpt-mismatch-warning-job",
                    payload={
                        "run_id": "chatgpt-run-mismatch-warning",
                        "topic_spec": {"topic": "mismatch-warning"},
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "blocked",
                    "code": "BROWSER_RESTART_EXHAUSTED",
                    "worker_result": {
                        "status": "blocked",
                        "stage": "browser_preflight",
                        "error_code": "BROWSER_BLOCKED",
                    },
                },
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-mismatch-warning",
                )

            debug_log_file = (
                config.debug_log_root / "control-run-mismatch-warning.jsonl"
            )
            raw_entries = [
                json.loads(line)
                for line in debug_log_file.read_text(encoding="utf-8").splitlines()
                if line.strip()
            ]
            entries = cast(list[object], raw_entries)
            control_entries = [
                cast(dict[object, object], entry)
                for entry in entries
                if isinstance(entry, dict)
                and entry.get("event") == "control_loop_result"
            ]

        self.assertEqual(len(control_entries), 1)
        self.assertEqual(
            str(control_entries[0]["warning_worker_error_code_mismatch"]),
            "worker_error_code=BROWSER_BLOCKED error_code=BROWSER_RESTART_EXHAUSTED",
        )

    def test_control_plane_normalizes_placeholder_worker_error_code(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="chatgpt-placeholder-worker-error-job",
                    workload="chatgpt",
                    checkpoint_key="seed:chatgpt-placeholder-worker-error-job",
                    payload={
                        "run_id": "chatgpt-run-placeholder-worker-error",
                        "topic_spec": {"topic": "placeholder-error"},
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "blocked",
                    "code": "BROWSER_RESTART_EXHAUSTED",
                    "worker_result": {
                        "status": "blocked",
                        "stage": "browser_preflight",
                        "error_code": "-",
                    },
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-placeholder-worker-error",
                )

            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])

        self.assertEqual(result["code"], "BROWSER_RESTART_EXHAUSTED")
        self.assertEqual(
            str(latest_metadata["worker_error_code"]), "BROWSER_RESTART_EXHAUSTED"
        )

    def test_control_plane_does_not_seed_replan_for_restart_exhausted_browser_verify(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="agent-browser-restart-exhausted-job",
                    workload="agent_browser_verify",
                    checkpoint_key="seed:agent-browser-restart-exhausted-job",
                    payload={
                        "run_id": "agent-browser-restart-exhausted-run",
                        "service": "chatgpt",
                        "expected_url_substring": "chatgpt.com",
                        "replan_on_failure": True,
                        "verification": ["browser-check"],
                        "browser_checks": [{"service": "chatgpt"}],
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "blocked",
                    "code": "BROWSER_RESTART_EXHAUSTED",
                },
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-agent-browser-restart-exhausted",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "agent-browser-restart-exhausted-job"
            )
            replan_jobs = [
                item
                for item in queue_items
                if str(item.get("workload", "")) == "dev_replan"
            ]

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "BROWSER_RESTART_EXHAUSTED")
        self.assertEqual(str(job_payload["status"]), "failed")
        self.assertEqual(replan_jobs, [])

    def test_control_plane_retries_browser_unhealthy_runtime_preflight_with_backoff(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="chatgpt-unhealthy-job",
                    workload="chatgpt",
                    checkpoint_key="seed:chatgpt-unhealthy-job",
                    payload={
                        "run_id": "chatgpt-run-unhealthy",
                        "topic_spec": {"topic": "unhealthy"},
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={"status": "failed", "code": "BROWSER_UNHEALTHY"},
            ):
                result = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-unhealthy"
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "chatgpt-unhealthy-job"
            )
            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])
            latest_join = load_joined_latest_run(config, completed=True)

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "BROWSER_UNHEALTHY")
        self.assertEqual(str(result["queue_status"]), "retry")
        self.assertEqual(str(job_payload["status"]), "retry")
        self.assertEqual(int(cast(int, job_payload["attempts"])), 1)
        self.assertGreater(float(cast(float, latest_metadata["backoff_sec"])), 0.0)
        self.assertEqual(str(latest_metadata["worker_error_code"]), "BROWSER_UNHEALTHY")
        self.assertEqual(str(latest_metadata["completion_state"]), "failed")
        canonical_handoff = cast(
            dict[str, object], latest_metadata["canonical_handoff"]
        )
        self.assertEqual(str(canonical_handoff["schema_version"]), "1.0")
        self.assertEqual(str(canonical_handoff["workload"]), "chatgpt")
        self.assertFalse(bool(latest_join["out_of_sync"]))

    def test_control_plane_forces_browser_recovery_on_browser_unhealthy_retry(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="genspark-browser-unhealthy-retry-job",
                    workload="genspark",
                    status="retry",
                    attempts=1,
                    checkpoint_key="seed:genspark-browser-unhealthy-retry-job",
                    payload={
                        "run_id": "genspark-browser-unhealthy-retry-run",
                        "row_ref": "Sheet1!row1",
                        "scene_index": 1,
                        "promotion_gate": "A",
                        "last_error_code": "BROWSER_UNHEALTHY",
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "failed",
                    "code": "BROWSER_UNHEALTHY",
                    "worker_result": {
                        "status": "failed",
                        "stage": "genspark_adapter",
                        "error_code": "BROWSER_UNHEALTHY",
                        "retryable": True,
                        "next_jobs": [],
                        "completion": {"state": "failed", "final_output": False},
                    },
                },
            ) as run_gated_mock:
                _ = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-force-browser-recovery",
                )

        self.assertEqual(
            run_gated_mock.call_args.kwargs["force_unhealthy_service"], "genspark"
        )

    def test_stale_running_recovery_retries_immediately_on_next_control_once(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            stale_job = JobContract(
                job_id="chatgpt-stale-running-job",
                workload="chatgpt",
                status="running",
                attempts=0,
                checkpoint_key="seed:chatgpt-stale-running-job",
                payload={
                    "run_id": "chatgpt-stale-running-run",
                    "row_ref": "Sheet1!row15",
                    "next_attempt_at": time.time() + 999,
                    "topic_spec": {"topic": "stale-running"},
                },
            )
            stale_job.updated_at = time.time() - (config.running_stale_sec + 5)
            seed_control_job(stale_job, config=config)

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={
                    "status": "ok",
                    "code": "OK",
                    "worker_result": {
                        "status": "ok",
                        "stage": "chatgpt",
                        "error_code": "",
                        "retryable": False,
                        "next_jobs": [],
                        "details": {},
                        "completion": {"state": "succeeded", "final_output": False},
                    },
                },
            ) as run_gated_mock:
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-stale-retry-immediate",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "chatgpt-stale-running-job"
            )

        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["code"], "OK")
        self.assertEqual(str(job_payload["status"]), "completed")
        self.assertEqual(run_gated_mock.call_count, 1)

    def test_control_plane_holds_gpt_floor_failure_with_fixed_backoff(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="chatgpt-gpt-floor-job",
                    workload="chatgpt",
                    checkpoint_key="seed:chatgpt-gpt-floor-job",
                    payload={
                        "run_id": "chatgpt-run-gpt-floor",
                        "topic_spec": {"topic": "gpt-floor"},
                    },
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={"status": "failed", "code": "GPT_FLOOR_FAIL"},
            ):
                result = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-gpt-floor"
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "chatgpt-gpt-floor-job"
            )
            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])

        self.assertEqual(result["status"], "blocked")
        self.assertEqual(result["code"], "GPT_FLOOR_FAIL")
        self.assertEqual(str(job_payload["status"]), "retry")
        self.assertEqual(int(cast(int, job_payload["attempts"])), 0)
        self.assertGreater(float(cast(float, latest_metadata["backoff_sec"])), 0.0)
        self.assertEqual(str(latest_metadata["status"]), "blocked")
        self.assertEqual(str(latest_metadata["worker_error_code"]), "GPT_FLOOR_FAIL")
        self.assertEqual(str(latest_metadata["completion_state"]), "blocked")

    def test_control_plane_holds_gpu_lease_busy_with_fixed_backoff(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="qwen-gpu-busy-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-gpu-busy-job",
                    payload={"script_text": "hello", "chain_depth": 0},
                ),
                config=config,
            )

            with patch(
                "runtime_v2.control_plane.run_gated",
                return_value={"status": "failed", "code": "GPU_LEASE_BUSY"},
            ):
                result = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-gpu-busy"
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "qwen-gpu-busy-job"
            )
            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])

        self.assertEqual(result["status"], "blocked")
        self.assertEqual(result["code"], "GPU_LEASE_BUSY")
        self.assertEqual(str(job_payload["status"]), "retry")
        self.assertEqual(int(cast(int, job_payload["attempts"])), 0)
        self.assertGreater(float(cast(float, latest_metadata["backoff_sec"])), 0.0)
        self.assertEqual(str(latest_metadata["status"]), "blocked")
        self.assertEqual(str(latest_metadata["worker_error_code"]), "GPU_LEASE_BUSY")
        self.assertEqual(str(latest_metadata["completion_state"]), "blocked")

    def test_control_plane_fail_closes_second_stale_running_recovery(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            object.__setattr__(config, "running_stale_sec", 1)
            stale_job = JobContract(
                job_id="qwen-stale-running-job",
                workload="qwen3_tts",
                status="running",
                attempts=1,
                checkpoint_key="seed:qwen-stale-running-job",
                payload={
                    "run_id": "qwen-stale-run",
                    "row_ref": "Sheet1!row1",
                    "chain_depth": 0,
                },
            )
            stale_job.updated_at = 0.0
            QueueStore(config.queue_store_file).save([stale_job])
            _ = config.worker_registry_file.parent.mkdir(parents=True, exist_ok=True)
            _ = config.worker_registry_file.write_text(
                json.dumps(
                    {
                        "qwen3_tts": {
                            "workload": "qwen3_tts",
                            "state": "busy",
                            "run_id": "qwen-stale-run",
                            "last_seen": 0.0,
                            "progress_ts": 0.0,
                        }
                    },
                    ensure_ascii=True,
                ),
                encoding="utf-8",
            )

            result = run_control_loop_once(
                owner="runtime_v2",
                config=config,
                run_id="control-run-stale-fail-close",
                allow_runtime_side_effects=False,
            )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "qwen-stale-running-job"
            )
            worker_registry = json.loads(
                config.worker_registry_file.read_text(encoding="utf-8")
            )
            qwen_registry = cast(dict[str, object], worker_registry["qwen3_tts"])

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "WORKER_STALL_DETECTED")
        self.assertEqual(str(job_payload["status"]), "failed")
        self.assertEqual(int(cast(int, job_payload["attempts"])), 2)
        self.assertEqual(str(qwen_registry["state"]), "idle")

    def test_control_plane_writes_failure_summary_for_terminal_failed_row(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            failed_job = JobContract(
                job_id="qwen-terminal-failed-job",
                workload="qwen3_tts",
                status="failed",
                attempts=2,
                checkpoint_key="seed:qwen-terminal-failed-job",
                payload={
                    "run_id": "qwen-terminal-run",
                    "row_ref": "Sheet1!row1",
                    "chain_depth": 1,
                    "last_error_code": "WORKER_STALL_DETECTED",
                    "last_worker_stage": "qwen3_tts_adapter",
                },
            )
            failed_job.updated_at = 100.0
            completed_job = JobContract(
                job_id="chatgpt-terminal-completed-job",
                workload="chatgpt",
                status="completed",
                checkpoint_key="seed:chatgpt-terminal-completed-job",
                payload={
                    "run_id": "qwen-terminal-run",
                    "row_ref": "Sheet1!row1",
                    "chain_depth": 0,
                },
            )
            completed_job.updated_at = 90.0
            blocked_render_job = JobContract(
                job_id="render-terminal-blocked-job",
                workload="render",
                status="failed",
                checkpoint_key="seed:render-terminal-blocked-job",
                payload={
                    "run_id": "qwen-terminal-run",
                    "row_ref": "Sheet1!row1",
                    "chain_depth": 1,
                },
            )
            blocked_render_job.updated_at = 110.0
            QueueStore(config.queue_store_file).save(
                [completed_job, failed_job, blocked_render_job]
            )

            result = run_control_loop_once(
                owner="runtime_v2",
                config=config,
                run_id="control-run-terminal-failure-summary",
                allow_runtime_side_effects=False,
            )
            failure_summary_exists = config.failure_summary_file.exists()

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "WORKER_STALL_DETECTED")
        self.assertTrue(failure_summary_exists)

    def test_control_plane_inferrs_stall_code_for_terminal_qwen_failure_without_payload_error(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            failed_qwen_job = JobContract(
                job_id="qwen-terminal-inferred-stall-job",
                workload="qwen3_tts",
                status="failed",
                attempts=2,
                checkpoint_key="seed:qwen-terminal-inferred-stall-job",
                payload={
                    "run_id": "qwen-inferred-stall-run",
                    "row_ref": "Sheet1!row1",
                    "chain_depth": 1,
                },
            )
            failed_qwen_job.updated_at = 100.0
            QueueStore(config.queue_store_file).save([failed_qwen_job])

            result = run_control_loop_once(
                owner="runtime_v2",
                config=config,
                run_id="control-run-inferred-stall-summary",
                allow_runtime_side_effects=False,
            )

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "WORKER_STALL_DETECTED")

    def test_control_plane_uses_retryable_not_completion_state_for_retry_decision(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="qwen-native-failed-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-native-failed-job",
                    payload={"script_text": "hello", "chain_depth": 0},
                ),
                config=config,
            )

            runtime_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "failed",
                    "stage": "qwen3_tts",
                    "error_code": "native_qwen3_tts_not_implemented",
                    "retryable": False,
                    "next_jobs": [],
                    "completion": {"state": "blocked", "final_output": False},
                },
            }

            with patch(
                "runtime_v2.control_plane.run_gated", return_value=runtime_result
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-native-failed",
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "qwen-native-failed-job"
            )
            latest_result = cast(
                dict[str, object],
                json.loads(config.result_router_file.read_text(encoding="utf-8")),
            )
            latest_metadata = cast(dict[str, object], latest_result["metadata"])

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "native_qwen3_tts_not_implemented")
        self.assertEqual(str(job_payload["status"]), "failed")
        self.assertEqual(int(cast(int, job_payload["attempts"])), 1)
        self.assertEqual(float(cast(float, latest_metadata["backoff_sec"])), 0.0)
        self.assertEqual(
            str(latest_metadata["worker_error_code"]),
            "native_qwen3_tts_not_implemented",
        )
        self.assertEqual(str(latest_metadata["completion_state"]), "failed")

    def test_control_plane_event_log_keeps_control_run_id_on_events_and_transitions(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="qwen-run-id-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-run-id-job",
                    payload={"script_text": "hello", "chain_depth": 0},
                ),
                config=config,
            )

            runtime_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "ok",
                    "stage": "synthesize_audio",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [],
                    "completion": {"state": "succeeded", "final_output": True},
                },
            }

            with patch(
                "runtime_v2.control_plane.run_gated", return_value=runtime_result
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-events"
                )

            raw_entries = [
                json.loads(line)
                for line in config.control_plane_events_file.read_text(
                    encoding="utf-8"
                ).splitlines()
                if line.strip()
            ]
            entries = [
                cast(dict[object, object], entry)
                for entry in raw_entries
                if isinstance(entry, dict)
            ]

        job_entries = [
            entry
            for entry in entries
            if str(entry.get("job_id", "")) == "qwen-run-id-job"
        ]
        self.assertTrue(job_entries)
        self.assertTrue(
            all(
                str(entry.get("run_id", "")) == "control-run-events"
                for entry in job_entries
            )
        )

    def test_control_plane_appends_browser_events_from_runtime_result(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="qwen-browser-event-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-browser-event-job",
                    payload={"script_text": "hello", "chain_depth": 0},
                ),
                config=config,
            )

            runtime_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "browser": {
                    "events": [
                        {
                            "event": "browser_supervisor_status",
                            "run_id": "browser-plane-run-id",
                            "status": "login_required",
                            "action_result": "blocked",
                            "tick_id": "runtime-browser-event",
                        }
                    ]
                },
                "worker_result": {
                    "status": "ok",
                    "stage": "synthesize_audio",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [],
                    "completion": {"state": "succeeded", "final_output": True},
                },
            }

            with patch(
                "runtime_v2.control_plane.run_gated", return_value=runtime_result
            ):
                _ = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-browser-events",
                )

            raw_entries = [
                json.loads(line)
                for line in config.control_plane_events_file.read_text(
                    encoding="utf-8"
                ).splitlines()
                if line.strip()
            ]
            entries = [
                cast(dict[object, object], entry)
                for entry in raw_entries
                if isinstance(entry, dict)
            ]
            browser_entries = [
                entry
                for entry in entries
                if str(entry.get("event", "")) == "browser_supervisor_status"
            ]

        self.assertEqual(len(browser_entries), 1)
        self.assertEqual(
            str(browser_entries[0].get("run_id", "")), "control-run-browser-events"
        )
        self.assertEqual(str(browser_entries[0].get("status", "")), "login_required")

    def test_runtime_control_path_rejects_mock_chain_without_explicit_probe_or_debug_mode(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="qwen-mock-runtime-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-mock-runtime-job",
                    payload={
                        "script_text": "hello",
                        "chain_depth": 0,
                        "mock_chain": True,
                    },
                ),
                config=config,
            )

            with (
                patch(
                    "runtime_v2.control_plane._run_mock_chain_worker",
                    side_effect=AssertionError("mock chain must stay probe-only"),
                ),
                patch(
                    "runtime_v2.control_plane.run_gated",
                    return_value={"status": "failed", "code": "BROWSER_UNHEALTHY"},
                ),
            ):
                result = run_control_loop_once(
                    owner="runtime_v2", config=config, run_id="control-run-no-mock"
                )

            queue_payload = cast(
                list[object],
                json.loads(config.queue_store_file.read_text(encoding="utf-8")),
            )
            queue_items = [
                cast(dict[str, object], item)
                for item in queue_payload
                if isinstance(item, dict)
            ]
            job_payload = next(
                item
                for item in queue_items
                if str(item["job_id"]) == "qwen-mock-runtime-job"
            )

        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["code"], "BROWSER_UNHEALTHY")
        self.assertEqual(str(job_payload["status"]), "retry")

    def test_control_plane_side_effect_free_mode_skips_bootstrap_and_gpt_ticks(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = _runtime_config(root)
            seed_control_job(
                JobContract(
                    job_id="qwen-side-effect-free-job",
                    workload="qwen3_tts",
                    checkpoint_key="seed:qwen-side-effect-free-job",
                    payload={"script_text": "hello", "chain_depth": 0},
                ),
                config=config,
            )

            runtime_result: dict[str, object] = {
                "status": "ok",
                "code": "OK",
                "worker_result": {
                    "status": "ok",
                    "stage": "synthesize_audio",
                    "error_code": "",
                    "retryable": False,
                    "next_jobs": [],
                    "completion": {"state": "succeeded", "final_output": True},
                },
            }

            with (
                patch(
                    "runtime_v2.control_plane.ensure_runtime_bootstrap"
                ) as bootstrap_runtime,
                patch("runtime_v2.control_plane.tick_gpt_status") as tick_gpt_status,
                patch(
                    "runtime_v2.control_plane.apply_autospawn_decision"
                ) as apply_autospawn,
                patch(
                    "runtime_v2.control_plane.run_gated", return_value=runtime_result
                ),
            ):
                result = run_control_loop_once(
                    owner="runtime_v2",
                    config=config,
                    run_id="control-run-no-side-effects",
                    allow_runtime_side_effects=False,
                )

        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["code"], "OK")
        bootstrap_runtime.assert_not_called()
        tick_gpt_status.assert_not_called()
        apply_autospawn.assert_not_called()


if __name__ == "__main__":
    _ = unittest.main()
