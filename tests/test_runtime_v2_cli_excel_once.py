from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from typing import cast
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from runtime_v2 import exit_codes
from runtime_v2.cli import main


def _write_excel_fixture(path: Path) -> Path:
    workbook = Workbook()
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "Sheet1"
    sheet.append(["Topic", "Status", "Video Plan", "Reason Code"])
    sheet.append(["Bridge topic", "", "", ""])
    workbook.save(path)
    workbook.close()
    return path


class RuntimeV2CliExcelOnceTests(unittest.TestCase):
    def test_excel_once_keeps_ok_status_blocked_by_default(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx")
            workbook = load_workbook(excel_path)
            try:
                sheet = cast(Worksheet, workbook["Sheet1"])
                status_cell = cast(Cell, sheet["B2"])
                status_cell.value = "OK"
                workbook.save(excel_path)
            finally:
                workbook.close()

            with patch(
                "sys.argv",
                [
                    "runtime_v2.cli",
                    "--excel-once",
                    "--excel-path",
                    str(excel_path),
                    "--sheet-name",
                    "Sheet1",
                    "--row-index",
                    "0",
                    "--runtime-root",
                    str(root / "runtime"),
                ],
            ):
                exit_code = main()

            job_file = (
                root / "runtime" / "inbox" / "chatgpt" / "chatgpt-sheet1-1.job.json"
            )

            self.assertEqual(exit_code, exit_codes.SUCCESS)
            self.assertFalse(job_file.exists())

    def test_excel_once_can_opt_in_ok_status_for_regular_rerun(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx")
            workbook = load_workbook(excel_path)
            try:
                sheet = cast(Worksheet, workbook["Sheet1"])
                status_cell = cast(Cell, sheet["B2"])
                status_cell.value = "OK"
                workbook.save(excel_path)
            finally:
                workbook.close()

            with patch(
                "sys.argv",
                [
                    "runtime_v2.cli",
                    "--excel-once",
                    "--excel-path",
                    str(excel_path),
                    "--sheet-name",
                    "Sheet1",
                    "--row-index",
                    "0",
                    "--runtime-root",
                    str(root / "runtime"),
                    "--accepted-statuses",
                    "ok,seeded",
                ],
            ):
                exit_code = main()

            job_file = (
                root / "runtime" / "inbox" / "chatgpt" / "chatgpt-sheet1-1.job.json"
            )

            self.assertEqual(exit_code, exit_codes.SUCCESS)
            self.assertTrue(job_file.exists())


if __name__ == "__main__":
    _ = unittest.main()
