from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from typing import cast

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from runtime_v2.manager import merge_stage1_result


def _write_merge_fixture(path: Path, *, headers: list[str], row: list[object]) -> Path:
    workbook = Workbook()
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "Sheet1"
    sheet.append(headers)
    sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def _row_values(path: Path) -> dict[str, object]:
    workbook = load_workbook(path)
    try:
        sheet = workbook["Sheet1"]
        headers = [str(cell.value) for cell in sheet[1]]
        values = [
            sheet.cell(row=2, column=index + 1).value for index in range(len(headers))
        ]
        return dict(zip(headers, values, strict=False))
    finally:
        workbook.close()


class RuntimeV2Stage1ExcelMergeTests(unittest.TestCase):
    def test_video_plan_merge_updates_only_allowed_columns_and_status(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            excel_path = _write_merge_fixture(
                Path(tmp_dir) / "topic.xlsx",
                headers=["Topic", "status", "video plan", "REASON CODE", "Owner"],
                row=["Bridge topic", "", "", "", "keep-me"],
            )

            merged = merge_stage1_result(
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                video_plan={
                    "topic": "Bridge topic",
                    "story_outline": ["a", "b"],
                    "reason_code": "ok",
                },
            )

            values = _row_values(excel_path)

        self.assertTrue(merged)
        self.assertEqual(values["status"], "OK")
        self.assertEqual(values["REASON CODE"], "ok")
        self.assertIn("Bridge topic", str(values["video plan"]))
        self.assertEqual(values["Owner"], "keep-me")

    def test_terminal_rows_are_not_overwritten_by_stage1_merge(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            excel_path = _write_merge_fixture(
                Path(tmp_dir) / "topic.xlsx",
                headers=["Topic", "Status", "Video Plan", "Reason Code"],
                row=["Bridge topic", "Done", "locked", "old"],
            )

            merged = merge_stage1_result(
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                video_plan={
                    "topic": "Bridge topic",
                    "story_outline": ["a", "b"],
                    "reason_code": "ok",
                },
            )

            values = _row_values(excel_path)

        self.assertFalse(merged)
        self.assertEqual(values["Status"], "Done")
        self.assertEqual(values["Video Plan"], "locked")
        self.assertEqual(values["Reason Code"], "old")

    def test_stage1_merge_rejects_stale_excel_snapshot(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            excel_path = _write_merge_fixture(
                Path(tmp_dir) / "topic.xlsx",
                headers=["Topic", "Status", "Video Plan", "Reason Code"],
                row=["Bridge topic", "partial", "", ""],
            )

            merged = merge_stage1_result(
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                video_plan={
                    "topic": "Bridge topic",
                    "story_outline": ["a", "b"],
                    "reason_code": "ok",
                    "evidence": {"excel_snapshot_hash": "stale-hash"},
                },
            )

            values = _row_values(excel_path)

        self.assertFalse(merged)
        self.assertEqual(values["Status"], "partial")
        self.assertIn(values["Video Plan"], {"", None})

    def test_stage1_merge_returns_false_when_sheet_is_missing(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            excel_path = _write_merge_fixture(
                Path(tmp_dir) / "topic.xlsx",
                headers=["Topic", "Status", "Video Plan", "Reason Code"],
                row=["Bridge topic", "", "", ""],
            )

            merged = merge_stage1_result(
                excel_path=excel_path,
                sheet_name="MissingSheet",
                row_index=0,
                video_plan={
                    "topic": "Bridge topic",
                    "story_outline": ["a", "b"],
                    "reason_code": "ok",
                },
            )

        self.assertFalse(merged)

    def test_stage1_merge_writes_parsed_fields_when_handoff_exists(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            excel_path = _write_merge_fixture(
                Path(tmp_dir) / "topic.xlsx",
                headers=[
                    "Topic",
                    "Status",
                    "Video Plan",
                    "Reason Code",
                    "Title",
                    "Title for Thumb",
                    "Description",
                    "Keywords",
                    "Voice",
                ],
                row=["Bridge topic", "", "", "", "", "", "", "", ""],
            )

            merged = merge_stage1_result(
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                video_plan={
                    "topic": "Bridge topic",
                    "story_outline": ["a", "b"],
                    "reason_code": "ok",
                    "stage1_handoff": {
                        "contract": {
                            "title": "Bridge title",
                            "title_for_thumb": "Bridge thumb",
                            "description": "Bridge description",
                            "keywords": ["bridge", "topic"],
                            "bgm": "calm piano",
                            "scene_prompts": ["scene one", "scene two"],
                            "voice_groups": [{"scene_index": 1, "voice": "narration"}],
                        }
                    },
                },
            )

            values = _row_values(excel_path)

        self.assertTrue(merged)
        self.assertEqual(values["Title"], "Bridge title")
        self.assertEqual(values["Title for Thumb"], "Bridge thumb")
        self.assertEqual(values["Description"], "Bridge description")
        self.assertEqual(values["Keywords"], "bridge, topic")

    def test_stage1_merge_writes_bgm_and_scene_columns_when_present(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            excel_path = _write_merge_fixture(
                Path(tmp_dir) / "topic.xlsx",
                headers=[
                    "Topic",
                    "Status",
                    "Video Plan",
                    "Reason Code",
                    "BGM",
                    "#01",
                    "#02",
                ],
                row=["Bridge topic", "", "", "", "", "", ""],
            )

            merged = merge_stage1_result(
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                video_plan={
                    "topic": "Bridge topic",
                    "story_outline": ["a", "b"],
                    "reason_code": "ok",
                    "stage1_handoff": {
                        "contract": {
                            "title": "Bridge title",
                            "title_for_thumb": "Bridge thumb",
                            "description": "Bridge description",
                            "keywords": ["bridge", "topic"],
                            "bgm": "calm piano",
                            "scene_prompts": ["scene one", "scene two"],
                            "voice_groups": [
                                {"scene_index": 1, "voice": "narration"},
                                {"scene_index": 2, "voice": "narration"},
                            ],
                        }
                    },
                },
            )

            values = _row_values(excel_path)

        self.assertTrue(merged)
        self.assertEqual(values["BGM"], "calm piano")
        self.assertEqual(values["#01"], "scene one")
        self.assertEqual(values["#02"], "scene two")


if __name__ == "__main__":
    _ = unittest.main()
