from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from typing import cast

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from runtime_v2.cli import _write_probe_result
from runtime_v2.config import RuntimeConfig
from runtime_v2.manager import (
    seed_excel_row,
    sync_failure_result,
    sync_final_video_result,
)
from runtime_v2.stage1.chatgpt_runner import run_stage1_chatgpt_job
from runtime_v2.stage2.json_builders import build_stage2_jobs


def _write_excel_fixture(path: Path) -> Path:
    workbook = Workbook()
    sheet = cast(Worksheet, workbook.active)
    sheet.title = "Sheet1"
    sheet.append(["Topic", "Status", "Video Plan", "Reason Code"])
    sheet.append(["Bridge topic", "", "", ""])
    workbook.save(path)
    workbook.close()
    return path


class RuntimeV2ExcelTopicEndToEndTests(unittest.TestCase):
    def test_seed_excel_row_carries_optional_legacy_fields_into_topic_spec(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = root / "topic.xlsx"
            workbook = Workbook()
            sheet = cast(Worksheet, workbook.active)
            sheet.title = "Sheet1"
            sheet.append(
                [
                    "Topic",
                    "Status",
                    "Ref Img 1",
                    "Ref Img 2",
                    "BGM",
                    "URL",
                    "Video1",
                    "Video2",
                ]
            )
            sheet.append(
                [
                    "Bridge topic",
                    "",
                    "ref prompt one",
                    "ref prompt two",
                    "serious piano",
                    "https://example.com/bridge",
                    "video clip one",
                    "video clip two",
                ]
            )
            workbook.save(excel_path)
            workbook.close()
            config = RuntimeConfig.from_root(root)

            seeded = seed_excel_row(
                config=config,
                run_id="carryover-run-1",
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
            )

        topic_spec = cast(dict[str, object], seeded["topic_spec"])
        self.assertEqual(topic_spec["ref_img_1"], "ref prompt one")
        self.assertEqual(topic_spec["ref_img_2"], "ref prompt two")
        self.assertEqual(topic_spec["bgm"], "serious piano")
        self.assertEqual(topic_spec["url"], "https://example.com/bridge")
        self.assertEqual(topic_spec["videos"], ["video clip one", "video clip two"])

    def test_excel_row1_topic_can_seed_stage1_and_finish_final_video_contracts(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx")
            config = RuntimeConfig.from_root(root)

            seeded = seed_excel_row(
                config=config,
                run_id="e2e-run-1",
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
            )
            topic_spec = cast(dict[str, object], seeded["topic_spec"])
            workspace = root / "artifacts" / "chatgpt"
            workspace.mkdir(parents=True, exist_ok=True)
            stage1_result = run_stage1_chatgpt_job(
                topic_spec,
                workspace,
                debug_log=str((root / "logs" / "e2e-run-1.jsonl").resolve()),
            )
            stage1_details = cast(dict[str, object], stage1_result["details"])
            video_plan = cast(dict[str, object], stage1_details["video_plan"])
            common_assets = root / "common_assets"
            common_assets.mkdir(parents=True, exist_ok=True)
            video_plan["asset_plan"] = {
                "asset_root": str(common_assets.resolve()),
                "common_asset_folder": str(common_assets.resolve()),
            }
            stage2_jobs, render_spec = build_stage2_jobs(video_plan)
            workbook = load_workbook(excel_path)
            try:
                status_cell = cast(Cell, workbook["Sheet1"].cell(row=2, column=2))
                status_cell.value = "Voice OK"
                workbook.save(excel_path)
            finally:
                workbook.close()
            final_video = root / "final_video.mp4"
            final_video.write_bytes(b"mp4")
            synced = sync_final_video_result(
                config=config,
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                worker_result={
                    "completion": {
                        "state": "completed",
                        "final_output": True,
                        "final_artifact": "final_video.mp4",
                        "final_artifact_path": str(final_video.resolve()),
                    },
                    "details": {"reason_code": "ok"},
                },
                run_id="e2e-run-1",
                artifact_root=root,
                debug_log=str((root / "logs" / "e2e-run-1.jsonl").resolve()),
            )

            workbook = load_workbook(excel_path)
            try:
                status_value = workbook["Sheet1"].cell(row=2, column=2).value
            finally:
                workbook.close()

        self.assertEqual(topic_spec["contract"], "topic_spec")
        self.assertEqual(video_plan["contract"], "video_plan")
        self.assertEqual(render_spec["contract"], "render_spec")
        self.assertTrue(bool(stage2_jobs))
        self.assertTrue(synced)
        self.assertEqual(status_value, "Done")

    def test_seed_excel_row_keeps_ok_status_blocked_by_default(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx")
            workbook = load_workbook(excel_path)
            try:
                sheet = cast(Worksheet, workbook["Sheet1"])
                status_cell = cast(Cell, sheet["B2"])
                status_cell.value = "OK"
                workbook.save(excel_path)
            finally:
                workbook.close()
            config = RuntimeConfig.from_root(root)

            seeded = seed_excel_row(
                config=config,
                run_id="ok-status-default-blocked",
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
            )

        self.assertEqual(seeded["status"], "no_work")
        self.assertEqual(seeded["code"], "NO_WORK")

    def test_seed_excel_row_can_opt_in_ok_status_for_closeout(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx")
            workbook = load_workbook(excel_path)
            try:
                sheet = cast(Worksheet, workbook["Sheet1"])
                status_cell = cast(Cell, sheet["B2"])
                status_cell.value = "OK"
                workbook.save(excel_path)
            finally:
                workbook.close()
            config = RuntimeConfig.from_root(root)

            seeded = seed_excel_row(
                config=config,
                run_id="ok-status-closeout",
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                accepted_statuses={"", "partial", "failed", "nan", "ok"},
            )

        self.assertEqual(seeded["status"], "seeded")
        self.assertEqual(seeded["code"], "SEEDED_JOB")

    def test_seed_excel_row_can_resume_seeded_status_idempotently(self) -> None:
        with tempfile.TemporaryDirectory(dir=r"D:\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx")
            workbook = load_workbook(excel_path)
            try:
                sheet = cast(Worksheet, workbook["Sheet1"])
                status_cell = cast(Cell, sheet["B2"])
                status_cell.value = "Seeded"
                workbook.save(excel_path)
            finally:
                workbook.close()
            config = RuntimeConfig.from_root(root)

            seeded = seed_excel_row(
                config=config,
                run_id="seeded-status-closeout",
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                accepted_statuses={"", "partial", "failed", "nan", "ok", "seeded"},
            )

        self.assertEqual(seeded["status"], "seeded")
        self.assertEqual(seeded["code"], "SEEDED_JOB")

    def test_control_plane_keeps_same_run_id_across_excel_stage1_stage2_final(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            excel_path = _write_excel_fixture(root / "topic.xlsx")
            config = RuntimeConfig.from_root(root)
            run_id = "e2e-run-2"
            seeded = seed_excel_row(
                config=config,
                run_id=run_id,
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
            )
            topic_spec = cast(dict[str, object], seeded["topic_spec"])
            workspace = root / "artifacts" / "chatgpt"
            workspace.mkdir(parents=True, exist_ok=True)
            stage1_result = run_stage1_chatgpt_job(
                topic_spec,
                workspace,
                debug_log=str((root / "logs" / "e2e-run-2.jsonl").resolve()),
            )
            video_plan = cast(
                dict[str, object],
                cast(dict[str, object], stage1_result["details"])["video_plan"],
            )
            common_assets = root / "common_assets"
            common_assets.mkdir(parents=True, exist_ok=True)
            video_plan["asset_plan"] = {
                "asset_root": str(common_assets.resolve()),
                "common_asset_folder": str(common_assets.resolve()),
            }
            stage2_jobs, _ = build_stage2_jobs(video_plan)
            final_video = root / "final_video.mp4"
            final_video.write_bytes(b"mp4")
            _ = sync_final_video_result(
                config=config,
                excel_path=excel_path,
                sheet_name="Sheet1",
                row_index=0,
                worker_result={
                    "completion": {
                        "state": "completed",
                        "final_output": True,
                        "final_artifact_path": str(final_video.resolve()),
                    },
                    "details": {"reason_code": "ok"},
                },
                run_id=run_id,
                artifact_root=root,
                debug_log=str((root / "logs" / "e2e-run-2.jsonl").resolve()),
            )
            latest_result = json.loads(
                config.result_router_file.read_text(encoding="utf-8")
            )

        self.assertEqual(topic_spec["run_id"], run_id)
        self.assertEqual(video_plan["run_id"], run_id)
        first_stage2 = cast(dict[str, object], stage2_jobs[0]["job"])
        self.assertEqual(
            cast(dict[str, object], first_stage2["payload"])["run_id"], run_id
        )
        self.assertEqual(latest_result["metadata"]["run_id"], run_id)

    def test_stage2_builder_adds_default_kenburns_motion_settings(self) -> None:
        video_plan: dict[str, object] = {
            "contract": "video_plan",
            "run_id": "stage2-run-kenburns-motion",
            "row_ref": "Sheet1!row1",
            "reason_code": "ok",
            "asset_plan": {
                "asset_root": str(Path("D:/YOUTUBEAUTO").resolve()),
                "common_asset_folder": str(Path("D:/YOUTUBEAUTO").resolve()),
            },
            "scene_plan": [
                {"scene_index": 1, "prompt": "scene one"},
                {"scene_index": 2, "prompt": "scene two"},
            ],
        }

        stage2_jobs, _ = build_stage2_jobs(video_plan)

        kenburns_jobs = [
            cast(dict[str, object], item)
            for item in stage2_jobs
            if cast(dict[str, object], item["job"])["worker"] == "kenburns"
        ]
        self.assertEqual(len(kenburns_jobs), 1)
        kenburns_payload = cast(dict[str, object], kenburns_jobs[0]["job"])["payload"]
        scenes = cast(
            list[object],
            cast(
                dict[str, object],
                cast(dict[str, object], kenburns_payload)["scene_bundle_map"],
            )["scenes"],
        )
        first_scene = cast(dict[str, object], scenes[0])
        second_scene = cast(dict[str, object], scenes[1])
        self.assertEqual(first_scene["pan_direction"], "left")
        self.assertEqual(second_scene["pan_direction"], "right")
        self.assertEqual(first_scene["zoom_mode"], "in")
        self.assertEqual(second_scene["zoom_mode"], "out")
        self.assertEqual(first_scene["pan_pct"], 0.05)
        self.assertEqual(first_scene["zoom_pct"], 0.4)

    def test_probe_result_uses_canonical_path_and_required_schema(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir) / "probe"
            output = _write_probe_result(
                root,
                {
                    "run_id": "probe-run-1",
                    "mode": "selftest",
                    "status": "ok",
                    "code": "OK",
                    "exit_code": 0,
                    "debug_log": "x",
                    "result_path": "y",
                    "ts": 1.0,
                },
            )
            payload = json.loads(output.read_text(encoding="utf-8"))

        self.assertEqual(output.name, "probe_result.json")
        self.assertEqual(payload["run_id"], "probe-run-1")
        self.assertEqual(payload["mode"], "selftest")
        self.assertIn("schema_version", payload)

    def test_probe_result_serializes_path_payload_safely(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir) / "probe"
            output = _write_probe_result(
                root,
                {
                    "run_id": "probe-run-2",
                    "mode": "selftest",
                    "status": "ok",
                    "code": "OK",
                    "exit_code": 0,
                    "debug_log": Path("logs/run.jsonl"),
                    "result_path": Path("evidence/result.json"),
                    "ts": 1.0,
                },
            )
            payload = json.loads(output.read_text(encoding="utf-8"))

        self.assertTrue(str(payload["debug_log"]).endswith("run.jsonl"))
        self.assertTrue(str(payload["result_path"]).endswith("result.json"))

    def test_failure_summary_keeps_one_reason_three_evidence_refs(self) -> None:
        with tempfile.TemporaryDirectory(dir="D:\\YOUTUBEAUTO") as tmp_dir:
            root = Path(tmp_dir)
            config = RuntimeConfig.from_root(root)
            output = sync_failure_result(
                config=config,
                run_id="fail-run-1",
                reason_code="one_reason",
                summary="boom",
                evidence_refs=["a", "b", "c", "d"],
                debug_log=str((root / "logs" / "fail-run-1.jsonl").resolve()),
                artifact_root=root,
            )
            failure_payload = json.loads(
                config.failure_summary_file.read_text(encoding="utf-8")
            )
            latest_payload = json.loads(
                config.result_router_file.read_text(encoding="utf-8")
            )
            gui_payload = json.loads(config.gui_status_file.read_text(encoding="utf-8"))

        self.assertTrue(output)
        self.assertEqual(failure_payload["reason_code"], "one_reason")
        self.assertEqual(len(failure_payload["evidence_refs"]), 3)
        self.assertEqual(
            latest_payload["metadata"]["failure_summary_path"],
            str(config.failure_summary_file),
        )
        self.assertEqual(gui_payload["run_id"], "fail-run-1")


if __name__ == "__main__":
    _ = unittest.main()
